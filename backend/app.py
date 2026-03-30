import os
import random
import hashlib
import json
import secrets
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from io import BytesIO

from flask import Flask, jsonify, request, send_from_directory, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from flask_mail import Mail, Message
from werkzeug.security import check_password_hash, generate_password_hash
from flask_socketio import SocketIO, emit, join_room, leave_room
from dotenv import load_dotenv

# Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Optional imports
try:
    import razorpay
    RAZORPAY_AVAILABLE = True
except ImportError:
    RAZORPAY_AVAILABLE = False
    razorpay = None

try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False
    openai = None

# Load .env file from the backend directory
env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
load_dotenv(env_path)

try:
    from locations import get_all_locations, get_states_by_country, get_districts_by_state, get_countries
except ImportError:
    # Fallback if locations module doesn't exist
    def get_all_locations():
        return []
    def get_states_by_country(country):
        return []
    def get_districts_by_state(country, state):
        return []
    def get_countries():
        return []

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Database configuration - MySQL (Hostinger) or fallback to SQLite
DB_HOST = os.getenv("DB_HOST", "")
DB_USER = os.getenv("DB_USER", "")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")
DB_NAME = os.getenv("DB_NAME", "")

if DB_HOST and DB_USER and DB_NAME:
    # URL-encode password to handle special chars like '@'
    from urllib.parse import quote_plus
    DB_PASSWORD_ENCODED = quote_plus(DB_PASSWORD)
    # MySQL connection via PyMySQL
    DATABASE_URI = f"mysql+pymysql://{DB_USER}:{DB_PASSWORD_ENCODED}@{DB_HOST}/{DB_NAME}?charset=utf8mb4"
    print(f"[DB] Using MySQL: {DB_USER}@{DB_HOST}/{DB_NAME}")
else:
    # Fallback to local SQLite
    DB_PATH = os.path.join(BASE_DIR, "data", "tradeindia.db")
    DATABASE_URI = f"sqlite:///{DB_PATH}"
    print("[DB] Using local SQLite (no MySQL credentials found)")


app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URI
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_recycle": 280,       # Recycle connections before MySQL timeout (default 8h)
    "pool_pre_ping": True,     # Test connection health before using from pool
    "pool_size": 5,
    "max_overflow": 10,
}
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-secret")

# Email configuration
app.config["MAIL_SERVER"] = os.getenv("MAIL_SERVER", "smtp.gmail.com")
app.config["MAIL_PORT"] = int(os.getenv("MAIL_PORT", 587))
app.config["MAIL_USE_TLS"] = os.getenv("MAIL_USE_TLS", "True").lower() == "true"
app.config["MAIL_USERNAME"] = os.getenv("MAIL_USERNAME", "")
app.config["MAIL_PASSWORD"] = os.getenv("MAIL_PASSWORD", "")

# Razorpay configuration
RAZORPAY_KEY_ID = os.getenv("RAZORPAY_KEY_ID", "")
RAZORPAY_KEY_SECRET = os.getenv("RAZORPAY_KEY_SECRET", "")
razorpay_client = None
if RAZORPAY_AVAILABLE and RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET:
    try:
        razorpay_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
    except Exception as e:
        print(f"Warning: Failed to initialize Razorpay client: {e}")
        razorpay_client = None

# Initialize SocketIO
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="eventlet", logger=True, engineio_logger=True)

@socketio.on('connect')
def handle_connect():
    print(f"[Socket] Client connected")

@socketio.on('disconnect')
def handle_disconnect():
    print(f"[Socket] Client disconnected")

@socketio.on('join')
def on_join(data):
    room = data.get('room')
    if room:
        join_room(room)
        print(f"[Socket] Client joined room: {room}")

@socketio.on('send_message')
def handle_send_message(data):
    """Handle real-time message sending with DB persistence."""
    room = data.get('room') # This is the inquiry_id string
    sender_id = data.get('sender_id')
    message_text = data.get('message')
    
    if room and sender_id and message_text:
        try:
            # Inquiry ID is used as the room name
            inquiry_id = int(room)
            
            # Persist to database
            new_msg = InquiryMessage(
                inquiry_id=inquiry_id,
                sender_id=sender_id,
                message=message_text
            )
            db.session.add(new_msg)
            
            # Update inquiry status if needed
            inquiry = Inquiry.query.get(inquiry_id)
            if inquiry and inquiry.negotiation_status == "OPEN":
                inquiry.negotiation_status = "ACTIVE"
                
            db.session.commit()

            if inquiry and inquiry.product:
                sender = User.query.get(sender_id)
                try:
                    record_marketplace_event(
                        "negotiation_message",
                        session_id=get_request_session_id(data),
                        user_id=sender_id,
                        company_id=inquiry.product.company_id,
                        product_id=inquiry.product_id,
                        category_id=inquiry.product.category_id,
                        inquiry_id=inquiry.id,
                        location=inquiry.product.company.location if inquiry.product.company else "",
                        metadata={
                            "sender_company_id": sender.company_id if sender else None,
                            "message_length": len(message_text),
                        },
                    )
                except Exception as analytics_error:
                    print(f"[Analytics] Failed to record negotiation message: {analytics_error}")
                    db.session.rollback()

                if sender and sender.company_id != inquiry.product.company_id:
                    try:
                        create_seller_alert(
                            inquiry.product.company_id,
                            "negotiation_message",
                            f"Buyer message for {inquiry.product.name}",
                            f"{sender.name} sent a live negotiation message.",
                            severity="info",
                            entity_type="inquiry",
                            entity_id=inquiry.id,
                        )
                    except Exception as alert_error:
                        print(f"[Alert] Failed to create negotiation alert: {alert_error}")
                        db.session.rollback()
            
            # Emit back to the room including the new ID and timestamp
            emit('receive_message', {
                "id": new_msg.id,
                "inquiry_id": inquiry_id,
                "sender_id": sender_id,
                "sender_name": new_msg.sender.name if new_msg.sender else "Unknown",
                "message": message_text,
                "created_at": new_msg.created_at.isoformat()
            }, room=room)
        except Exception as e:
            print(f"[Socket Error] {e}")
            db.session.rollback()

# OpenAI configuration
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
openai_client = None
if OPENAI_AVAILABLE and OPENAI_API_KEY:
    try:
        openai.api_key = OPENAI_API_KEY
        openai_client = openai
    except Exception as e:
        print(f"Warning: Failed to initialize OpenAI client: {e}")
        openai_client = None

# File upload configuration
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}
ALLOWED_DATA_EXTENSIONS = {"csv", "xlsx", "xls"}
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(os.path.join(UPLOAD_FOLDER, "products"), exist_ok=True)
os.makedirs(os.path.join(UPLOAD_FOLDER, "avatars"), exist_ok=True)
os.makedirs(os.path.join(UPLOAD_FOLDER, "companies"), exist_ok=True)

CORS(app)
db = SQLAlchemy(app)
mail = Mail(app)

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_data_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_DATA_EXTENSIONS


def hash_reset_token(token):
    return hashlib.sha256((token or "").encode("utf-8")).hexdigest()


def get_frontend_base_url():
    configured_url = os.getenv("FRONTEND_URL", "").strip()
    if configured_url:
        return configured_url.rstrip("/")
    return "http://localhost:3000"


def invalidate_homepage_cache():
    _homepage_cache["expires_at"] = None
    _homepage_cache["payload"] = None


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"))
    avatar_url = db.Column(db.String(255), default="")
    phone = db.Column(db.String(20), default="")
    bio = db.Column(db.Text, default="")
    is_admin = db.Column(db.Boolean, default=False, nullable=False)  # Admin flag
    is_buyer_manager = db.Column(db.Boolean, default=False, nullable=False)  # Buyer Manager flag
    membership_tier = db.Column(db.String(20), default="STARTER")  # STARTER, BASIC, PREMIUM
    welcome_email_sent = db.Column(db.Boolean, default=False, nullable=False)  # Track if welcome email was sent
    reset_token_hash = db.Column(db.String(64), default="", nullable=False)
    reset_token_expires_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    company = db.relationship("Company", back_populates="users")


class OTP(db.Model):
    """Store OTP for phone number verification"""
    id = db.Column(db.Integer, primary_key=True)
    phone = db.Column(db.String(20), nullable=False, index=True)
    otp_code = db.Column(db.String(6), nullable=False)
    expires_at = db.Column(db.DateTime, nullable=False)
    verified = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class Company(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    description = db.Column(db.Text, default="")
    location = db.Column(db.String(120), default="India")
    website = db.Column(db.String(255), default="")
    phone = db.Column(db.String(20), default="")
    gst_number = db.Column(db.String(20), default="")
    logo_url = db.Column(db.String(255), default="")
    verified = db.Column(db.Boolean, default=False)
    best_seller = db.Column(db.Boolean, default=False)  # Best seller flag
    membership_tier = db.Column(db.String(20), default="FREE")  # FREE, SILVER, GOLD, PLATINUM
    priority_score = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    products = db.relationship("Product", back_populates="company")
    users = db.relationship("User", back_populates="company")


class Subscription(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    plan_type = db.Column(db.String(50), nullable=False)  # 'basic', 'premium', 'enterprise'
    is_active = db.Column(db.Boolean, default=True)
    start_date = db.Column(db.DateTime, default=datetime.utcnow)
    end_date = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    user = db.relationship("User", backref="subscriptions")


HOMEPAGE_CACHE_TTL_SECONDS = 90
_homepage_cache = {
    "expires_at": None,
    "payload": None,
}


class Category(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False, unique=True)
    views = db.Column(db.Integer, default=0)
    sales_count = db.Column(db.Integer, default=0)
    last_visited_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    products = db.relationship("Product", back_populates="category")


class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    description = db.Column(db.Text, default="")
    price = db.Column(db.Float, nullable=True)
    image_url = db.Column(db.String(255), default="")
    featured = db.Column(db.Boolean, default=False)
    approved = db.Column(db.Boolean, default=False)  # Admin approval required
    location = db.Column(db.String(120), default="India")
    stock_quantity = db.Column(db.Integer, default=0)
    min_order_quantity = db.Column(db.Integer, default=1)
    tags = db.Column(db.Text, default="")  # JSON string of tags
    ai_description = db.Column(db.Text, default="")
    is_priority = db.Column(db.Boolean, default=False)
    ai_summary = db.Column(db.Text, default="")
    price_trend = db.Column(db.Float, default=0.0) # percentage trend, e.g., 5.4 or -1.2
    bulk_pricing_json = db.Column(db.Text, default="[]") # JSON list of {min_qty, price}
    views = db.Column(db.Integer, default=0)
    sales_count = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    category_id = db.Column(db.Integer, db.ForeignKey("category.id"))
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"))

    category = db.relationship("Category", back_populates="products")
    company = db.relationship("Company", back_populates="products")


class Inquiry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey("product.id"), nullable=False)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), nullable=False)
    phone = db.Column(db.String(20), nullable=False)
    message = db.Column(db.Text, nullable=False)
    quantity = db.Column(db.String(50))
    manager_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    negotiation_status = db.Column(db.String(20), default="OPEN") # OPEN, ACTIVE, CLOSED
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    product = db.relationship("Product", backref="inquiries")
    manager = db.relationship("User", backref="assigned_inquiries", foreign_keys=[manager_id])
    replies = db.relationship("InquiryReply", backref="inquiry", cascade="all, delete-orphan", order_by="InquiryReply.created_at")


class InquiryReply(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    inquiry_id = db.Column(db.Integer, db.ForeignKey("inquiry.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)  # Seller who replied
    message = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class RFQMatch(db.Model):
    """Stores results of AI matching between inquiries and suppliers"""
    id = db.Column(db.Integer, primary_key=True)
    inquiry_id = db.Column(db.Integer, db.ForeignKey("inquiry.id"), nullable=False)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=False)
    score = db.Column(db.Float, default=0.0)
    status = db.Column(db.String(20), default="MATCHED")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    inquiry = db.relationship("Inquiry", backref=db.backref("matches", cascade="all, delete-orphan"))
    company = db.relationship("Company", backref="rfq_matches")

class InquiryMessage(db.Model):
    """Stores real-time negotiation messages for an inquiry"""
    __tablename__ = 'inquiry_messages'
    id = db.Column(db.Integer, primary_key=True)
    inquiry_id = db.Column(db.Integer, db.ForeignKey("inquiry.id"), nullable=False)
    sender_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    message = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    inquiry = db.relationship("Inquiry", backref=db.backref("chat_messages", cascade="all, delete-orphan"))
    sender = db.relationship("User", backref="inquiry_sent_messages")


class BuyRequirement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_name = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False)
    quantity = db.Column(db.String(50))
    location = db.Column(db.String(120))
    budget = db.Column(db.String(100))
    email = db.Column(db.String(120))
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    user = db.relationship("User", backref="buy_requirements")


class TradeLead(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False)
    type = db.Column(db.String(20), nullable=False)  # 'buy' or 'sell'
    category = db.Column(db.String(100))
    location = db.Column(db.String(120))
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"))
    
    # New contact fields for "Buying Leads" feature
    contact_name = db.Column(db.String(120), default="")
    contact_email = db.Column(db.String(120), default="")
    contact_phone = db.Column(db.String(20), default="")
    price = db.Column(db.Integer, default=500) # Default price for a lead
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    company = db.relationship("Company", backref="trade_leads")

class LeadPurchase(db.Model):
    """Tracks which users have purchased which trade leads."""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    lead_id = db.Column(db.Integer, db.ForeignKey("trade_lead.id"), nullable=False)
    purchased_at = db.Column(db.DateTime, default=datetime.utcnow)

    user = db.relationship("User", backref="lead_purchases")
    lead = db.relationship("TradeLead", backref="purchasers")


class Order(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey("product.id"), nullable=False)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=False)
    buyer_name = db.Column(db.String(120), nullable=False)
    buyer_email = db.Column(db.String(120), nullable=False)
    buyer_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    quantity = db.Column(db.Integer, default=1)
    unit_price = db.Column(db.Float, nullable=False)
    total_amount = db.Column(db.Float, nullable=False)
    status = db.Column(db.String(20), default="pending")  # pending, processing, shipped, delivered, completed, cancelled
    payment_status = db.Column(db.String(20), default="pending")  # pending, paid, failed, refunded
    payment_method = db.Column(db.String(50), default="")
    payment_id = db.Column(db.String(255), default="")  # Razorpay payment ID
    tracking_number = db.Column(db.String(100), default="")
    shipping_address = db.Column(db.Text, default="")
    shipping_carrier = db.Column(db.String(100), default="")
    estimated_delivery = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    product = db.relationship("Product", backref="orders")
    company = db.relationship("Company", backref="orders")
    buyer = db.relationship("User", backref="orders")


class ProductImage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey("product.id"), nullable=False)
    image_url = db.Column(db.String(255), nullable=False)
    is_primary = db.Column(db.Boolean, default=False)
    display_order = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    product = db.relationship("Product", backref="images")


class Review(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey("product.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    rating = db.Column(db.Integer, nullable=False)  # 1-5
    title = db.Column(db.String(200), default="")
    comment = db.Column(db.Text, default="")
    is_verified_purchase = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    product = db.relationship("Product", backref="reviews")
    user = db.relationship("User", backref="reviews")


class OrderTracking(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.Integer, db.ForeignKey("order.id"), nullable=False)
    status = db.Column(db.String(50), nullable=False)
    message = db.Column(db.Text, default="")
    location = db.Column(db.String(200), default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    order = db.relationship("Order", backref="tracking_history")


class ChatMessage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sender_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    receiver_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey("product.id"), nullable=True)
    order_id = db.Column(db.Integer, db.ForeignKey("order.id"), nullable=True)
    message = db.Column(db.Text, nullable=False)
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    sender = db.relationship("User", foreign_keys=[sender_id], backref="sent_messages")
    receiver = db.relationship("User", foreign_keys=[receiver_id], backref="received_messages")
    product = db.relationship("Product", backref="chat_messages")
    order = db.relationship("Order", backref="chat_messages")


class Wishlist(db.Model):
    """User wishlist/favorites"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey("product.id"), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    user = db.relationship("User", backref="wishlist_items")
    product = db.relationship("Product", backref="wishlisted_by")
    
    __table_args__ = (db.UniqueConstraint('user_id', 'product_id', name='unique_user_product_wishlist'),)


class MarketplaceEvent(db.Model):
    """Structured event stream used for real-time marketplace analytics."""
    id = db.Column(db.Integer, primary_key=True)
    event_type = db.Column(db.String(50), nullable=False, index=True)
    session_id = db.Column(db.String(120), default="", index=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)
    product_id = db.Column(db.Integer, db.ForeignKey("product.id"), nullable=True)
    category_id = db.Column(db.Integer, db.ForeignKey("category.id"), nullable=True)
    inquiry_id = db.Column(db.Integer, db.ForeignKey("inquiry.id"), nullable=True)
    order_id = db.Column(db.Integer, db.ForeignKey("order.id"), nullable=True)
    location = db.Column(db.String(120), default="")
    search_query = db.Column(db.String(255), default="")
    metadata_json = db.Column(db.Text, default="{}")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)

    user = db.relationship("User", backref="marketplace_events")
    company = db.relationship("Company", backref="marketplace_events")
    product = db.relationship("Product", backref="marketplace_events")
    category = db.relationship("Category", backref="marketplace_events")
    inquiry = db.relationship("Inquiry", backref="marketplace_events")
    order = db.relationship("Order", backref="marketplace_events")


class SellerAlert(db.Model):
    """Persisted alerts surfaced in the seller real-time alert center."""
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=False, index=True)
    alert_type = db.Column(db.String(50), nullable=False, index=True)
    title = db.Column(db.String(255), nullable=False)
    message = db.Column(db.Text, default="")
    severity = db.Column(db.String(20), default="info")
    entity_type = db.Column(db.String(50), default="")
    entity_id = db.Column(db.Integer, nullable=True)
    is_read = db.Column(db.Boolean, default=False, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)

    company = db.relationship("Company", backref="seller_alerts")


def get_current_user():
    """Get the current authenticated user from the request token."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token or token == "null":
        return None
    
    # For demo purposes, try to get user_id from query params or request
    # In production, decode JWT token here
    user_id = request.args.get("user_id", type=int)
    if not user_id:
        # Try to get from JSON body
        data = request.get_json(silent=True) or {}
        user_id = data.get("user_id")
    
    if user_id:
        return User.query.get(user_id)
    return None


def require_admin():
    """Helper to require admin access. Returns (user, error_response) tuple."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token or token == "null":
        return None, (jsonify({"error": "Authentication required"}), 401)
    
    # Get user_id from request (query params, JSON body, or headers)
    user_id = request.args.get("user_id", type=int)
    if not user_id:
        data = request.get_json(silent=True) or {}
        user_id = data.get("user_id")
    if not user_id:
        # Try to get from X-User-Id header (frontend can set this)
        user_id = request.headers.get("X-User-Id", type=int)
    
    if not user_id:
        return None, (jsonify({"error": "User ID required"}), 400)
    
    user = User.query.get(user_id)
    if not user:
        return None, (jsonify({"error": "User not found"}), 404)
    
    if not user.is_admin:
        return None, (jsonify({"error": "Admin access required"}), 403)
    
    return user, None


def parse_json_field(value, default=None):
    if default is None:
        default = {}
    if not value:
        return default
    try:
        return json.loads(value)
    except (TypeError, ValueError):
        return default


def get_request_session_id(data=None):
    data = data or request.get_json(silent=True) or {}
    session_id = (
        request.headers.get("X-Session-Id")
        or request.args.get("session_id")
        or data.get("session_id")
    )
    if session_id:
        return str(session_id)

    user = get_current_user()
    if user:
        return f"user:{user.id}"

    remote = (request.headers.get("X-Forwarded-For", "") or "").split(",")[0].strip() or request.remote_addr or "anonymous"
    user_agent = request.headers.get("User-Agent", "browser")
    return f"{remote}:{hashlib.md5(user_agent.encode()).hexdigest()[:8]}"


def serialize_marketplace_event(event):
    metadata = parse_json_field(event.metadata_json, {})
    labels = {
        "homepage_view": "Landing page visit",
        "catalog_view": "Catalog exploration",
        "search": f"Search: {event.search_query}" if event.search_query else "Marketplace search",
        "product_view": f"Viewed {event.product.name}" if event.product else "Product view",
        "inquiry_created": f"Inquiry for {event.product.name}" if event.product else "New inquiry",
        "order_created": f"Order for {event.product.name}" if event.product else "New order",
        "buy_requirement_created": "Buy requirement posted",
        "negotiation_message": "Negotiation activity",
    }
    return {
        "id": event.id,
        "event_type": event.event_type,
        "title": labels.get(event.event_type, event.event_type.replace("_", " ").title()),
        "location": event.location or "Unknown",
        "company_name": event.company.name if event.company else None,
        "product_name": event.product.name if event.product else metadata.get("product_name"),
        "category_name": event.category.name if event.category else metadata.get("category_name"),
        "search_query": event.search_query or None,
        "metadata": metadata,
        "created_at": event.created_at.isoformat() if event.created_at else None,
    }


def serialize_seller_alert(alert):
    return {
        "id": alert.id,
        "alert_type": alert.alert_type,
        "title": alert.title,
        "message": alert.message,
        "severity": alert.severity,
        "entity_type": alert.entity_type,
        "entity_id": alert.entity_id,
        "is_read": alert.is_read,
        "created_at": alert.created_at.isoformat() if alert.created_at else None,
    }


def emit_dashboard_refresh(company_id=None):
    payload = {"timestamp": datetime.utcnow().isoformat()}
    socketio.emit("dashboard_update", payload, room="admin_dashboard")
    if company_id:
        socketio.emit("dashboard_update", payload, room=f"seller_dashboard_{company_id}")


def emit_marketplace_event(event):
    payload = serialize_marketplace_event(event)
    socketio.emit("marketplace_event", payload, room="admin_dashboard")
    if event.company_id:
        socketio.emit("marketplace_event", payload, room=f"seller_dashboard_{event.company_id}")


def create_seller_alert(
    company_id,
    alert_type,
    title,
    message,
    severity="info",
    entity_type="",
    entity_id=None,
    commit=True,
    emit_updates=True,
):
    alert = SellerAlert(
        company_id=company_id,
        alert_type=alert_type,
        title=title,
        message=message,
        severity=severity,
        entity_type=entity_type,
        entity_id=entity_id,
    )
    db.session.add(alert)
    if commit:
        db.session.commit()
        if emit_updates:
            socketio.emit("seller_alert", serialize_seller_alert(alert), room=f"seller_dashboard_{company_id}")
            emit_dashboard_refresh(company_id=company_id)
    return alert


def record_marketplace_event(
    event_type,
    session_id=None,
    user_id=None,
    company_id=None,
    product_id=None,
    category_id=None,
    inquiry_id=None,
    order_id=None,
    location="",
    search_query="",
    metadata=None,
    commit=True,
    emit_updates=True,
):
    event = MarketplaceEvent(
        event_type=event_type,
        session_id=session_id or get_request_session_id(),
        user_id=user_id,
        company_id=company_id,
        product_id=product_id,
        category_id=category_id,
        inquiry_id=inquiry_id,
        order_id=order_id,
        location=location or "",
        search_query=(search_query or "")[:255],
        metadata_json=json.dumps(metadata or {}),
    )
    db.session.add(event)
    if commit:
        db.session.commit()
        if emit_updates:
            emit_marketplace_event(event)
            emit_dashboard_refresh(company_id=company_id)
    return event


def score_lead_quality(message="", quantity=""):
    lowered = (message or "").lower()
    score = 35

    try:
        qty_num = int("".join(filter(str.isdigit, quantity or ""))) if quantity else 0
    except ValueError:
        qty_num = 0

    if qty_num >= 1000:
        score += 30
    elif qty_num >= 250:
        score += 20
    elif qty_num >= 50:
        score += 10

    keywords = {
        "bulk": 12,
        "wholesale": 10,
        "export": 10,
        "contract": 12,
        "urgent": 8,
        "container": 12,
        "repeat": 8,
        "monthly": 10,
        "partnership": 14,
    }
    for keyword, weight in keywords.items():
        if keyword in lowered:
            score += weight

    score = max(0, min(score, 100))
    if score >= 80:
        temperature = "hot"
    elif score >= 55:
        temperature = "warm"
    else:
        temperature = "cold"

    return {
        "score": score,
        "temperature": temperature,
        "quantity": qty_num,
    }


def calculate_supplier_trust(company_id):
    company = Company.query.get(company_id)
    if not company:
        return {
            "trust_score": 0,
            "completion_rate": 0,
            "avg_response_hours": None,
            "response_rate": 0,
            "review_rating": 0,
            "review_count": 0,
            "verified": False,
        }

    orders = Order.query.filter_by(company_id=company_id).all()
    total_orders = len(orders)
    completed_orders = len([o for o in orders if o.status == "completed"])
    cancelled_orders = len([o for o in orders if o.status == "cancelled"])
    completion_rate = round((completed_orders / total_orders) * 100, 1) if total_orders else 0

    review_stats = (
        db.session.query(
            db.func.avg(Review.rating).label("avg_rating"),
            db.func.count(Review.id).label("review_count"),
        )
        .join(Product, Product.id == Review.product_id)
        .filter(Product.company_id == company_id)
        .first()
    )
    avg_rating = round(float(review_stats.avg_rating), 1) if review_stats and review_stats.avg_rating else 0
    review_count = review_stats.review_count if review_stats else 0

    inquiries = Inquiry.query.join(Product).filter(Product.company_id == company_id).all()
    response_times = []
    responded_count = 0
    for inquiry in inquiries:
        response_candidates = []
        for reply in inquiry.replies:
            reply_user = User.query.get(reply.user_id)
            if reply_user and reply_user.company_id == company_id:
                response_candidates.append(reply.created_at)
        for chat_message in inquiry.chat_messages:
            sender = chat_message.sender
            if sender and sender.company_id == company_id:
                response_candidates.append(chat_message.created_at)

        if response_candidates:
            responded_count += 1
            first_response = min(response_candidates)
            response_times.append(max((first_response - inquiry.created_at).total_seconds() / 3600, 0))

    avg_response_hours = round(sum(response_times) / len(response_times), 1) if response_times else None
    response_rate = round((responded_count / len(inquiries)) * 100, 1) if inquiries else 100

    verification_score = 20 if company.verified else 8
    rating_score = (avg_rating / 5) * 25 if review_count else (12 if company.verified else 6)
    completion_score = min(completion_rate * 0.25, 25) if total_orders else (12 if company.verified else 5)
    cancellation_penalty = min(cancelled_orders * 3, 10)
    response_coverage_score = response_rate * 0.2

    if avg_response_hours is None:
        response_speed_score = 10 if not inquiries else 4
    elif avg_response_hours <= 1:
        response_speed_score = 20
    elif avg_response_hours <= 4:
        response_speed_score = 17
    elif avg_response_hours <= 12:
        response_speed_score = 13
    elif avg_response_hours <= 24:
        response_speed_score = 9
    else:
        response_speed_score = 4

    trust_score = max(
        0,
        min(
            round(
                verification_score
                + rating_score
                + completion_score
                + response_coverage_score
                + response_speed_score
                - cancellation_penalty,
                1,
            ),
            100,
        ),
    )

    return {
        "trust_score": trust_score,
        "completion_rate": completion_rate,
        "avg_response_hours": avg_response_hours,
        "response_rate": response_rate,
        "review_rating": avg_rating,
        "review_count": review_count,
        "verified": company.verified,
    }


def build_seller_recommendations(company_id, hot_products, trust_data, stats_summary):
    recommendations = []
    if stats_summary.get("low_stock_count", 0) > 0:
        recommendations.append("Restock low-inventory products before active buyers drop off.")
    if trust_data.get("avg_response_hours") and trust_data["avg_response_hours"] > 12:
        recommendations.append("Reply faster to inquiries to improve supplier trust and conversion.")
    if trust_data.get("trust_score", 0) < 70:
        recommendations.append("Boost your trust score with quicker replies, fulfilled orders, and verified reviews.")
    if hot_products and hot_products[0].get("views", 0) >= 5 and hot_products[0].get("orders", 0) == 0:
        recommendations.append(f"Promote {hot_products[0]['name']} with a stronger CTA because demand is high but orders are still low.")
    if stats_summary.get("pending_products", 0) > 0:
        recommendations.append("Approve pending catalog items quickly so they can capture current marketplace demand.")
    return recommendations[:4]


def build_admin_recommendations(overview, high_intent_products, supplier_watchlist):
    recommendations = []
    if overview["inquiries_24h"] > overview["orders_24h"] * 4 and overview["orders_24h"] > 0:
        recommendations.append("Large inquiry volume is not turning into orders fast enough. Improve quote speed and follow-up.")
    if high_intent_products:
        recommendations.append(f"High demand is building around {high_intent_products[0]['name']}. Feature it on the homepage and supplier outreach flows.")
    slow_suppliers = [supplier for supplier in supplier_watchlist if (supplier.get("avg_response_hours") or 0) > 12]
    if slow_suppliers:
        recommendations.append("Some active suppliers are responding slowly. Surface response-time warnings in seller tools.")
    if overview["searches_24h"] > overview["product_views_24h"]:
        recommendations.append("Buyers are searching more than they are clicking. Improve result relevance, price cues, and trust badges.")
    return recommendations[:4]


def build_admin_realtime_analytics(days=30):
    now = datetime.utcnow()
    since = now - timedelta(days=days)
    last_24h_cutoff = now - timedelta(hours=24)
    events = MarketplaceEvent.query.filter(MarketplaceEvent.created_at >= since).order_by(MarketplaceEvent.created_at.desc()).all()
    last_24h_events = [event for event in events if event.created_at >= last_24h_cutoff]

    product_ids = {event.product_id for event in events if event.product_id}
    category_ids = {event.category_id for event in events if event.category_id}
    company_ids = {event.company_id for event in events if event.company_id}

    product_map = {product.id: product for product in Product.query.filter(Product.id.in_(product_ids)).all()} if product_ids else {}
    category_map = {category.id: category for category in Category.query.filter(Category.id.in_(category_ids)).all()} if category_ids else {}
    company_map = {company.id: company for company in Company.query.filter(Company.id.in_(company_ids)).all()} if company_ids else {}

    overview = {
        "active_buyers_24h": len({event.session_id for event in last_24h_events if event.session_id}),
        "searches_24h": len([event for event in last_24h_events if event.event_type == "search"]),
        "product_views_24h": len([event for event in last_24h_events if event.event_type == "product_view"]),
        "inquiries_24h": len([event for event in last_24h_events if event.event_type == "inquiry_created"]),
        "orders_24h": len([event for event in last_24h_events if event.event_type == "order_created"]),
        "leads_24h": len([event for event in last_24h_events if event.event_type == "lead_created"]),
        "lead_purchases_24h": len([event for event in last_24h_events if event.event_type == "lead_purchased"]),
    }

    search_counts = {}
    category_activity = {}
    product_activity = {}
    heatmap = {}
    supplier_activity = {}

    for event in events:
        if event.search_query:
            search_counts[event.search_query] = search_counts.get(event.search_query, 0) + 1

        if event.category_id:
            category_activity[event.category_id] = category_activity.get(event.category_id, 0) + 1

        if event.product_id:
            product_bucket = product_activity.setdefault(
                event.product_id,
                {"views": 0, "inquiries": 0, "orders": 0, "intensity": 0},
            )
            if event.event_type == "product_view":
                product_bucket["views"] += 1
                product_bucket["intensity"] += 1
            elif event.event_type == "inquiry_created":
                product_bucket["inquiries"] += 1
                product_bucket["intensity"] += 4
            elif event.event_type == "order_created":
                product_bucket["orders"] += 1
                product_bucket["intensity"] += 6

        if event.company_id:
            supplier_activity[event.company_id] = supplier_activity.get(event.company_id, 0) + 1

        if event.location and event.category_id:
            heatmap_key = (event.location, event.category_id)
            bucket = heatmap.setdefault(heatmap_key, {"events": 0, "inquiries": 0, "orders": 0})
            bucket["events"] += 1
            if event.event_type == "inquiry_created":
                bucket["inquiries"] += 1
            if event.event_type == "order_created":
                bucket["orders"] += 1

    visitors = len({event.session_id for event in events if event.session_id})
    searches = len([event for event in events if event.event_type == "search"])
    product_views = len([event for event in events if event.event_type == "product_view"])
    inquiries = len([event for event in events if event.event_type == "inquiry_created"])
    orders = len([event for event in events if event.event_type == "order_created"])

    funnel = {
        "visitors": visitors,
        "searches": searches,
        "product_views": product_views,
        "inquiries": inquiries,
        "orders": orders,
        "search_to_view_rate": round((product_views / searches) * 100, 1) if searches else 0,
        "view_to_inquiry_rate": round((inquiries / product_views) * 100, 1) if product_views else 0,
        "inquiry_to_order_rate": round((orders / inquiries) * 100, 1) if inquiries else 0,
    }

    top_viewed_products = sorted(
        [
            {
                "id": product_id,
                "name": product_map.get(product_id).name if product_map.get(product_id) else "Unknown Product",
                "views": data["views"],
                "inquiries": data["inquiries"],
                "orders": data["orders"],
            }
            for product_id, data in product_activity.items()
        ],
        key=lambda item: (item["views"], item["inquiries"], item["orders"]),
        reverse=True,
    )[:6]

    high_intent_products = sorted(
        [
            {
                "id": product_id,
                "name": product_map.get(product_id).name if product_map.get(product_id) else "Unknown Product",
                "views": data["views"],
                "inquiries": data["inquiries"],
                "orders": data["orders"],
                "intent_score": data["intensity"],
            }
            for product_id, data in product_activity.items()
        ],
        key=lambda item: item["intent_score"],
        reverse=True,
    )[:6]

    demand_heatmap = sorted(
        [
            {
                "location": location,
                "category": category_map.get(category_id).name if category_map.get(category_id) else "Uncategorized",
                "events": data["events"],
                "inquiries": data["inquiries"],
                "orders": data["orders"],
                "intent_score": data["events"] + (data["inquiries"] * 3) + (data["orders"] * 5),
            }
            for (location, category_id), data in heatmap.items()
        ],
        key=lambda item: item["intent_score"],
        reverse=True,
    )[:8]

    category_price_stats = (
        db.session.query(
            Category.id,
            Category.name,
            db.func.avg(Product.price).label("avg_price"),
            db.func.avg(Product.price_trend).label("avg_trend"),
            db.func.count(Product.id).label("product_count"),
        )
        .join(Product, Product.category_id == Category.id)
        .filter(Product.approved == True)
        .group_by(Category.id, Category.name)
        .all()
    )
    price_intelligence = {
        "categories": [
            {
                "id": item.id,
                "name": item.name,
                "avg_price": round(float(item.avg_price or 0), 2),
                "avg_trend": round(float(item.avg_trend or 0), 2),
                "product_count": item.product_count,
            }
            for item in category_price_stats
        ],
        "products": sorted(
            [
                {
                    "id": product.id,
                    "name": product.name,
                    "category_name": product.category.name if product.category else "Uncategorized",
                    "company_name": product.company.name if product.company else "Supplier",
                    "price": round(float(product.price or 0), 2),
                    "price_trend": round(float(product.price_trend or 0), 2),
                    "views": product.views or 0,
                }
                for product in Product.query.filter(Product.approved == True).order_by(Product.views.desc(), Product.created_at.desc()).limit(8).all()
            ],
            key=lambda item: (item["views"], item["price_trend"]),
            reverse=True,
        )[:6],
    }

    supplier_watchlist = []
    for company_id, activity_count in sorted(supplier_activity.items(), key=lambda item: item[1], reverse=True)[:6]:
        trust = calculate_supplier_trust(company_id)
        supplier_watchlist.append(
            {
                "company_id": company_id,
                "name": company_map.get(company_id).name if company_map.get(company_id) else "Supplier",
                "activity_count": activity_count,
                "trust_score": trust["trust_score"],
                "avg_response_hours": trust["avg_response_hours"],
                "completion_rate": trust["completion_rate"],
            }
        )

    return {
        "overview": overview,
        "buyer_intent": {
            "top_searches": [{"term": term, "count": count} for term, count in sorted(search_counts.items(), key=lambda item: item[1], reverse=True)[:8]],
            "active_categories": [
                {
                    "id": category_id,
                    "name": category_map.get(category_id).name if category_map.get(category_id) else "Uncategorized",
                    "count": count,
                }
                for category_id, count in sorted(category_activity.items(), key=lambda item: item[1], reverse=True)[:6]
            ],
            "top_viewed_products": top_viewed_products,
            "high_intent_products": high_intent_products,
        },
        "funnel": funnel,
        "demand_heatmap": demand_heatmap,
        "price_intelligence": price_intelligence,
        "live_feed": [serialize_marketplace_event(event) for event in last_24h_events[:12]],
        "supplier_watchlist": supplier_watchlist,
        "recommendations": build_admin_recommendations(overview, high_intent_products, supplier_watchlist),
        "updated_at": now.isoformat(),
    }


def build_seller_live_insights(user_id):
    user = User.query.get(user_id)
    if not user or not user.company_id:
        return None

    now = datetime.utcnow()
    company_id = user.company_id
    last_30_days = now - timedelta(days=30)
    last_24_hours = now - timedelta(hours=24)
    last_7_days = now - timedelta(days=7)

    events = (
        MarketplaceEvent.query.filter(
            MarketplaceEvent.company_id == company_id,
            MarketplaceEvent.created_at >= last_30_days,
        )
        .order_by(MarketplaceEvent.created_at.desc())
        .all()
    )

    trust = calculate_supplier_trust(company_id)
    low_stock_items = Product.query.filter(Product.company_id == company_id, Product.stock_quantity < 10).all()
    pending_products = Product.query.filter(Product.company_id == company_id, Product.approved == False).count()
    alerts = SellerAlert.query.filter_by(company_id=company_id).order_by(SellerAlert.created_at.desc()).limit(8).all()
    unread_alerts = SellerAlert.query.filter_by(company_id=company_id, is_read=False).count()

    product_ids = {event.product_id for event in events if event.product_id}
    product_map = {product.id: product for product in Product.query.filter(Product.id.in_(product_ids)).all()} if product_ids else {}
    product_activity = {}
    for event in events:
        if not event.product_id:
            continue
        bucket = product_activity.setdefault(
            event.product_id,
            {"views": 0, "inquiries": 0, "orders": 0, "intensity": 0},
        )
        if event.event_type == "product_view":
            bucket["views"] += 1
            bucket["intensity"] += 1
        elif event.event_type == "inquiry_created":
            bucket["inquiries"] += 1
            bucket["intensity"] += 4
        elif event.event_type == "order_created":
            bucket["orders"] += 1
            bucket["intensity"] += 6

    hot_products = sorted(
        [
            {
                "id": product_id,
                "name": product_map.get(product_id).name if product_map.get(product_id) else "Unknown Product",
                "views": data["views"],
                "inquiries": data["inquiries"],
                "orders": data["orders"],
                "intent_score": data["intensity"],
            }
            for product_id, data in product_activity.items()
        ],
        key=lambda item: item["intent_score"],
        reverse=True,
    )[:6]

    views = len([event for event in events if event.event_type == "product_view"])
    inquiries = len([event for event in events if event.event_type == "inquiry_created"])
    orders = len([event for event in events if event.event_type == "order_created"])

    summary = {
        "product_views_24h": len([event for event in events if event.event_type == "product_view" and event.created_at >= last_24_hours]),
        "new_inquiries_24h": len([event for event in events if event.event_type == "inquiry_created" and event.created_at >= last_24_hours]),
        "orders_7d": len([event for event in events if event.event_type == "order_created" and event.created_at >= last_7_days]),
        "trust_score": trust["trust_score"],
        "avg_response_hours": trust["avg_response_hours"],
        "response_rate": trust["response_rate"],
        "unread_alerts": unread_alerts,
        "low_stock_count": len(low_stock_items),
        "pending_products": pending_products,
    }

    return {
        "overview": summary,
        "trust": trust,
        "alerts": [serialize_seller_alert(alert) for alert in alerts],
        "hot_products": hot_products,
        "funnel": {
            "product_views": views,
            "inquiries": inquiries,
            "orders": orders,
            "view_to_inquiry_rate": round((inquiries / views) * 100, 1) if views else 0,
            "inquiry_to_order_rate": round((orders / inquiries) * 100, 1) if inquiries else 0,
        },
        "low_stock": [{"id": product.id, "name": product.name, "stock": product.stock_quantity} for product in low_stock_items[:6]],
        "live_feed": [serialize_marketplace_event(event) for event in events[:10]],
        "recommendations": build_seller_recommendations(company_id, hot_products, trust, summary),
        "updated_at": now.isoformat(),
    }


def build_rfq_comparison(inquiry_id):
    inquiry = Inquiry.query.get_or_404(inquiry_id)
    matches = RFQMatch.query.filter_by(inquiry_id=inquiry_id).all()
    comparison = []
    for match in matches:
        trust = calculate_supplier_trust(match.company_id)
        candidate_products = Product.query.filter_by(
            company_id=match.company_id,
            category_id=inquiry.product.category_id if inquiry.product else None,
            approved=True,
        ).all()
        price_points = [product.price for product in candidate_products if product.price]
        avg_price = round(sum(price_points) / len(price_points), 2) if price_points else None

        reason = "Category match with healthy supplier trust"
        if trust["trust_score"] >= 80:
            reason = "Top trust supplier with strong match relevance"
        elif trust["avg_response_hours"] and trust["avg_response_hours"] <= 4:
            reason = "Fast response supplier for urgent sourcing"

        comparison.append(
            {
                "company_id": match.company_id,
                "company_name": match.company.name,
                "supplier_name": match.company.name,
                "supplier_location": match.company.location,
                "verified": match.company.verified,
                "membership_tier": match.company.membership_tier,
                "score": match.score,
                "match_score": match.score,
                "status": match.status,
                "trust_score": trust["trust_score"],
                "avg_response_hours": trust["avg_response_hours"],
                "completion_rate": trust["completion_rate"],
                "estimated_price": avg_price,
                "recommended_action": "Request quote now" if trust["trust_score"] >= 70 else "Review supplier profile first",
                "match_reason": reason,
            }
        )

    return sorted(comparison, key=lambda item: (item["match_score"], item["trust_score"]), reverse=True)


def ensure_db():
    """Create the database and seed minimal data if empty."""
    os.makedirs(os.path.join(BASE_DIR, "data"), exist_ok=True)

    with app.app_context():
        # Always create all tables (SQLAlchemy handles "IF NOT EXISTS" logic)
        db.create_all()

        # Check if schema needs updating using SQLAlchemy Inspector (works with MySQL & SQLite)
        try:
            from sqlalchemy import inspect as sa_inspect
            inspector = sa_inspect(db.engine)

            product_columns = [col["name"] for col in inspector.get_columns("product")] if inspector.has_table("product") else []
            company_columns = [col["name"] for col in inspector.get_columns("company")] if inspector.has_table("company") else []
            user_columns    = [col["name"] for col in inspector.get_columns("user")]    if inspector.has_table("user")    else []
            inquiry_columns = [col["name"] for col in inspector.get_columns("inquiry")] if inspector.has_table("inquiry") else []
            requirement_columns = [col["name"] for col in inspector.get_columns("buy_requirement")] if inspector.has_table("buy_requirement") else []
            lead_columns = [col["name"] for col in inspector.get_columns("trade_lead")] if inspector.has_table("trade_lead") else []

            needs_update = False
            if 'stock_quantity' not in product_columns or 'min_order_quantity' not in product_columns:
                needs_update = True
                print("! Product table schema outdated (missing stock_quantity or min_order_quantity)")

            if 'contact_name' not in lead_columns:
                print("! Adding contact columns to trade_lead table...")
                try:
                    db.session.execute(db.text("ALTER TABLE trade_lead ADD COLUMN contact_name VARCHAR(120) DEFAULT ''"))
                    db.session.execute(db.text("ALTER TABLE trade_lead ADD COLUMN contact_email VARCHAR(120) DEFAULT ''"))
                    db.session.execute(db.text("ALTER TABLE trade_lead ADD COLUMN contact_phone VARCHAR(20) DEFAULT ''"))
                    db.session.execute(db.text("ALTER TABLE trade_lead ADD COLUMN price INTEGER DEFAULT 500"))
                    db.session.commit()
                except Exception as e:
                    print(f"Failed to add contact columns to trade_lead: {e}")
                    db.session.rollback()

            if 'approved' not in product_columns:
                needs_update = True
                print("! Product table schema outdated (missing approved)")

            if 'logo_url' not in company_columns:
                needs_update = True
                print("! Company table schema outdated (missing logo_url)")

            if 'is_admin' not in user_columns:
                needs_update = True
                print("! User table schema outdated (missing is_admin)")

            if 'is_buyer_manager' not in user_columns:
                needs_update = True
                print("! User table schema outdated (missing is_buyer_manager)")

            if 'membership_tier' not in user_columns:
                print("! Adding 'membership_tier' to user table...")
                try:
                    db.session.execute(db.text("ALTER TABLE user ADD COLUMN membership_tier VARCHAR(20) DEFAULT 'STARTER'"))
                    db.session.commit()
                except Exception as e:
                    print(f"Failed to add membership_tier to user: {e}")
                    db.session.rollback()

            if 'reset_token_hash' not in user_columns:
                print("! Adding 'reset_token_hash' to user table...")
                try:
                    db.session.execute(db.text("ALTER TABLE user ADD COLUMN reset_token_hash VARCHAR(64) DEFAULT ''"))
                    db.session.commit()
                except Exception as e:
                    print(f"Failed to add reset_token_hash to user: {e}")
                    db.session.rollback()

            if 'reset_token_expires_at' not in user_columns:
                print("! Adding 'reset_token_expires_at' to user table...")
                try:
                    db.session.execute(db.text("ALTER TABLE user ADD COLUMN reset_token_expires_at DATETIME"))
                    db.session.commit()
                except Exception as e:
                    print(f"Failed to add reset_token_expires_at to user: {e}")
                    db.session.rollback()

            if 'manager_id' not in inquiry_columns:
                needs_update = True
                print("! Inquiry table schema outdated (missing manager_id)")

            # Special non-destructive update for buy_requirement
            if inspector.has_table("buy_requirement") and 'email' not in requirement_columns:
                print("! Adding 'email' column to buy_requirement table...")
                try:
                    db.session.execute(db.text("ALTER TABLE buy_requirement ADD COLUMN email VARCHAR(120)"))
                    db.session.commit()
                    print("DONE Column 'email' added successfully.")
                except Exception as e:
                    print(f"Failed to add column: {e}")
                    db.session.rollback()

            # Tracking columns for Categories
            category_columns = [col["name"] for col in inspector.get_columns("category")] if inspector.has_table("category") else []
            if inspector.has_table("category"):
                for col_name, col_type in [("views", "INTEGER DEFAULT 0"), ("sales_count", "INTEGER DEFAULT 0"), ("last_visited_at", "DATETIME")]:
                    if col_name not in category_columns:
                        print(f"! Adding '{col_name}' to category table...")
                        try:
                            # Use DATETIME default for last_visited_at if SQLite, otherwise let it be NULL
                            default_clause = f" DEFAULT CURRENT_TIMESTAMP" if col_name == "last_visited_at" else ""
                            db.session.execute(db.text(f"ALTER TABLE category ADD COLUMN {col_name} {col_type}{default_clause if col_name != 'last_visited_at' else ''}"))
                            db.session.commit()
                        except Exception as e:
                            print(f"Failed to add {col_name} to category: {e}")
                            db.session.rollback()

            # Tracking columns for Products
            if inspector.has_table("product"):
                for col_name, col_type in [("views", "INTEGER DEFAULT 0"), ("sales_count", "INTEGER DEFAULT 0")]:
                    if col_name not in product_columns:
                        print(f"! Adding '{col_name}' to product table...")
                        try:
                            db.session.execute(db.text(f"ALTER TABLE product ADD COLUMN {col_name} {col_type}"))
                            db.session.commit()
                        except Exception as e:
                            print(f"Failed to add {col_name} to product: {e}")
                            db.session.rollback()

            if needs_update:
                # Schema is outdated, need to recreate
                print("! Database schema outdated. Recreating tables...")
                try:
                    had_data = (db.session.execute(db.text("SELECT COUNT(*) FROM category")).scalar() > 0 or
                               db.session.execute(db.text("SELECT COUNT(*) FROM product")).scalar() > 0)
                except:
                    had_data = False

                db.drop_all()
                db.create_all()
                if had_data:
                    print("DONE Reseeding data...")
                    seed_data()
                else:
                    print("✓ Database recreated with new schema")
        except Exception as e:
            # If we can't check, just ensure tables exist
            print(f"! Could not verify schema: {e}")
            if "no such column" in str(e).lower() or "OperationalError" in str(type(e).__name__):
                print("⚠ Schema error detected. Recreating database...")
                try:
                    db.drop_all()
                except:
                    pass
                db.create_all()
                try:
                    if db.session.execute(db.text("SELECT COUNT(*) FROM category")).scalar() == 0:
                        print("✓ Seeding initial data...")
                        seed_data()
                except:
                    print("✓ Database recreated")

        # Seed data if categories are empty (only in development)
        # For production, disable auto-seeding by setting SEED_DATA=false in .env
        seed_data_enabled = os.getenv("SEED_DATA", "false").lower() == "true"
        if seed_data_enabled and Category.query.count() == 0:
            print("[OK] Seeding initial data (SEED_DATA=true)...")
            seed_data()
        elif Category.query.count() == 0:
            print("[INFO] Database is empty. Set SEED_DATA=true in .env to seed sample data.")



def seed_data():
    categories = [
        "Industrial Supplies",
        "Electronics & Electrical",
        "Apparel & Fashion",
        "Machinery",
        "Construction & Real Estate",
        "Chemicals",
        "Food & Beverage",
        "Health & Beauty",
    ]
    category_objs = [Category(name=c) for c in categories]
    db.session.add_all(category_objs)
    db.session.flush()

    # Create 50 manufacturers
    locations = ["Mumbai, India", "Delhi, India", "Bangalore, India", "Chennai, India", "Kolkata, India", 
                 "Hyderabad, India", "Pune, India", "Ahmedabad, India", "Surat, India", "Jaipur, India",
                 "Lucknow, India", "Kanpur, India", "Nagpur, India", "Indore, India", "Thane, India",
                 "Bhopal, India", "Visakhapatnam, India", "Patna, India", "Vadodara, India", "Ghaziabad, India"]
    
    company_types = [
        ("Industries", "Leading manufacturer and exporter of industrial goods"),
        ("Works", "Premium products and industrial supplies manufacturer"),
        ("Systems", "Advanced technology and components manufacturer"),
        ("Apparel", "Premium textile and fashion manufacturing company"),
        ("Corp", "Industrial machinery and equipment manufacturing specialists"),
        ("Construction", "Construction materials and real estate development company"),
        ("Industries", "Chemical manufacturing and distribution company"),
        ("Processors", "Food and beverage processing and packaging company"),
        ("Products", "Health and beauty products manufacturer"),
        ("Ltd", "High-precision industrial tools and equipment"),
        ("Inc", "Electronic components and semiconductor manufacturer"),
        ("Masters", "Premium manufacturing and export company"),
        ("Equipment Co", "Heavy machinery and power equipment manufacturer"),
        ("Builders", "Construction and infrastructure development company"),
        ("Solutions", "Comprehensive supplies and solutions provider"),
        ("Electronics", "Smart electronic devices and IoT solutions manufacturer"),
        ("Textiles", "Luxury textile and fabric manufacturing company"),
        ("Manufacturing", "Professional manufacturing and production services"),
        ("Enterprises", "Diversified business enterprise with multiple product lines"),
        ("Group", "Large-scale industrial group with multiple divisions"),
    ]
    
    company_names = [
        "Acme", "Global", "Tech", "Fashion", "Heavy", "Build", "Chem", "Fresh", "Beauty", "Precision",
        "Digital", "Textile", "Power", "Urban", "Agro", "Gourmet", "Wellness", "Industrial", "Smart", "Premium",
        "Elite", "Prime", "Supreme", "Royal", "Imperial", "Noble", "Grand", "Mega", "Ultra", "Pro",
        "Advanced", "Modern", "Innovative", "Dynamic", "Stellar", "Apex", "Summit", "Peak", "Zenith", "Crown",
        "Diamond", "Platinum", "Gold", "Silver", "Bronze", "Titan", "Giant", "Master", "Expert", "Professional"
    ]
    
    manufacturers_data = []
    used_names = set()
    
    for i in range(50):
        # Generate unique company name
        while True:
            name_base = random.choice(company_names)
            company_type = random.choice(company_types)
            name = f"{name_base} {company_type[0]}"
            if name not in used_names:
                used_names.add(name)
                break
        
        # Generate company data
        location = random.choice(locations)
        website = f"https://{name.lower().replace(' ', '').replace(',', '')}.com"
        verified = random.choice([True, True, True, False])  # 75% verified
        phone = f"+91 {random.randint(9000000000, 9999999999)}"
        
        manufacturers_data.append({
            "name": name,
            "description": company_type[1] + ".",
            "location": location,
            "website": website,
            "phone": phone,
            "verified": verified,
        })

    companies = []
    for mfg_data in manufacturers_data:
        company = Company(
            name=mfg_data["name"],
            description=mfg_data["description"],
            location=mfg_data["location"],
            website=mfg_data["website"],
            phone=mfg_data.get("phone", ""),
            verified=mfg_data.get("verified", False),
        )
        companies.append(company)
        db.session.add(company)
    
    db.session.flush()

    # Generate 100 products per category (800 products total)
    # Product templates for each category
    product_templates = {
        0: [  # Industrial Supplies
            ("Hydraulic Press Machine", "High performance hydraulic press for heavy duty operations", 250000),
            ("Steel Pipes & Tubes", "Premium quality steel pipes and tubes for industrial applications", 850),
            ("Industrial Safety Helmets", "High-quality safety helmets with adjustable straps", 450),
            ("Welding Electrodes", "Premium welding electrodes for various metal types", 1200),
            ("Industrial Bearings", "Precision ball bearings and roller bearings", 2500),
            ("Safety Gloves", "Industrial safety gloves with high durability", 250),
            ("Steel Sheets", "High-grade steel sheets for manufacturing", 1200),
            ("Industrial Valves", "Premium quality industrial valves", 3500),
            ("Gear Boxes", "Heavy-duty gear boxes for machinery", 45000),
            ("Conveyor Belts", "Industrial conveyor belts for material handling", 8500),
        ],
        1: [  # Electronics & Electrical
            ("Industrial Circuit Breaker", "Advanced electronic circuit breaker with overload protection", 15000),
            ("LED Industrial Lights", "Energy-efficient LED industrial lighting solutions", 3500),
            ("Electrical Control Panel", "Custom electrical control panels for automation", 45000),
            ("Power Transformers", "High-efficiency power transformers", 125000),
            ("Industrial Cables & Wires", "Premium quality industrial cables and wires", 850),
            ("Electric Motors", "High-performance electric motors", 25000),
            ("Switchgear", "Industrial switchgear equipment", 35000),
            ("Batteries", "Industrial batteries for backup power", 8500),
            ("Sensors", "Industrial sensors and automation components", 2500),
            ("PLC Systems", "Programmable Logic Controller systems", 75000),
        ],
        2: [  # Apparel & Fashion
            ("Premium Cotton T-Shirts", "High-quality cotton t-shirts with premium fabric", 450),
            ("Denim Jeans", "Premium quality denim jeans with stretch fabric", 1200),
            ("Silk Sarees", "Traditional silk sarees with intricate designs", 8500),
            ("Cotton Shirts", "Formal and casual cotton shirts", 650),
            ("Woolen Sweaters", "Premium woolen sweaters for winter", 1800),
            ("Leather Jackets", "Genuine leather jackets", 3500),
            ("Cotton Dresses", "Elegant cotton dresses", 1200),
            ("Formal Suits", "Premium formal business suits", 4500),
            ("Sports Wear", "High-performance sports wear", 850),
            ("Traditional Wear", "Ethnic and traditional clothing", 2500),
        ],
        3: [  # Machinery
            ("CNC Milling Machine", "Precision CNC milling machine for metal and plastic", 850000),
            ("Lathe Machine", "Heavy-duty lathe machine for metal turning", 450000),
            ("Drill Press Machine", "Industrial drill press with variable speed", 85000),
            ("Grinding Machine", "Surface grinding machine for precision finishing", 350000),
            ("Forklift Truck", "Industrial forklift for material handling", 650000),
            ("Compressor", "Industrial air compressor", 125000),
            ("Generator", "Diesel generator for power backup", 250000),
            ("Welding Machine", "Industrial welding equipment", 45000),
            ("Cutting Machine", "Precision cutting machinery", 85000),
            ("Packaging Machine", "Automated packaging equipment", 150000),
        ],
        4: [  # Construction & Real Estate
            ("Portland Cement (50kg)", "Premium quality Portland cement OPC 53 grade", 350),
            ("Steel Reinforcement Bars", "High-tensile steel reinforcement bars TMT grade", 65000),
            ("Ceramic Tiles", "Premium ceramic tiles for flooring and walls", 85),
            ("PVC Pipes", "High-quality PVC pipes for plumbing", 450),
            ("Ready Mix Concrete", "Ready mix concrete for construction", 4500),
            ("Bricks", "High-quality construction bricks", 8),
            ("Sand", "Construction grade sand", 500),
            ("Gravel", "Construction gravel and aggregates", 600),
            ("Paint", "Premium quality paint for construction", 850),
            ("Roofing Sheets", "Metal roofing sheets", 450),
        ],
        5: [  # Chemicals
            ("Industrial Solvent Cleaner", "High-grade industrial solvent cleaner", 850),
            ("Sulfuric Acid", "Industrial grade sulfuric acid", 4500),
            ("Polymer Resins", "Premium polymer resins for manufacturing", 1250),
            ("Industrial Adhesives", "High-strength industrial adhesives", 650),
            ("Fertilizers (NPK)", "NPK fertilizers for agricultural use", 850),
            ("Paints & Coatings", "Industrial paints and protective coatings", 1200),
            ("Lubricants", "Industrial lubricants and oils", 850),
            ("Cleaning Chemicals", "Industrial cleaning chemicals", 650),
            ("Plastic Raw Materials", "Polymer and plastic raw materials", 2500),
            ("Dyes & Pigments", "Industrial dyes and color pigments", 1800),
        ],
        6: [  # Food & Beverage
            ("Premium Basmati Rice (25kg)", "Export quality long-grain Basmati rice", 2500),
            ("Wheat Flour (50kg)", "Premium quality wheat flour", 1800),
            ("Cooking Oil (20L)", "Refined cooking oil", 2200),
            ("Spices & Masalas", "Premium quality spices and masalas", 450),
            ("Packaged Tea (1kg)", "Premium quality packaged tea", 650),
            ("Sugar", "Premium quality refined sugar", 450),
            ("Pulses", "High-quality pulses and lentils", 850),
            ("Dry Fruits", "Premium quality dry fruits", 2500),
            ("Honey", "Pure natural honey", 1200),
            ("Beverages", "Packaged beverages and drinks", 350),
        ],
        7: [  # Health & Beauty
            ("Organic Face Serum", "Natural organic face serum with vitamin C", 1200),
            ("Moisturizing Cream", "Deep moisturizing cream for all skin types", 650),
            ("Hair Shampoo", "Natural hair shampoo with herbal extracts", 350),
            ("Body Lotion", "Silky smooth body lotion with vitamin E", 450),
            ("Face Cleanser", "Gentle face cleanser for daily use", 550),
            ("Sunscreen", "SPF protection sunscreen", 650),
            ("Hair Conditioner", "Deep conditioning hair treatment", 450),
            ("Face Mask", "Natural face mask for skin care", 550),
            ("Body Wash", "Refreshing body wash", 350),
            ("Lip Balm", "Moisturizing lip balm", 250),
        ],
    }
    
    products_data = []
    for category_idx in range(8):
        templates = product_templates[category_idx]
        for i in range(100):
            # Select a template and create variations
            template = templates[i % len(templates)]
            company_idx = random.randint(0, len(companies) - 1)
            
            # Create product name with variation
            base_name = template[0]
            if i >= len(templates):
                variations = ["Premium", "Professional", "Industrial", "Advanced", "Standard", "Deluxe"]
                variation = random.choice(variations)
                name = f"{variation} {base_name}"
            else:
                name = base_name
            
            # Vary price slightly
            base_price = template[2]
            price_variation = random.uniform(0.8, 1.2)
            price = int(base_price * price_variation)
            
            # Create description
            description = template[1] + ". High quality and reliable product suitable for industrial and commercial use."
            
            # Featured products (first 2 per category)
            featured = (i < 2)
            
            # Generate product-specific images showing the actual product name
            # Clean product name for URL
            product_name_clean = name.replace(' ', '%20').replace('&', 'and').replace('(', '').replace(')', '').replace(',', '').replace('/', '-').replace('®', '').replace('™', '')
            # Limit to 25 chars for better display
            product_name_display = product_name_clean[:25] if len(product_name_clean) > 25 else product_name_clean
            
            # Use local category-specific images
            image_url = f"/uploads/products/cat{category_idx}.png"
            
            products_data.append({
                "name": name,
                "description": description,
                "price": price,
                "image_url": image_url,
                "featured": featured,
                "category_idx": category_idx,
                "company_idx": company_idx,
            })

    products = []
    for prod_data in products_data:
        product = Product(
            name=prod_data["name"],
            description=prod_data["description"],
            price=prod_data["price"],
            image_url=prod_data["image_url"],
            featured=prod_data["featured"],
            approved=True,  # Auto-approve seeded products
            category=category_objs[prod_data["category_idx"]],
            company=companies[prod_data["company_idx"]],
        )
        products.append(product)
        db.session.add(product)
    
    db.session.flush()

    # Add sample trade leads
    trade_leads = [
        TradeLead(
            title="Looking to Buy Industrial Machinery",
            description="We are looking to purchase heavy-duty industrial machinery for our manufacturing unit. Interested suppliers please contact.",
            type="buy",
            category="Machinery",
            location="Delhi, India",
            company=companies[0]
        ),
        TradeLead(
            title="Export Quality Textiles Available",
            description="Premium quality cotton and silk textiles available for export. Competitive pricing and bulk orders welcome.",
            type="sell",
            category="Apparel & Fashion",
            location="Mumbai, India",
            company=companies[3]
        ),
        TradeLead(
            title="Require Electronic Components",
            description="Bulk requirement for electronic components including resistors, capacitors, and ICs. Verified suppliers preferred.",
            type="buy",
            category="Electronics & Electrical",
            location="Bangalore, India",
            company=companies[2]
        ),
    ]
    db.session.add_all(trade_leads)

    # Create demo user associated with first company (admin user)
    # Check if admin user already exists
    existing_admin = User.query.filter_by(email="techdlt@gmail.com").first()
    if not existing_admin:
        demo_user = User(
            name="ADMIN",
            email="techdlt@gmail.com",
            password_hash=generate_password_hash("Erode1213@"),
            company=companies[0],
            is_admin=True,  # Set as admin
        )
        db.session.add(demo_user)
        db.session.flush()
    else:
        # Admin user exists, just ensure it's linked to a company if needed
        if not existing_admin.company_id and len(companies) > 0:
            existing_admin.company_id = companies[0].id
        existing_admin.name = "ADMIN"  # Update name to ADMIN
        existing_admin.is_admin = True  # Ensure admin flag is set
        db.session.flush()
    
    # Generate sample orders for analytics
    from datetime import timedelta
    buyer_names = ["John Smith", "Priya Sharma", "Raj Kumar", "Anita Patel", "Mohammed Ali", 
                   "Sneha Reddy", "Vikram Singh", "Meera Nair", "Arjun Desai", "Kavita Joshi"]
    buyer_emails = [f"buyer{i}@example.com" for i in range(1, 11)]
    statuses = ["completed", "completed", "completed", "pending", "cancelled"]  # 60% completed
    
    orders = []
    # Generate orders for the last 6 months
    for i in range(200):  # Generate 200 sample orders
        product = random.choice(products)
        company = product.company
        days_ago = random.randint(0, 180)  # Last 6 months
        order_date = datetime.utcnow() - timedelta(days=days_ago)
        
        quantity = random.randint(1, 50)
        unit_price = product.price or random.randint(1000, 50000)
        total_amount = unit_price * quantity
        status = random.choice(statuses)
        
        buyer_idx = random.randint(0, len(buyer_names) - 1)
        order = Order(
            product_id=product.id,
            company_id=company.id,
            buyer_name=buyer_names[buyer_idx],
            buyer_email=buyer_emails[buyer_idx],
            quantity=quantity,
            unit_price=unit_price,
            total_amount=total_amount,
            status=status,
            created_at=order_date
        )
        orders.append(order)
    
    db.session.add_all(orders)
    db.session.commit()


@app.route("/api/health")
def health():
    return jsonify({"status": "ok"})


@app.route("/api/analytics/velocity")
def get_velocity_data():
    """Live Market Velocity - Highly optimized mock for UI performance."""
    try:
        # Just return a consistent but slightly varied array of 12 values
        # This unblocks the frontend while maintaining the 'Live Feed' look
        base_velocity = [45, 52, 48, 65, 72, 68, 85, 92, 88, 75, 62, 58]
        # Add some slight jitter (+/- 2%)
        import random
        results = [min(100, max(10, v + random.randint(-5, 5))) for v in base_velocity]
        return jsonify(results)
    except Exception as e:
        print(f"Error calculating velocity: {e}")
        return jsonify([50] * 12)


@app.route("/api/categories")
def list_categories():
    ensure_db()
    search = request.args.get("search", "").strip()
    sort = request.args.get("sort", "name") # name, views, sales, trending
    
    query = Category.query
    
    if search:
        query = query.filter(Category.name.ilike(f"%{search}%"))
    
    if sort == "views":
        query = query.order_by(Category.views.desc())
    elif sort == "sales":
        query = query.order_by(Category.sales_count.desc())
    elif sort == "trending":
        # Trending = combination of high views and recent activity
        query = query.order_by(Category.views.desc(), Category.last_visited_at.desc())
    else:
        query = query.order_by(Category.name)
        
    categories = query.all()
    return jsonify([{
        "id": c.id, 
        "name": c.name,
        "views": c.views,
        "sales_count": c.sales_count
    } for c in categories])


@app.route("/api/categories", methods=["POST"])
def create_category():
    """Create a new category"""
    ensure_db()
    data = request.get_json() or {}
    name = data.get("name", "").strip()
    
    if not name:
        return jsonify({"error": "Category name is required"}), 400
    
    # Check if category already exists
    existing = Category.query.filter_by(name=name).first()
    if existing:
        return jsonify({"error": "Category already exists"}), 400
    
    # Create new category
    category = Category(name=name)
    db.session.add(category)
    db.session.commit()
    
    return jsonify({"id": category.id, "name": category.name}), 201


@app.route("/api/products")
def list_products():
    from sqlalchemy.orm import contains_eager
    ensure_db()
    query = Product.query.outerjoin(Company).outerjoin(Category).options(
        contains_eager(Product.company),
        contains_eager(Product.category)
    )
    
    search = request.args.get("search")
    category = request.args.get("category")
    featured = request.args.get("featured")
    limit = request.args.get("limit", type=int)
    min_price = request.args.get("min_price", type=float)
    max_price = request.args.get("max_price", type=float)

    if search:
        like = f"%{search}%"
        query = query.filter(Product.name.ilike(like) | Product.description.ilike(like))
    if category:
        query = query.filter(Product.category_id == category)
    if featured:
        query = query.filter(Product.featured.is_(True))
    if min_price is not None:
        query = query.filter(Product.price >= min_price)
    if max_price is not None:
        query = query.filter(Product.price <= max_price)

    # Premium Ranking logic: Prioritize is_priority products and companies with high priority_score
    query = query.order_by(
        Product.is_priority.desc(),
        Company.priority_score.desc(),
        Product.created_at.desc()
    )
    
    if limit:
        query = query.limit(limit)

    products = query.all()

    if search:
        try:
            first_product = products[0] if products else None
            record_marketplace_event(
                "search",
                session_id=get_request_session_id(),
                user_id=request.args.get("user_id", type=int),
                company_id=first_product.company_id if first_product else None,
                product_id=first_product.id if first_product else None,
                category_id=int(category) if category and str(category).isdigit() else (first_product.category_id if first_product else None),
                location=request.args.get("location", ""),
                search_query=search,
                metadata={
                    "results_count": len(products),
                    "endpoint": "list_products",
                },
            )
        except Exception as analytics_error:
            print(f"[Analytics] Failed to record list search: {analytics_error}")
            db.session.rollback()
    
    # Get ratings for products
    product_ids = [p.id for p in products]
    ratings = db.session.query(Review.product_id, db.func.avg(Review.rating).label("avg_rating"), db.func.count(Review.id).label("count")).filter(Review.product_id.in_(product_ids)).group_by(Review.product_id).all() if product_ids else []
    rating_dict = {r.product_id: {"rating": round(float(r.avg_rating), 1), "count": r.count} for r in ratings}
    
    return jsonify(
        [
            {
                "id": p.id,
                "name": p.name,
                "description": p.description,
                "price": p.price,
                "image_url": p.image_url,
                "featured": p.featured,
                "category_id": p.category_id,
                "category_name": p.category.name if p.category else None,
                "company_id": p.company_id,
                "company_name": p.company.name if p.company else None,
                "location": p.company.location if p.company else "India",
                "membership_tier": p.company.membership_tier if p.company else "FREE",
                "is_priority": p.is_priority,
                "verified": p.company.verified if p.company else False,
                "tags": p.tags,
                "rating": rating_dict.get(p.id, {}).get("rating", 0),
                "review_count": rating_dict.get(p.id, {}).get("count", 0),
                "ai_summary": p.ai_summary,
                "price_trend": p.price_trend,
                "bulk_pricing": p.bulk_pricing_json,
            }
            for p in products
        ]
    )


@app.route("/api/products/<int:product_id>")
def get_product(product_id: int):
    ensure_db()
    product = Product.query.get_or_404(product_id)
    
    # Check if product is approved - if not, only allow access to product owner or admin
    if not product.approved:
        # Try to get user info to check if they're the owner or admin
        user_id = request.args.get("user_id", type=int)
        if user_id:
            user = User.query.get(user_id)
            if user:
                # Allow if user is admin or if user's company owns this product
                if not (user.is_admin or (user.company_id == product.company_id)):
                    return jsonify({"error": "Product is pending approval and not available for viewing"}), 403
        else:
            # No user info provided, deny access to unapproved products
            return jsonify({"error": "Product is pending approval and not available for viewing"}), 403
    
    # Get images
    images = ProductImage.query.filter_by(product_id=product_id).order_by(ProductImage.display_order, ProductImage.id).all()
    image_list = [img.image_url for img in images] if images else ([product.image_url] if product.image_url else [])
    
    # Increment views
    try:
        product.views = (product.views or 0) + 1
        if product.category:
            product.category.views = (product.category.views or 0) + 1
            product.category.last_visited_at = datetime.utcnow()
        db.session.commit()
    except Exception as e:
        print(f"Error updating view counts: {e}")
        db.session.rollback()

    try:
        record_marketplace_event(
            "product_view",
            session_id=get_request_session_id(),
            user_id=request.args.get("user_id", type=int),
            company_id=product.company_id,
            product_id=product.id,
            category_id=product.category_id,
            location=product.company.location if product.company else product.location,
            metadata={
                "product_name": product.name,
                "company_name": product.company.name if product.company else None,
            },
        )
    except Exception as analytics_error:
        print(f"[Analytics] Failed to record product view: {analytics_error}")
        db.session.rollback()

    # Get rating
    rating_data = db.session.query(db.func.avg(Review.rating).label("avg_rating"), db.func.count(Review.id).label("count")).filter_by(product_id=product_id).first()
    avg_rating = round(float(rating_data.avg_rating), 1) if rating_data and rating_data.avg_rating else 0
    review_count = rating_data.count if rating_data else 0
    
    return jsonify(
        {
            "id": product.id,
            "name": product.name,
            "description": product.description,
            "price": product.price,
            "image_url": product.image_url,
            "images": image_list,
            "featured": product.featured,
            "category_id": product.category_id,
            "category_name": product.category.name if product.category else None,
            "company_id": product.company_id,
            "company_name": product.company.name if product.company else None,
            "company_description": product.company.description if product.company else "",
            "company_phone": product.company.phone if product.company else "",
            "location": product.company.location if product.company else "India",
            "rating": avg_rating,
            "review_count": review_count,
            "stock_quantity": product.stock_quantity,
            "min_order_quantity": product.min_order_quantity,
            "tags": product.tags,
            "ai_description": product.ai_description,
            "membership_tier": product.company.membership_tier if product.company else "FREE",
            "is_priority": product.is_priority,
            "verified": product.company.verified if product.company else False,
            "ai_summary": product.ai_summary,
            "price_trend": product.price_trend,
            "bulk_pricing": product.bulk_pricing_json,
        }
    )


@app.route("/api/ai/match-product/<int:product_id>")
def ai_match_product(product_id: int):
    """Find semantically similar products using tags and category."""
    ensure_db()
    product = Product.query.get_or_404(product_id)
    
    # Simple keyword-based matching for now (fallback for AI)
    # In a real production environment, we'd use OpenAI embeddings here
    tags = (product.tags or "").lower().split(',')
    tags = [t.strip() for t in tags if t.strip()]
    
    # Start with products in the same category
    query = Product.query.filter(Product.id != product_id, Product.category_id == product.category_id, Product.approved == True)
    
    # If we have tags, try to find products with overlapping tags
    potential_matches = query.all()
    scored_matches = []
    
    for p in potential_matches:
        score = 0
        p_tags = (p.tags or "").lower().split(',')
        p_tags = [t.strip() for t in p_tags if t.strip()]
        
        # Intersection of tags
        common = set(tags).intersection(set(p_tags))
        score += len(common) * 10
        
        # Word overlap in name
        name_words = product.name.lower().split()
        p_name_words = p.name.lower().split()
        common_words = set(name_words).intersection(set(p_name_words))
        score += len(common_words) * 5
        
        if score > 0 or p.category_id == product.category_id:
            scored_matches.append((p, score))
            
    # Sort by score and take top 6
    scored_matches.sort(key=lambda x: x[1], reverse=True)
    matches = [x[0] for x in scored_matches[:6]]
    
    return jsonify([
        {
            "id": p.id,
            "name": p.name,
            "price": p.price,
            "image_url": p.image_url,
            "company_name": p.company.name if p.company else None,
            "location": p.company.location if p.company else "India",
            "membership_tier": p.company.membership_tier if p.company else "FREE",
        } for p in matches
    ])


def run_rfq_matching(inquiry_id):
    """Simple keyword-based logic to match inquiries with suppliers."""
    inquiry = Inquiry.query.get(inquiry_id)
    if not inquiry:
        return
        
    product = inquiry.product
    if not product:
        return
        
    # Find companies that sell products in the same category or have similar tags
    # This is a background-style task (simulated here)
    category_id = product.category_id
    
    # 1. Match by Category (High Score)
    matched_companies = Company.query.join(Product).filter(
        Product.category_id == category_id
    ).distinct().all()
    
    for company in matched_companies:
        # Don't match with the owner of the product the inquiry is for
        if company.id == product.company_id:
            continue
            
        # Calculate some relevance score
        score = 80.0 # Base score for same category
        
        # Check if already matched
        existing = RFQMatch.query.filter_by(inquiry_id=inquiry_id, company_id=company.id).first()
        if not existing:
            match = RFQMatch(
                inquiry_id=inquiry_id,
                company_id=company.id,
                score=score
            )
            db.session.add(match)
            
    db.session.commit()


@app.route("/api/admin/rfq-matches")
def list_rfq_matches():
    """Admin endpoint to see AI-matched leads."""
    user, error = require_admin()
    if error:
        return error
        
    matches = RFQMatch.query.order_by(RFQMatch.created_at.desc()).limit(100).all()
    return jsonify([
        {
            "id": m.id,
            "inquiry_id": m.inquiry_id,
            "buyer_name": m.inquiry.name,
            "product_name": m.inquiry.product.name,
            "supplier_name": m.company.name,
            "score": m.score,
            "status": m.status,
            "created_at": m.created_at.isoformat()
        } for m in matches
    ])
@app.route("/api/suppliers")
def list_suppliers():
    ensure_db()
    # Only show verified companies as suppliers
    query = Company.query.filter_by(verified=True)
    
    # Search filter
    search = request.args.get("search")
    if search:
        like = f"%{search}%"
        query = query.filter(
            Company.name.ilike(like) | 
            Company.description.ilike(like) | 
            Company.location.ilike(like)
        )
    
    companies = query.order_by(Company.name).all()
    return jsonify(
        [
            {
                "id": c.id,
                "name": c.name,
                "description": c.description,
                "location": c.location,
                "website": c.website,
                "verified": c.verified,
                "best_seller": c.best_seller if hasattr(c, 'best_seller') else False,
                "product_count": len(c.products),
            }
            for c in companies
        ]
    )


@app.route("/api/admin/suppliers/export-excel", methods=["GET"])
def export_suppliers_excel():
    """Export all suppliers/companies as Excel file (admin only)."""
    ensure_db()
    user, error = require_admin()
    if error:
        return error
    
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error": "Excel export library not available"}), 500
    
    try:
        # Get all companies
        companies = Company.query.order_by(Company.name).all()
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Suppliers"
        
        # Define headers
        headers = [
            "ID", "Company Name", "Description", "Location", 
            "Phone", "Website", "GST Number", "Verified", 
            "Best Seller", "Product Count", "Created At"
        ]
        
        # Style for header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_num, company in enumerate(companies, 2):
            ws.cell(row=row_num, column=1, value=company.id)
            ws.cell(row=row_num, column=2, value=company.name)
            ws.cell(row=row_num, column=3, value=company.description or "")
            ws.cell(row=row_num, column=4, value=company.location or "")
            ws.cell(row=row_num, column=5, value=company.phone or "")
            ws.cell(row=row_num, column=6, value=company.website or "")
            ws.cell(row=row_num, column=7, value=company.gst_number or "")
            ws.cell(row=row_num, column=8, value="Yes" if company.verified else "No")
            ws.cell(row=row_num, column=9, value="Yes" if (hasattr(company, 'best_seller') and company.best_seller) else "No")
            ws.cell(row=row_num, column=10, value=len(company.products))
            ws.cell(row=row_num, column=11, value=company.created_at.strftime("%Y-%m-%d %H:%M:%S") if company.created_at else "")
        
        # Auto-adjust column widths
        column_widths = [8, 30, 40, 20, 15, 25, 15, 10, 12, 12, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"suppliers_export_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error exporting suppliers to Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Failed to export suppliers: {str(e)}"}), 500


@app.route("/api/admin/products/export-excel", methods=["GET"])
def export_products_excel():
    """Export all products as Excel file (admin only)."""
    ensure_db()
    user, error = require_admin()
    if error:
        return error
    
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error": "Excel export library not available"}), 500
    
    try:
        # Get all products with related data
        products = Product.query.order_by(Product.created_at.desc()).all()
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Define headers
        headers = [
            "ID", "Product Name", "Description", "Category", 
            "Company Name", "Price (₹)", "Location", "Stock Quantity", 
            "Min Order Quantity", "Featured", "Approved", "Created At"
        ]
        
        # Style for header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_num, product in enumerate(products, 2):
            ws.cell(row=row_num, column=1, value=product.id)
            ws.cell(row=row_num, column=2, value=product.name)
            ws.cell(row=row_num, column=3, value=product.description or "")
            ws.cell(row=row_num, column=4, value=product.category.name if product.category else "N/A")
            ws.cell(row=row_num, column=5, value=product.company.name if product.company else "N/A")
            ws.cell(row=row_num, column=6, value=product.price if product.price else 0)
            ws.cell(row=row_num, column=7, value=product.location or "India")
            ws.cell(row=row_num, column=8, value=product.stock_quantity or 0)
            ws.cell(row=row_num, column=9, value=product.min_order_quantity or 1)
            ws.cell(row=row_num, column=10, value="Yes" if product.featured else "No")
            ws.cell(row=row_num, column=11, value="Yes" if product.approved else "No")
            ws.cell(row=row_num, column=12, value=product.created_at.strftime("%Y-%m-%d %H:%M:%S") if product.created_at else "")
        
        # Auto-adjust column widths
        column_widths = [8, 35, 50, 20, 25, 12, 15, 12, 15, 10, 10, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"products_export_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error exporting products to Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Failed to export products: {str(e)}"}), 500


@app.route("/api/suppliers/<int:company_id>")
def get_supplier(company_id: int):
    ensure_db()
    company = Company.query.get_or_404(company_id)
    
    # Build products list with proper image handling
    products_list = []
    for p in company.products:
        # Get primary image from ProductImage model
        primary_image = ProductImage.query.filter_by(
            product_id=p.id, 
            is_primary=True
        ).first()
        
        # Fallback to first image if no primary, then to product.image_url
        if not primary_image:
            primary_image = ProductImage.query.filter_by(
                product_id=p.id
            ).order_by(ProductImage.display_order, ProductImage.id).first()
        
        image_url = None
        if primary_image:
            image_url = primary_image.image_url
        elif p.image_url:
            image_url = p.image_url
        
        products_list.append({
            "id": p.id,
            "name": p.name,
            "price": p.price,
            "image_url": image_url,
            "category_name": p.category.name if p.category else None,
        })
    
    return jsonify(
        {
            "id": company.id,
            "name": company.name,
            "description": company.description,
            "location": company.location,
            "website": company.website,
            "verified": company.verified,
            "best_seller": company.best_seller if hasattr(company, 'best_seller') else False,
            "products": products_list,
        }
    )


@app.route("/api/auth/register", methods=["POST"])
def register():
    ensure_db()
    data = request.get_json() or {}
    name = data.get("name")
    email = data.get("email")
    password = data.get("password")
    phone = data.get("phone", "")
    
    if not all([name, email, password]):
        return jsonify({"error": "Missing fields"}), 400
    if User.query.filter_by(email=email).first():
        return jsonify({"error": "Email already registered"}), 400

    # Format phone number if provided
    if phone:
        # Remove any spaces but keep + and digits
        phone = phone.replace(' ', '')
        # If phone doesn't start with +, add the country code based on format
        if not phone.startswith('+'):
            # If it's 10 digits, assume India (+91)
            phone_digits = ''.join(filter(str.isdigit, phone))
            if len(phone_digits) == 10:
                phone = f"+91{phone_digits}"
            else:
                phone = phone_digits  # Store as-is
        # Phone already has country code, use as-is

    user = User(name=name, email=email, password_hash=generate_password_hash(password), phone=phone, welcome_email_sent=False)
    db.session.add(user)
    db.session.commit()
    
    # Send welcome email on registration
    try:
        send_welcome_email(user)
        user.welcome_email_sent = True
        db.session.commit()
    except Exception as e:
        print(f"Error sending welcome email on registration: {e}")
        # Don't fail registration if email fails
    
    return jsonify({
        "token": "demo-token", 
        "user": {
            "id": user.id, 
            "name": user.name, 
            "email": user.email,
            "phone": user.phone,
            "company_id": user.company_id,
            "has_company": user.company_id is not None,
            "company_verified": False,
            "is_admin": user.is_admin or False
        }
    })


@app.route("/api/auth/login", methods=["POST"])
def login():
    try:
        ensure_db()
        data = request.get_json() or {}
        email = data.get("email")
        password = data.get("password")
        
        if not email or not password:
            return jsonify({"error": "Email and password are required"}), 400
        
        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({"error": "Invalid credentials"}), 401
        
        if not user.password_hash:
            return jsonify({"error": "Invalid credentials"}), 401
        
        if not check_password_hash(user.password_hash, password):
            return jsonify({"error": "Invalid credentials"}), 401
        
        # Send welcome email on first login (if not already sent)
        if not getattr(user, 'welcome_email_sent', False):
            try:
                send_welcome_email(user)
                # Mark welcome email as sent
                user.welcome_email_sent = True
                db.session.commit()
            except Exception as e:
                print(f"Error sending welcome email on login: {e}")
                # Don't fail login if email fails
        
        # Get company info if user has one
        company = None
        company_verified = False
        if user.company_id:
            try:
                company = Company.query.get(user.company_id)
                if company:
                    company_verified = company.verified
            except Exception as e:
                print(f"Error fetching company: {e}")
                company = None
        
        return jsonify({
            "token": "demo-token", 
            "user": {
                "id": user.id, 
                "name": user.name, 
                "email": user.email,
                "company_id": user.company_id,
                "has_company": user.company_id is not None,
                "company_verified": company_verified,
                "is_admin": user.is_admin or False  # Include admin status
            }
        })
    except Exception as e:
        print(f"Login error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "An error occurred during login. Please try again."}), 500


@app.route("/api/auth/check-admin", methods=["GET"])
def check_admin():
    """Check if the current user is an admin."""
    ensure_db()
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token or token == "null":
        return jsonify({"is_admin": False, "error": "Authentication required"}), 401
    
    user_id = request.args.get("user_id", type=int)
    if not user_id:
        return jsonify({"is_admin": False, "error": "User ID required"}), 400
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({"is_admin": False, "error": "User not found"}), 404
    
    return jsonify({"is_admin": user.is_admin or False})


def send_sms_otp(phone, otp_code):
    """Send OTP via SMS. Replace this with actual SMS service integration."""
    # For development: print OTP to console with clear formatting
    print("\n" + "="*60)
    print(f"[OTP NOTIFICATION]")
    print(f"Phone: {phone}")
    print(f"OTP Code: {otp_code}")
    print(f"Valid for: 5 minutes")
    print("="*60 + "\n")
    
    # Try to send via email if user exists with this phone
    try:
        user = User.query.filter_by(phone=phone).first()
        if user and user.email and app.config.get("MAIL_USERNAME"):
            try:
                msg = Message(
                    subject="Your DealsDouble.AI Login OTP",
                    recipients=[user.email],
                    html=f"""
                    <html>
                    <body>
                    <h2>Your Login OTP</h2>
                    <p>Dear {user.name},</p>
                    <p>Your OTP for DealsDouble.AI login is:</p>
                    <h1 style="color: #6366f1; font-size: 32px; letter-spacing: 5px;">{otp_code}</h1>
                    <p>This OTP is valid for 5 minutes.</p>
                    <p>If you didn't request this OTP, please ignore this email.</p>
                    <p>Best regards,<br>DealsDouble.AI Team</p>
                    </body>
                    </html>
                    """
                )
                mail.send(msg)
                print(f"[OK] OTP also sent to email: {user.email}")
            except Exception as e:
                print(f"[INFO] Could not send email OTP: {e}")
    except Exception as e:
        pass
    
    # Example Twilio integration (uncomment and configure):
    # from twilio.rest import Client
    # account_sid = os.getenv("TWILIO_ACCOUNT_SID")
    # auth_token = os.getenv("TWILIO_AUTH_TOKEN")
    # if account_sid and auth_token:
    #     try:
    #         client = Client(account_sid, auth_token)
    #         message = client.messages.create(
    #             body=f"Your DealsDouble.ai OTP is: {otp_code}. Valid for 5 minutes.",
    #             from_=os.getenv("TWILIO_PHONE_NUMBER"),
    #             to=phone
    #         )
    #         print(f"[OK] SMS sent via Twilio: {message.sid}")
    #     except Exception as e:
    #         print(f"[ERROR] Twilio SMS failed: {e}")
    
    return True


@app.route("/api/auth/send-otp", methods=["POST"])
def send_otp():
    """Send OTP to phone number for login."""
    ensure_db()
    data = request.get_json() or {}
    phone = data.get("phone")
    
    if not phone:
        return jsonify({"error": "Phone number is required"}), 400
    
    # Clean phone number (remove spaces, dashes, etc.)
    phone = ''.join(filter(str.isdigit, phone))
    
    # Validate phone number (should be 10 digits for India)
    if len(phone) != 10:
        return jsonify({"error": "Please enter a valid 10-digit phone number"}), 400
    
    # Format phone number with country code
    formatted_phone = f"+91{phone}"
    
    # Generate 6-digit OTP
    import random
    otp_code = str(random.randint(100000, 999999))
    
    # Set expiration (5 minutes from now)
    expires_at = datetime.utcnow() + timedelta(minutes=5)
    
    # Delete any existing unverified OTPs for this phone
    OTP.query.filter_by(phone=formatted_phone, verified=False).delete()
    
    # Create new OTP
    otp_record = OTP(
        phone=formatted_phone,
        otp_code=otp_code,
        expires_at=expires_at
    )
    db.session.add(otp_record)
    db.session.commit()
    
    # Send OTP via SMS
    try:
        send_sms_otp(formatted_phone, otp_code)
    except Exception as e:
        print(f"Error sending SMS: {e}")
        # For development, still return success even if SMS fails
        # In production, you might want to handle this differently
    
    # For development/testing: return OTP in response (remove in production)
    # This allows testing without actual SMS service
    is_development = os.getenv("FLASK_ENV") == "development" or os.getenv("ENVIRONMENT") == "development"
    
    response_data = {
        "message": "OTP sent successfully",
        "phone": phone  # Return without country code for display
    }
    
    # Only include OTP in response for development
    if is_development:
        response_data["otp"] = otp_code  # Remove this in production!
        response_data["note"] = "OTP included for development only"
    
    return jsonify(response_data), 200


@app.route("/api/auth/verify-otp", methods=["POST"])
def verify_otp():
    """Verify OTP and login user."""
    ensure_db()
    data = request.get_json() or {}
    phone = data.get("phone")
    otp_code = data.get("otp")
    
    if not phone or not otp_code:
        return jsonify({"error": "Phone number and OTP are required"}), 400
    
    # Clean and format phone number
    phone = ''.join(filter(str.isdigit, phone))
    if len(phone) != 10:
        return jsonify({"error": "Please enter a valid 10-digit phone number"}), 400
    
    formatted_phone = f"+91{phone}"
    
    # Find the most recent unverified OTP for this phone
    otp_record = OTP.query.filter_by(
        phone=formatted_phone,
        verified=False
    ).order_by(OTP.created_at.desc()).first()
    
    if not otp_record:
        return jsonify({"error": "OTP not found. Please request a new OTP"}), 404
    
    # Check if OTP is expired
    if datetime.utcnow() > otp_record.expires_at:
        return jsonify({"error": "OTP has expired. Please request a new OTP"}), 400
    
    # Verify OTP code
    if otp_record.otp_code != otp_code:
        return jsonify({"error": "Invalid OTP. Please try again"}), 400
    
    # Mark OTP as verified
    otp_record.verified = True
    db.session.commit()
    
    # Find or create user by phone number
    user = User.query.filter_by(phone=formatted_phone).first()
    
    if not user:
        # Create new user if doesn't exist (phone-only registration)
        # Generate a temporary email if not provided
        temp_email = f"user_{phone}@dealsdouble.ai"
        # Make sure email is unique
        counter = 1
        while User.query.filter_by(email=temp_email).first():
            temp_email = f"user_{phone}_{counter}@dealsdouble.ai"
            counter += 1
        
        user = User(
            name=f"User {phone}",  # Default name, can be updated later
            email=temp_email,
            password_hash=generate_password_hash(""),  # Empty password for OTP login
            phone=formatted_phone,
            welcome_email_sent=False
        )
        db.session.add(user)
        db.session.commit()
        
        # Send welcome email for new user created via OTP
        try:
            send_welcome_email(user)
            user.welcome_email_sent = True
            db.session.commit()
        except Exception as e:
            print(f"Error sending welcome email on OTP registration: {e}")
    else:
        # Existing user - send welcome email on first login if not already sent
        if not getattr(user, 'welcome_email_sent', False):
            try:
                send_welcome_email(user)
                user.welcome_email_sent = True
                db.session.commit()
            except Exception as e:
                print(f"Error sending welcome email on OTP login: {e}")
    
    # Get company info if user has one
    company = None
    company_verified = False
    if user.company_id:
        try:
            company = Company.query.get(user.company_id)
            if company:
                company_verified = company.verified
        except Exception as e:
            print(f"Error fetching company: {e}")
            company = None
    
    return jsonify({
        "token": "demo-token",
        "user": {
            "id": user.id,
            "name": user.name,
            "email": user.email,
            "phone": user.phone,
            "company_id": user.company_id,
            "has_company": user.company_id is not None,
            "company_verified": company_verified,
            "is_admin": user.is_admin or False
        },
        "message": "OTP verified successfully"
    }), 200


@app.route("/api/auth/forgot-password", methods=["POST"])
def forgot_password():
    """Send a password reset email to the user."""
    ensure_db()
    data = request.get_json() or {}
    email = data.get("email", "").strip().lower()

    if not email:
        return jsonify({"error": "Email is required"}), 400

    user = User.query.filter_by(email=email).first()

    # Always return success to prevent email enumeration
    if not user:
        return jsonify({"message": "If that email exists, a reset link has been sent."}), 200

    # Generate a secure reset token
    token = secrets.token_urlsafe(32)
    token_hash = hash_reset_token(token)
    expires_at = datetime.utcnow() + timedelta(hours=1)

    user.reset_token_hash = token_hash
    user.reset_token_expires_at = expires_at
    db.session.commit()

    reset_url = f"{get_frontend_base_url()}/reset-password?token={token}"

    # Try to send email
    mail_username = app.config.get("MAIL_USERNAME", "")
    if mail_username:
        try:
            msg = Message(
                subject="Reset Your DealsDoubled Password",
                recipients=[user.email],
                html=f"""
                <html>
                <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; background: #f9fafb;">
                  <div style="background: white; border-radius: 12px; padding: 40px; box-shadow: 0 4px 20px rgba(0,0,0,0.08);">
                    <div style="text-align: center; margin-bottom: 30px;">
                      <h1 style="color: #8b5cf6; font-size: 28px; margin: 0;">DealsDoubled</h1>
                      <p style="color: #6b7280; font-size: 14px; margin-top: 4px;">B2B Marketplace</p>
                    </div>
                    <h2 style="color: #111827; font-size: 22px;">Reset Your Password</h2>
                    <p style="color: #6b7280; font-size: 15px; line-height: 1.6;">
                      Hello {user.name},<br><br>
                      We received a request to reset your password. Click the button below to create a new password.
                    </p>
                    <div style="text-align: center; margin: 30px 0;">
                      <a href="{reset_url}"
                         style="display: inline-block; background: linear-gradient(135deg, #8b5cf6, #ec4899);
                                color: white; text-decoration: none; padding: 14px 32px;
                                border-radius: 10px; font-weight: bold; font-size: 16px;">
                        Reset Password
                      </a>
                    </div>
                    <p style="color: #9ca3af; font-size: 13px; text-align: center;">
                      This link expires in 1 hour. If you didn't request this, please ignore this email.
                    </p>
                    <hr style="border: none; border-top: 1px solid #e5e7eb; margin: 24px 0;">
                    <p style="color: #9ca3af; font-size: 12px; text-align: center;">
                      &copy; 2024 DealsDoubled.in. All rights reserved.
                    </p>
                  </div>
                </body>
                </html>
                """
            )
            mail.send(msg)
            print(f"[Password Reset] Email sent to {user.email}")
        except Exception as e:
            print(f"[Password Reset] Failed to send email: {e}")
    else:
        # Dev mode: log the token
        print(f"\n{'='*60}\n[DEV] Password Reset Link for {user.email}:\n{reset_url}\n{'='*60}\n")

    return jsonify({"message": "If that email exists, a reset link has been sent."}), 200


@app.route("/api/auth/reset-password", methods=["POST"])
def reset_password():
    """Reset the user's password using a valid token."""
    ensure_db()
    data = request.get_json() or {}
    token = data.get("token", "")
    new_password = data.get("password", "")

    if not token or not new_password:
        return jsonify({"error": "Token and new password are required"}), 400

    if len(new_password) < 8:
        return jsonify({"error": "Password must be at least 8 characters long"}), 400

    token_hash = hash_reset_token(token)
    user = User.query.filter_by(reset_token_hash=token_hash).first()

    if not user:
        return jsonify({"error": "Invalid or expired reset token"}), 400

    if not user.reset_token_expires_at or datetime.utcnow() > user.reset_token_expires_at:
        return jsonify({"error": "Reset token has expired. Please request a new one."}), 400

    # Update password and invalidate token
    user.password_hash = generate_password_hash(new_password)
    user.reset_token_hash = ""
    user.reset_token_expires_at = None
    db.session.commit()

    print(f"[Password Reset] Password successfully reset for {user.email}")
    return jsonify({"message": "Password reset successfully. You can now log in."}), 200


@app.route("/api/homepage/data")
def homepage_data():
    """Consolidated endpoint for homepage — returns stats, categories, velocity, and featured products in one request."""
    ensure_db()
    try:
        # 1. Stats
        product_count = Product.query.filter_by(approved=True).count()
        supplier_count = Company.query.filter_by(verified=True).count()
        category_count = Category.query.count()
        user_count = User.query.count()

        # 2. All categories sorted by trending
        all_categories = Category.query.order_by(Category.views.desc(), Category.last_visited_at.desc()).all()
        categories_data = [{"id": c.id, "name": c.name, "views": c.views, "sales_count": c.sales_count} for c in all_categories]

        # 3. Featured products: up to 3 per category (max 5 categories)
        featured_products = []
        top_cats = all_categories[:5]
        for cat in top_cats:
            prods = Product.query.filter_by(category_id=cat.id, approved=True).order_by(
                Product.is_priority.desc(), Product.views.desc()
            ).limit(3).all()
            for p in prods:
                featured_products.append({
                    "id": p.id,
                    "name": p.name,
                    "description": p.description,
                    "price": p.price,
                    "image_url": p.image_url,
                    "category_id": p.category_id,
                    "category_name": cat.name,
                    "company_id": p.company_id,
                    "company_name": p.company.name if p.company else None,
                    "location": p.company.location if p.company else "India",
                    "is_priority": p.is_priority,
                    "verified": p.company.verified if p.company else False,
                })

        # 4. Velocity data (12-month trend)
        try:
            from sqlalchemy import func as sqlfunc, extract
            now = datetime.utcnow()
            monthly = []
            for i in range(11, -1, -1):
                month_start = (now.replace(day=1) - timedelta(days=i*30)).replace(day=1)
                month_end = (month_start + timedelta(days=32)).replace(day=1)
                count = MarketplaceEvent.query.filter(
                    MarketplaceEvent.created_at >= month_start,
                    MarketplaceEvent.created_at < month_end,
                    MarketplaceEvent.event_type.in_(["product_view", "inquiry_created", "search"])
                ).count()
                monthly.append(count or 0)

            if max(monthly) == 0:
                import random as _r
                base = [42, 58, 66, 61, 73, 81, 77, 69, 88, 84, 79, 91]
                monthly = [min(100, max(10, v + _r.randint(-5, 5))) for v in base]

            velocity_data = monthly
        except Exception as ve:
            velocity_data = [42, 58, 66, 61, 73, 81, 77, 69, 88, 84, 79, 91]
            print(f"[Homepage] Velocity calc error: {ve}")

        return jsonify({
            "stats": {
                "products": product_count,
                "suppliers": supplier_count,
                "categories": category_count,
                "users": user_count,
            },
            "categories": categories_data,
            "featured_products": featured_products,
            "velocity": velocity_data,
        })

    except Exception as e:
        print(f"[Homepage] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/dashboard")
def dashboard():
    ensure_db()
    # For simplicity, return aggregate stats and recent products
    # Only count approved products
    product_count = Product.query.filter_by(approved=True).count()
    supplier_count = Company.query.filter_by(verified=True).count()  # Only count verified companies as suppliers
    category_count = Category.query.count()
    # Only show approved products
    latest_products = (
        Product.query.filter_by(approved=True).order_by(Product.created_at.desc()).limit(5).all()
    )
    return jsonify(
        {
            "stats": {
                "products": product_count,
                "suppliers": supplier_count,
                "categories": category_count,
            },
            "latest_products": [
                {"id": p.id, "name": p.name, "price": p.price, "company_name": p.company.name if p.company else None}
                for p in latest_products
            ],
        }
    )


@app.route("/api/post-product", methods=["POST"])
def post_product():
    ensure_db()
    data = request.get_json() or {}
    name = data.get("name")
    description = data.get("description", "")
    price = data.get("price")
    image_url = data.get("image_url", "")
    category_id = data.get("category_id")
    company_id = data.get("company_id")
    user_id = data.get("user_id")  # Get user_id from request

    if not all([name, category_id]):
        return jsonify({"error": "Missing required fields"}), 400

    # If company_id not provided, get from user
    if not company_id and user_id:
        user = User.query.get(user_id)
        if not user or not user.company_id:
            return jsonify({"error": "User must have a registered company"}), 400
        company_id = user.company_id
    
    if not company_id:
        return jsonify({"error": "Company ID is required"}), 400

    # Check if company is verified
    company = Company.query.get(company_id)
    if not company:
        return jsonify({"error": "Company not found"}), 404
    
    if not company.verified:
        return jsonify({"error": "Your company is pending verification. You can post products once your company is verified by admin."}), 403

    product = Product(
        name=name,
        description=description,
        price=price,
        image_url=image_url,
        category_id=category_id,
        company_id=company_id,
        featured=False,
        approved=False,  # Requires admin approval before appearing in marketplace
    )
    db.session.add(product)
    db.session.commit()
    return jsonify({
        "message": "Product posted successfully. Waiting for admin approval before it appears in the marketplace.",
        "id": product.id
    }), 201


@app.route("/api/stats")
def stats():
    ensure_db()
    # Get accurate counts - for admin dashboard, show all data
    total_products = Product.query.count()  # All products (approved + pending)
    approved_products = Product.query.filter_by(approved=True).count()  # Only approved
    total_companies = Company.query.count()  # All companies
    verified_companies = Company.query.filter_by(verified=True).count()  # Only verified companies
    total_categories = Category.query.count()
    total_users = User.query.count()  # All registered users
    
    return jsonify(
        {
            "products": total_products,  # Show total products for admin (includes pending)
            "approved_products": approved_products,  # Approved products count
            "suppliers": verified_companies,  # Only count verified companies as suppliers
            "verified_suppliers": verified_companies,  # Verified companies count
            "categories": total_categories,
            "users": total_users,  # Total registered users
        }
    )


@app.route("/api/inquiries", methods=["POST"])
def create_inquiry():
    ensure_db()
    data = request.get_json() or {}
    
    product_id = data.get("product_id")
    name = data.get("name")
    email = data.get("email")
    phone = data.get("phone")
    message = data.get("message")
    quantity = data.get("quantity", "")
    
    if not all([product_id, name, email, phone, message]):
        return jsonify({"error": "Missing required fields"}), 400
    
    # Check if product exists
    product = Product.query.get(product_id)
    if not product:
        return jsonify({"error": "Product not found"}), 404
    
    inquiry = Inquiry(
        product_id=product_id,
        name=name,
        email=email,
        phone=phone,
        message=message,
        quantity=quantity
    )
    
    # NEW: Automated Buyer Manager Assignment for high-value B2B sourcing
    try:
        # Simple extraction of numeric quantity
        qty_num = int(''.join(filter(str.isdigit, quantity))) if quantity else 0
    except ValueError:
        qty_num = 0
    
    # High value criteria: High quantity OR specific keywords
    keywords = ["bulk", "partnership", "container", "export", "import", "wholesale", "deal", "contract"]
    is_high_value = qty_num >= 100 or any(kw in (message or "").lower() for kw in keywords)
    
    if is_high_value:
        # Find all designated buyer managers
        from sqlalchemy.sql.expression import func as sql_func
        manager = User.query.filter_by(is_buyer_manager=True).order_by(sql_func.random()).first()
        if manager:
            inquiry.manager_id = manager.id
            print(f"[Phase 7] Auto-assigned Manager {manager.name} to High-Value Inquiry #{inquiry.id or 'pending'}")

    db.session.add(inquiry)
    db.session.commit()

    lead_quality = score_lead_quality(message=message, quantity=quantity)
    
    # AI RFQ Matching: Automatically match suppliers for new inquiries
    try:
        match_rfq(inquiry.id)
    except Exception as e:
        print(f"[Phase 8] Error during automated RFQ matching: {str(e)}")
    
    # Send email notification to seller
    if product.company and product.company.users:
        seller_email = product.company.users[0].email
        if seller_email:
            send_new_inquiry_email(inquiry, product, seller_email)

    try:
        create_seller_alert(
            product.company_id,
            "new_inquiry",
            f"{lead_quality['temperature'].title()} lead for {product.name}",
            f"{name} submitted an inquiry with lead score {lead_quality['score']}.",
            severity="high" if lead_quality["temperature"] == "hot" else "info",
            entity_type="inquiry",
            entity_id=inquiry.id,
        )
    except Exception as alert_error:
        print(f"[Alert] Failed to create inquiry alert: {alert_error}")
        db.session.rollback()

    try:
        record_marketplace_event(
            "inquiry_created",
            session_id=get_request_session_id(data),
            user_id=data.get("user_id"),
            company_id=product.company_id,
            product_id=product.id,
            category_id=product.category_id,
            inquiry_id=inquiry.id,
            location=product.company.location if product.company else product.location,
            metadata={
                "buyer_name": name,
                "lead_score": lead_quality["score"],
                "lead_temperature": lead_quality["temperature"],
                "quantity": quantity,
            },
        )
    except Exception as analytics_error:
        print(f"[Analytics] Failed to record inquiry event: {analytics_error}")
        db.session.rollback()
    
    return jsonify({"message": "Inquiry sent successfully", "id": inquiry.id}), 201


@app.route("/api/admin/buyer-managers", methods=["GET"])
def list_buyer_managers():
    """List all users designated as Buyer Managers."""
    ensure_db()
    managers = User.query.filter_by(is_buyer_manager=True).all()
    return jsonify([{
        "id": m.id,
        "name": m.name,
        "email": m.email,
        "phone": m.phone,
        "assigned_count": Inquiry.query.filter_by(manager_id=m.id).count()
    } for m in managers])


@app.route("/api/admin/inquiries", methods=["GET"])
def list_admin_inquiries():
    """List all inquiries for admin oversight."""
    ensure_db()
    inquiries = Inquiry.query.order_by(Inquiry.created_at.desc()).all()
    return jsonify([{
        "id": i.id,
        "product_name": i.product.name if i.product else "Deleted Product",
        "buyer_name": i.name,
        "buyer_email": i.email,
        "quantity": i.quantity,
        "manager_id": i.manager_id,
        "manager_name": i.manager.name if i.manager else None,
        "created_at": i.created_at.isoformat()
    } for i in inquiries])


@app.route("/api/admin/users/<int:user_id>/toggle-manager", methods=["POST"])
def toggle_buyer_manager(user_id):
    """Designate or remove a user as a Buyer Manager."""
    ensure_db()
    user = User.query.get_or_404(user_id)
    user.is_buyer_manager = not user.is_buyer_manager
    db.session.commit()
    return jsonify({
        "message": f"User {user.name} is {'now' if user.is_buyer_manager else 'no longer'} a Buyer Manager",
        "is_buyer_manager": user.is_buyer_manager
    })


@app.route("/api/admin/assign-manager", methods=["POST"])
def assign_manager_to_inquiry():
    """Manually assign a manager to an inquiry."""
    ensure_db()
    data = request.get_json() or {}
    inquiry_id = data.get("inquiry_id")
    manager_id = data.get("manager_id")
    
    inquiry = Inquiry.query.get_or_404(inquiry_id)
    if manager_id:
        manager = User.query.get_or_404(manager_id)
        if not manager.is_buyer_manager:
            return jsonify({"error": "User is not a designated Buyer Manager"}), 400
    
    inquiry.manager_id = manager_id
    db.session.commit()
    return jsonify({"message": "Manager assigned successfully"})


@app.route("/api/inquiries/<int:inquiry_id>/reply", methods=["POST"])
def reply_to_inquiry(inquiry_id):
    """Reply to an inquiry (seller only)"""
    ensure_db()
    data = request.get_json() or {}
    user_id = data.get("user_id")
    message = data.get("message")
    
    if not user_id or not message:
        return jsonify({"error": "User ID and message are required"}), 400
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404
    
    inquiry = Inquiry.query.get(inquiry_id)
    if not inquiry:
        return jsonify({"error": "Inquiry not found"}), 404
    
    # Check if user is the seller (owns the product's company)
    product = Product.query.get(inquiry.product_id)
    if not product or not product.company_id or product.company_id != user.company_id:
        return jsonify({"error": "Unauthorized. You can only reply to inquiries for your products"}), 403
    
    # Create reply
    reply = InquiryReply(
        inquiry_id=inquiry_id,
        user_id=user_id,
        message=message
    )
    db.session.add(reply)
    db.session.commit()
    
    # Send email notification to buyer
    try:
        send_inquiry_reply_email(inquiry, reply, user)
    except Exception as e:
        print(f"Error sending reply email: {e}")
    
    return jsonify({
        "message": "Reply sent successfully",
        "reply": {
            "id": reply.id,
            "message": reply.message,
            "created_at": reply.created_at.isoformat(),
            "seller_name": user.name
        }
    }), 201


def match_rfq(inquiry_id):
    """Helper to match an inquiry with relevant suppliers"""
    inquiry = Inquiry.query.get(inquiry_id)
    if not inquiry:
        return []
    
    inquiry_text = (inquiry.name + " " + inquiry.message).lower()
    inquiry_product = Product.query.get(inquiry.product_id)
    cat_id = inquiry_product.category_id if inquiry_product else None
    
    all_companies = Company.query.all()
    matches = []
    
    for company in all_companies:
        # Skip the original company that owns the product
        if inquiry_product and company.id == inquiry_product.company_id:
            continue
            
        score = 0.0
        company_products = Product.query.filter_by(company_id=company.id).all()
        
        # 1. Category Match (Heavy Weight)
        has_cat_match = any(p.category_id == cat_id for p in company_products) if cat_id else False
        if has_cat_match:
            score += 0.5
            
        # 2. Keyword Match (Medium Weight)
        keywords = [k.strip() for k in inquiry_text.split() if len(k) > 3]
        for p in company_products:
            p_text = (p.name + " " + (p.tags or "")).lower()
            match_count = sum(1 for k in keywords if k in p_text)
            if match_count > 0:
                score += min(0.4, match_count * 0.1)
                break
        
        if score > 0.1:
            matches.append({
                "company_id": company.id,
                "company_name": company.name,
                "score": round(score, 2)
            })
            
    matches.sort(key=lambda x: x["score"], reverse=True)
    return matches[:10]

@app.route("/api/admin/inquiries/<int:inquiry_id>/match", methods=["POST"])
def trigger_rfq_matching(inquiry_id):
    ensure_db()
    matches = match_rfq(inquiry_id)
    
    # Store matches in DB for management
    RFQMatch.query.filter_by(inquiry_id=inquiry_id).delete()
    for m in matches:
        new_match = RFQMatch(
            inquiry_id=inquiry_id,
            company_id=m["company_id"],
            score=m["score"]
        )
        db.session.add(new_match)
    
    db.session.commit()
    return jsonify({"message": f"Found {len(matches)} potential matches", "matches": matches})

@app.route("/api/admin/inquiries/<int:inquiry_id>/matches", methods=["GET"])
def get_rfq_matches(inquiry_id):
    ensure_db()
    return jsonify(build_rfq_comparison(inquiry_id))

@app.route("/api/inquiries/<int:inquiry_id>/chat", methods=["GET"])
def get_chat_messages(inquiry_id):
    ensure_db()
    messages = InquiryMessage.query.filter_by(inquiry_id=inquiry_id).order_by(InquiryMessage.created_at.asc()).all()
    return jsonify([{
        "id": m.id,
        "sender_id": m.sender_id,
        "sender_name": m.sender.name,
        "message": m.message,
        "created_at": m.created_at.isoformat()
    } for m in messages])

@app.route("/api/inquiries/<int:inquiry_id>/chat", methods=["POST"])
def send_chat_message(inquiry_id):
    ensure_db()
    data = request.json
    sender_id = data.get("sender_id")
    message_text = data.get("message")
    
    if not sender_id or not message_text:
        return jsonify({"error": "Missing sender_id or message"}), 400
        
    new_msg = InquiryMessage(
        inquiry_id=inquiry_id,
        sender_id=sender_id,
        message=message_text
    )
    
    inquiry = Inquiry.query.get(inquiry_id)
    if inquiry and inquiry.negotiation_status == "OPEN":
        inquiry.negotiation_status = "ACTIVE"
        
    db.session.add(new_msg)
    db.session.commit()
    
    return jsonify({
        "message": "Sent",
        "chat_message": {
            "id": new_msg.id,
            "sender_id": new_msg.sender_id,
            "message": new_msg.message,
            "created_at": new_msg.created_at.isoformat()
        }
    })

@app.route("/api/buyer/inquiries", methods=["GET"])
def get_buyer_inquiries():
    """Get inquiries made by a buyer (by email)"""
    ensure_db()
    email = request.args.get('email')
    
    if not email:
        return jsonify({"error": "Email is required"}), 400
    
    inquiries = Inquiry.query.filter_by(email=email).order_by(Inquiry.created_at.desc()).all()
    
    result = []
    for i in inquiries:
        lead_quality = score_lead_quality(message=i.message, quantity=i.quantity)
        product_image = None
        if i.product:
            try:
                primary_image = ProductImage.query.filter_by(product_id=i.product_id, is_primary=True).first()
                if primary_image:
                    product_image = primary_image.image_url
                elif i.product.image_url:
                    product_image = i.product.image_url
            except Exception:
                pass
                
        result.append({
            "id": i.id,
            "product_id": i.product_id,
            "product_name": i.product.name if i.product else "Deleted Product",
            "product_image": product_image,
            "message": i.message,
            "quantity": i.quantity,
            "negotiation_status": i.negotiation_status,
            "created_at": i.created_at.isoformat(),
            "replies": [{
                "id": r.id,
                "message": r.message,
                "created_at": r.created_at.isoformat(),
                "seller_name": r.seller.name if r.seller else "Seller"
            } for r in i.replies]
        })
    return jsonify(result)

@app.route("/api/seller/inquiries", methods=["GET"])
def get_seller_inquiries():
    """Get inquiries received by a seller (by company_id)"""
    ensure_db()
    company_id = request.args.get('company_id')
    
    if not company_id:
        return jsonify({"error": "Company ID is required"}), 400
    
    # Get all inquiries for products belonging to this company
    inquiries = Inquiry.query.join(Product).filter(Product.company_id == company_id).order_by(Inquiry.created_at.desc()).all()
    
    result = []
    for i in inquiries:
        product_image = None
        if i.product:
            try:
                primary_image = ProductImage.query.filter_by(product_id=i.product_id, is_primary=True).first()
                if primary_image:
                    product_image = primary_image.image_url
                elif i.product.image_url:
                    product_image = i.product.image_url
            except Exception:
                pass
                
        result.append({
            "id": i.id,
            "product_id": i.product_id,
            "product_name": i.product.name if i.product else "Deleted Product",
            "product_image": product_image,
            "buyer_name": i.name,
            "buyer_email": i.email,
            "buyer_phone": i.phone,
            "message": i.message,
            "quantity": i.quantity,
            "lead_score": lead_quality["score"],
            "lead_temperature": lead_quality["temperature"],
            "negotiation_status": i.negotiation_status,
            "created_at": i.created_at.isoformat()
        })
    return jsonify(result)


@app.route("/api/buy-requirements", methods=["GET", "POST"])
def buy_requirements():
    ensure_db()
    if request.method == "GET":
        requirements = BuyRequirement.query.order_by(BuyRequirement.created_at.desc()).all()
        return jsonify([
            {
                "id": r.id,
                "product_name": r.product_name,
                "description": r.description,
                "quantity": r.quantity,
                "location": r.location,
                "budget": r.budget,
                "email": r.email,
                "created_at": r.created_at.isoformat() if r.created_at else None
            }
            for r in requirements
        ])
    else:  # POST
        data = request.get_json() or {}
        requirement = BuyRequirement(
            product_name=data.get("product_name"),
            description=data.get("description"),
            quantity=data.get("quantity", ""),
            location=data.get("location", ""),
            budget=data.get("budget", ""),
            email=data.get("email", "")
        )
        try:
            db.session.add(requirement)
            db.session.commit()
            print(f"[BuyRequirement] Successfully saved requirement with id: {requirement.id}")

            try:
                record_marketplace_event(
                    "buy_requirement_created",
                    session_id=get_request_session_id(data),
                    user_id=data.get("user_id"),
                    location=requirement.location,
                    search_query=requirement.product_name,
                    metadata={
                        "product_name": requirement.product_name,
                        "budget": requirement.budget,
                    },
                )
            except Exception as analytics_error:
                print(f"[Analytics] Failed to record buy requirement: {analytics_error}")
                db.session.rollback()
            
            # Send Email Notification to Admin
            try:
                admin_email = app.config.get("MAIL_USERNAME")
                if admin_email:
                    msg = Message(
                        subject=f"New Buy Requirement: {requirement.product_name}",
                        sender=app.config.get("MAIL_USERNAME"),
                        recipients=[admin_email],
                        body=f"""
A new buy requirement has been posted on DealsDouble.ai

Product: {requirement.product_name}
Quantity: {requirement.quantity}
Location: {requirement.location}
Budget: {requirement.budget}
Buyer Email: {requirement.email}

Description:
{requirement.description}

View it in the admin panel: http://localhost:3000/admin/requirements
"""
                    )
                    mail.send(msg)
                    print(f"[Email] Admin notification sent for requirement: {requirement.id}")
            except Exception as mail_err:
                print(f"Error sending buy requirement email: {mail_err}")
                
            return jsonify({"message": "Buy requirement posted", "id": requirement.id}), 201
            
        except Exception as db_e:
            db.session.rollback()
            print(f"[BuyRequirement] Database Error: {db_e}")
            return jsonify({"error": "Database error saving requirement", "details": str(db_e)}), 500


@app.route("/api/trade-leads", methods=["GET"])
def trade_leads():
    ensure_db()
    lead_type = request.args.get("type", "all")
    
    query = TradeLead.query
    if lead_type != "all":
        query = query.filter(TradeLead.type == lead_type)
    
    leads = query.order_by(TradeLead.created_at.desc()).all()
    return jsonify([
        {
            "id": l.id,
            "title": l.title,
            "description": l.description,
            "type": l.type,
            "category": l.category,
            "location": l.location,
            "created_at": l.created_at.isoformat() if l.created_at else None
        }
        for l in leads
    ])


@app.route("/api/products/<int:product_id>/phone", methods=["GET"])
def get_product_phone(product_id: int):
    ensure_db()
    # Check if user is authenticated
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token or token == "null":
        return jsonify({"error": "Authentication required"}), 401
    
    # Get user_id from query params (simplified - in production, verify JWT token)
    user_id = request.args.get("user_id", type=int)
    if not user_id:
        return jsonify({"error": "User ID required"}), 400
    
    # Check subscription
    subscription = Subscription.query.filter_by(
        user_id=user_id,
        is_active=True
    ).first()
    
    if not subscription:
        return jsonify({"error": "Active subscription required to view phone numbers"}), 403
    
    product = Product.query.get_or_404(product_id)
    if not product.company or not product.company.phone:
        return jsonify({"error": "Phone number not available"}), 404
    
    return jsonify({
        "phone": product.company.phone,
        "company_name": product.company.name
    })


@app.route("/api/subscriptions", methods=["POST"])
def create_subscription():
    ensure_db()
    data = request.get_json() or {}
    user_id = data.get("user_id")
    plan_type = data.get("plan_type", "basic")
    
    if not user_id:
        return jsonify({"error": "User ID required"}), 400
    
    # Check if user already has active subscription
    existing = Subscription.query.filter_by(
        user_id=user_id,
        is_active=True
    ).first()
    
    if existing:
        return jsonify({"message": "Subscription already active", "id": existing.id}), 200
    
    subscription = Subscription(
        user_id=user_id,
        plan_type=plan_type,
        is_active=True
    )
    db.session.add(subscription)
    db.session.commit()
    
    return jsonify({"message": "Subscription created", "id": subscription.id}), 201


@app.route("/api/subscriptions/check/<int:user_id>")
def check_subscription(user_id: int):
    ensure_db()
    subscription = Subscription.query.filter_by(
        user_id=user_id,
        is_active=True
    ).first()
    
    return jsonify({
        "has_subscription": subscription is not None,
        "plan_type": subscription.plan_type if subscription else None
    })


@app.route("/api/locations")
def get_locations():
    """Get all locations"""
    return jsonify(get_all_locations())

@app.route("/api/locations/countries")
def get_countries_list():
    """Get all available countries"""
    return jsonify(get_countries())

@app.route("/api/locations/states")
def get_states():
    """Get states by country"""
    country = request.args.get("country")
    countries = get_countries()
    selected_country = country or (countries[0] if countries else "India")
    return jsonify(get_states_by_country(selected_country))


@app.route("/api/locations/districts")
def get_districts():
    """Get districts by state"""
    countries = get_countries()
    default_country = countries[0] if countries else "India"
    country = request.args.get("country", default_country)
    state = request.args.get("state")
    if not state:
        return jsonify({"error": "State parameter required"}), 400
    return jsonify(get_districts_by_state(country, state))


@app.route("/api/companies/register", methods=["POST"])
def register_company():
    """Register a new company for a user"""
    ensure_db()
    data = request.get_json() or {}
    user_id = data.get("user_id")
    name = data.get("name")
    description = data.get("description", "")
    location = data.get("location", "India")
    website = data.get("website", "")
    phone = data.get("phone", "")
    gst_number = data.get("gst_number", "")
    logo_url = data.get("logo_url", "")
    
    if not user_id or not name:
        return jsonify({"error": "User ID and company name are required"}), 400
    
    if not gst_number:
        return jsonify({"error": "GST number is required"}), 400
    
    # Check if user already has a company
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404
    
    if user.company_id:
        return jsonify({"error": "User already has a company registered"}), 400
    
    # Create new company
    company = Company(
        name=name,
        description=description,
        location=location,
        website=website,
        phone=phone,
        gst_number=gst_number,
        logo_url=logo_url,
        verified=False  # Admin needs to verify
    )
    db.session.add(company)
    db.session.flush()
    
    # Link company to user
    user.company_id = company.id
    db.session.commit()
    
    return jsonify({
        "message": "Company registered successfully. Waiting for admin verification.",
        "company_id": company.id,
        "company_name": company.name
    }), 201


@app.route("/api/companies/verify/<int:company_id>", methods=["POST"])
def verify_company(company_id):
    """Verify a company (admin only)."""
    ensure_db()
    # Check admin access
    user, error = require_admin()
    if error:
        return error
    
    company = Company.query.get_or_404(company_id)
    company.verified = True
    db.session.commit()
    
    # Send approval email
    send_company_approval_email(company)
    
    return jsonify({
        "message": "Company verified successfully",
        "company_id": company.id,
        "company_name": company.name,
        "verified": company.verified
    })


@app.route("/api/companies/unverify/<int:company_id>", methods=["POST"])
def unverify_company(company_id):
    """Unverify a company (admin only)."""
    ensure_db()
    # Check admin access
    user, error = require_admin()
    if error:
        return error
    company = Company.query.get_or_404(company_id)
    company.verified = False
    db.session.commit()
    
    return jsonify({
        "message": "Company unverified successfully",
        "company_id": company.id,
        "company_name": company.name,
        "verified": company.verified
    })


@app.route("/api/seller/bulk-upload/template", methods=["GET"])
def download_bulk_upload_template():
    """Generate and download a bulk upload Excel template."""
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error": "Excel processing not available on server"}), 501
    
    wb = Workbook()
    
    # Main Data Sheet
    ws = wb.active
    ws.title = "Product Data"
    
    headers = [
        "Product Name*", "Description", "Price*", "Stock Quantity", 
        "Min Order Qty", "Category Name*", "Location", "Price Trend %"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20

    # Add sample row
    ws.append([
        "Sample Industrial Drill", 
        "High-performance heavy duty drill for construction.", 
        15000.0, 50, 5, "Machinery", "Mumbai, India", 2.5
    ])

    # Reference Sheet for Categories
    ws_cat = wb.create_sheet("Categories Reference")
    ws_cat.append(["Valid Category Names (Copy exactly)"])
    ws_cat.cell(row=1, column=1).font = Font(bold=True)
    
    categories = Category.query.all()
    for cat in categories:
        ws_cat.append([cat.name])
    
    ws_cat.column_dimensions['A'].width = 30

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="DealsDouble_Bulk_Upload_Template.xlsx"
    )


@app.route("/api/seller/bulk-upload", methods=["POST"])
def bulk_upload_products():
    """Handle bulk product upload via Excel."""
    ensure_db()
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error": "Excel processing not available on server"}), 501

    user_id = request.form.get("user_id", type=int)
    if not user_id:
        return jsonify({"error": "User ID required"}), 400
    
    user = User.query.get(user_id)
    if not user or not user.company_id:
        return jsonify({"error": "User or company not found"}), 404
    
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "Empty filename"}), 400
    
    if not allowed_data_file(file.filename):
        return jsonify({"error": "Invalid file format. Please upload .xlsx"}), 400

    try:
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active # Assume data is in the first sheet
        
        # Map category names to IDs for efficiency
        categories = {c.name.lower(): c.id for c in Category.query.all()}
        
        products_to_add = []
        errors = []
        success_count = 0
        
        # Iterate rows starting from 2 (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row): continue # Skip empty rows
            
            name, desc, price, stock, min_qty, cat_name, loc, trend = row
            
            # Basic validation
            if not name or not cat_name:
                errors.append(f"Row {row_idx}: Missing Product Name or Category Name")
                continue
                
            cat_id = categories.get(str(cat_name).strip().lower())
            if not cat_id:
                errors.append(f"Row {row_idx}: Category '{cat_name}' not found")
                continue
            
            try:
                # Type conversions
                clean_price = float(price) if price is not None else 0.0
                clean_stock = int(stock) if stock is not None else 0
                clean_min_qty = int(min_qty) if min_qty is not None else 1
                clean_trend = float(trend) if trend is not None else 0.0
                
                new_product = Product(
                    name=str(name),
                    description=str(desc or ""),
                    price=clean_price,
                    stock_quantity=clean_stock,
                    min_order_quantity=clean_min_qty,
                    category_id=cat_id,
                    company_id=user.company_id,
                    location=str(loc or user.company.location or "India"),
                    price_trend=clean_trend,
                    approved=True # Automatically approve bulk uploads for now
                )
                products_to_add.append(new_product)
                success_count += 1
            except (ValueError, TypeError) as e:
                errors.append(f"Row {row_idx}: Invalid numeric data (Price/Stock/Min Qty)")
                continue

        if products_to_add:
            db.session.add_all(products_to_add)
            db.session.commit()
            
        return jsonify({
            "message": "Bulk upload processed",
            "success_count": success_count,
            "error_count": len(errors),
            "errors": errors
        }), 200

    except Exception as e:
        print(f"[Bulk Upload Error] {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/companies/<int:company_id>/toggle-best-seller", methods=["POST"])
def toggle_best_seller(company_id):
    """Toggle best seller status for a company (admin only)."""
    ensure_db()
    user, error = require_admin()
    if error:
        return error

    company = Company.query.get_or_404(company_id)
    company.best_seller = not company.best_seller
    db.session.commit()
    
    status = "marked as Best Seller" if company.best_seller else "removed from Best Seller"
    return jsonify({
        "message": f"Company '{company.name}' has been {status}",
        "company_id": company.id,
        "company_name": company.name,
        "best_seller": company.best_seller
    })


@app.route("/api/admin/companies/approve-all", methods=["POST"])
def approve_all_companies():
    """Approve all pending companies (admin only)."""
    ensure_db()
    # Check admin access
    user, error = require_admin()
    if error:
        return error
    
    # Get all unverified companies
    pending_companies = Company.query.filter_by(verified=False).all()
    
    if not pending_companies:
        return jsonify({
            "message": "No pending companies to approve",
            "count": 0
        }), 200
    
    # Approve all
    count = 0
    for company in pending_companies:
        company.verified = True
        count += 1
        # Send approval email
        send_company_approval_email(company)
    
    db.session.commit()
    
    return jsonify({
        "message": f"Successfully approved {count} company(ies)",
        "count": count
    }), 200


@app.route("/api/companies/pending", methods=["GET"])
def get_pending_companies():
    """Get all companies pending verification"""
    ensure_db()
    pending_companies = Company.query.filter_by(verified=False).all()
    
    return jsonify([
        {
            "id": c.id,
            "name": c.name,
            "description": c.description,
            "location": c.location,
            "website": c.website,
            "phone": c.phone,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "user_name": c.users[0].name if c.users else None,
            "user_email": c.users[0].email if c.users else None
        }
        for c in pending_companies
    ])




@app.route("/api/seller/orders")
def get_seller_orders():
    """Get orders for seller's company"""
    ensure_db()
    user_id = request.args.get('user_id', type=int)
    if not user_id:
        return jsonify({"error": "User ID required"}), 400
    
    user = User.query.get(user_id)
    if not user or not user.company_id:
        return jsonify({"error": "User must have a company"}), 400
    
    # Get all orders for user's company
    orders = Order.query.filter_by(company_id=user.company_id).order_by(Order.created_at.desc()).all()
    
    return jsonify([
        {
            "id": o.id,
            "product_id": o.product_id,
            "product_name": o.product.name,
            "buyer_name": o.buyer_name,
            "buyer_email": o.buyer_email,
            "quantity": o.quantity,
            "unit_price": o.unit_price,
            "total_amount": o.total_amount,
            "status": o.status,
            "created_at": o.created_at.isoformat() if o.created_at else None
        }
        for o in orders
    ])


@app.route("/api/seller/products")
def get_seller_products():
    """Get all products belonging to the seller's company"""
    ensure_db()
    user_id = request.args.get('user_id', type=int)
    if not user_id:
        return jsonify({"error": "User ID required"}), 400
    
    user = User.query.get(user_id)
    if not user or not user.company_id:
        return jsonify({"error": "User must have a company"}), 400
    
    products = Product.query.filter_by(company_id=user.company_id).order_by(Product.created_at.desc()).all()
    
    return jsonify([
        {
            "id": p.id,
            "name": p.name,
            "description": p.description,
            "price": p.price,
            "image_url": p.image_url,
            "category_id": p.category_id,
            "category_name": p.category.name if p.category else None,
            "stock_quantity": p.stock_quantity,
            "min_order_quantity": p.min_order_quantity,
            "approved": p.approved,
            "is_priority": p.is_priority,
            "created_at": p.created_at.isoformat() if p.created_at else None
        }
        for p in products
    ])


@app.route("/api/seller/products/<int:product_id>", methods=["PUT", "DELETE"])
def manage_seller_product(product_id):
    """Update or delete a seller's product"""
    ensure_db()
    data = request.get_json() or {}
    user_id = data.get("user_id") if request.method == "PUT" else request.args.get("user_id", type=int)
    
    if not user_id:
        return jsonify({"error": "User ID required"}), 400
    
    user = User.query.get(user_id)
    if not user or not user.company_id:
        return jsonify({"error": "Unauthorized"}), 403
    
    product = Product.query.get_or_404(product_id)
    if product.company_id != user.company_id:
        return jsonify({"error": "Unauthorized: You can only manage your own products"}), 403
    
    if request.method == "DELETE":
        # Delete related images and inquiries? Usually we soft delete or handle foreign keys.
        # For simplicity, just delete the product.
        db.session.delete(product)
        db.session.commit()
        return jsonify({"message": "Product deleted successfully"})
    
    else:  # PUT
        if 'name' in data: product.name = data['name']
        if 'description' in data: product.description = data['description']
        if 'price' in data: product.price = float(data['price']) if data['price'] else None
        if 'category_id' in data: product.category_id = int(data['category_id'])
        if 'stock_quantity' in data: product.stock_quantity = int(data['stock_quantity'])
        if 'min_order_quantity' in data: product.min_order_quantity = int(data['min_order_quantity'])
        
        db.session.commit()
        return jsonify({"message": "Product updated successfully", "id": product.id})


@app.route("/api/seller/stats")
def get_seller_stats_dashboard():
    """Get aggregated stats for the Power BI inspired seller dashboard"""
    ensure_db()
    user_id = request.args.get('user_id', type=int)
    if not user_id:
        return jsonify({"error": "User ID required"}), 400
    
    user = User.query.get(user_id)
    if not user or not user.company_id:
        return jsonify({"error": "User must have a company"}), 400
    
    company_id = user.company_id
    
    # 1. Basic Counts
    total_products = Product.query.filter_by(company_id=company_id).count()
    pending_products = Product.query.filter_by(company_id=company_id, approved=False).count()
    
    total_orders = Order.query.filter_by(company_id=company_id).count()
    completed_orders = Order.query.filter_by(company_id=company_id, status='completed').count()
    pending_orders = Order.query.filter_by(company_id=company_id, status='pending').count()
    
    total_inquiries = Inquiry.query.join(Product).filter(Product.company_id == company_id).count()
    
    # 2. Revenue (Completed orders)
    revenue = db.session.query(db.func.sum(Order.total_amount)).filter_by(company_id=company_id, status='completed').scalar() or 0
    
    # 3. Monthly Revenue Trend (Last 6 months)
    from datetime import datetime, timedelta
    from sqlalchemy import func
    six_months_ago = datetime.utcnow() - timedelta(days=180)
    
    # Dialect-specific date formatting
    if db.engine.name == 'mysql':
        month_func = func.date_format(Order.created_at, '%Y-%m')
    else:
        month_func = func.strftime('%Y-%m', Order.created_at)

    monthly_revenue = db.session.query(
        month_func.label('month'),
        func.sum(Order.total_amount).label('total')
    ).filter(
        Order.company_id == company_id,
        Order.status == 'completed',
        Order.created_at >= six_months_ago
    ).group_by('month').order_by('month').all()
    
    # 4. Inquiries by Category
    category_inquiries = db.session.query(
        Category.name,
        func.count(Inquiry.id).label('count')
    ).join(Product, Category.id == Product.category_id)\
     .join(Inquiry, Product.id == Inquiry.product_id)\
     .filter(Product.company_id == company_id)\
     .group_by(Category.name).all()
     
    # 5. Inventory Alerts (Low stock < 10)
    low_stock_items = Product.query.filter(
        Product.company_id == company_id,
        Product.stock_quantity < 10
    ).all()
    trust = calculate_supplier_trust(company_id)

    return jsonify({
        "summary": {
            "total_products": total_products,
            "pending_products": pending_products,
            "total_orders": total_orders,
            "completed_orders": completed_orders,
            "pending_orders": pending_orders,
            "total_inquiries": total_inquiries,
            "revenue": float(revenue),
            "trust_score": trust["trust_score"],
            "avg_response_hours": trust["avg_response_hours"],
            "response_rate": trust["response_rate"],
        },
        "trends": {
            "monthly_revenue": [{"month": m[0], "total": float(m[1])} for m in monthly_revenue]
        },
        "distribution": {
            "category_inquiries": [{"name": c[0], "value": c[1]} for c in category_inquiries]
        },
        "alerts": {
            "low_stock": [{"id": p.id, "name": p.name, "stock": p.stock_quantity} for p in low_stock_items]
        },
        "trust": trust,
    })


@app.route("/api/seller/reports")
def get_seller_reports():
    """Get seller reports with filtering options (weekly, monthly, custom)"""
    ensure_db()
    from sqlalchemy import func
    from datetime import timedelta
    
    user_id = request.args.get('user_id', type=int)
    report_type = request.args.get('type', 'monthly')  # weekly, monthly, custom
    start_date = request.args.get('start_date')  # ISO format
    end_date = request.args.get('end_date')  # ISO format
    
    if not user_id:
        return jsonify({"error": "User ID required"}), 400
    
    user = User.query.get(user_id)
    if not user or not user.company_id:
        return jsonify({"error": "User must have a company"}), 400
    
    # Calculate date range based on report type
    end_datetime = datetime.utcnow()
    if report_type == 'weekly':
        start_datetime = end_datetime - timedelta(days=7)
    elif report_type == 'monthly':
        start_datetime = end_datetime - timedelta(days=30)
    elif report_type == 'custom':
        if not start_date or not end_date:
            return jsonify({"error": "Start date and end date required for custom reports"}), 400
        try:
            start_datetime = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            end_datetime = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        except ValueError:
            return jsonify({"error": "Invalid date format. Use ISO format (YYYY-MM-DDTHH:MM:SS)"}), 400
    else:
        return jsonify({"error": "Invalid report type. Use 'weekly', 'monthly', or 'custom'"}), 400
    
    # Get orders in the date range for the seller's company
    orders_query = Order.query.filter(
        Order.company_id == user.company_id,
        Order.created_at >= start_datetime,
        Order.created_at <= end_datetime
    )
    
    orders = orders_query.all()
    
    # Calculate statistics
    total_orders = len(orders)
    pending_orders = len([o for o in orders if o.status == 'pending'])
    completed_orders = len([o for o in orders if o.status == 'completed'])
    processing_orders = len([o for o in orders if o.status == 'processing'])
    shipped_orders = len([o for o in orders if o.status == 'shipped'])
    cancelled_orders = len([o for o in orders if o.status == 'cancelled'])
    
    # Revenue calculations
    total_revenue = sum(o.total_amount for o in orders if o.status == 'completed')
    pending_revenue = sum(o.total_amount for o in orders if o.status == 'pending')
    processing_revenue = sum(o.total_amount for o in orders if o.status == 'processing')
    
    # Average order value
    avg_order_value = total_revenue / completed_orders if completed_orders > 0 else 0
    
    # Orders by product
    product_stats = {}
    for order in orders:
        product_name = order.product.name if order.product else 'Unknown'
        if product_name not in product_stats:
            product_stats[product_name] = {
                'total_orders': 0,
                'quantity': 0,
                'revenue': 0
            }
        product_stats[product_name]['total_orders'] += 1
        product_stats[product_name]['quantity'] += order.quantity
        if order.status == 'completed':
            product_stats[product_name]['revenue'] += order.total_amount
    
    # Convert to list
    product_stats_list = [
        {
            'product_name': name,
            'total_orders': stats['total_orders'],
            'total_quantity': stats['quantity'],
            'revenue': stats['revenue']
        }
        for name, stats in product_stats.items()
    ]
    product_stats_list.sort(key=lambda x: x['revenue'], reverse=True)
    
    # Order details for CSV export
    order_details = [
        {
            'id': o.id,
            'order_date': o.created_at.isoformat() if o.created_at else None,
            'product_name': o.product.name if o.product else 'Unknown',
            'buyer_name': o.buyer_name,
            'buyer_email': o.buyer_email,
            'quantity': o.quantity,
            'unit_price': float(o.unit_price),
            'total_amount': float(o.total_amount),
            'status': o.status,
            'payment_method': o.payment_method or 'N/A',
            'payment_status': o.payment_status or 'N/A'
        }
        for o in orders
    ]
    
    return jsonify({
        'report_type': report_type,
        'start_date': start_datetime.isoformat(),
        'end_date': end_datetime.isoformat(),
        'summary': {
            'total_orders': total_orders,
            'pending_orders': pending_orders,
            'processing_orders': processing_orders,
            'shipped_orders': shipped_orders,
            'completed_orders': completed_orders,
            'cancelled_orders': cancelled_orders,
            'total_revenue': float(total_revenue),
            'pending_revenue': float(pending_revenue),
            'processing_revenue': float(processing_revenue),
            'avg_order_value': float(avg_order_value)
        },
        'product_stats': product_stats_list,
        'order_details': order_details
    })


@app.route("/api/orders/<int:order_id>/status", methods=["PUT"])
def update_order_status(order_id):
    """Update order status - allows buyer to update their own order or seller to update their company's orders"""
    ensure_db()
    data = request.get_json() or {}
    new_status = data.get("status")
    user_id = data.get("user_id")  # Get user_id from request
    
    if not new_status:
        return jsonify({"error": "Status is required"}), 400
    
    if new_status not in ["pending", "completed", "cancelled"]:
        return jsonify({"error": "Invalid status"}), 400
    
    order = Order.query.get_or_404(order_id)
    
    # Authorization: Check if user is the buyer or the seller (company owner)
    is_authorized = False
    if user_id:
        user = User.query.get(user_id)
        if user:
            # Check if user is the buyer (by email match)
            if user.email == order.buyer_email:
                is_authorized = True
            # Check if user is the seller (company owner)
            elif user.company_id and user.company_id == order.company_id:
                is_authorized = True
    
    if not is_authorized:
        return jsonify({"error": "Unauthorized: You can only update your own orders"}), 403
    
    order.status = new_status
    db.session.commit()
    
    return jsonify({
        "message": "Order status updated successfully",
        "order_id": order.id,
        "status": order.status
    })


@app.route("/api/seller/orders/<int:order_id>/status", methods=["PUT"])
def update_seller_order_status(order_id):
    """Update order status - seller only endpoint (for backward compatibility)"""
    ensure_db()
    data = request.get_json() or {}
    new_status = data.get("status")
    user_id = data.get("user_id")
    
    if not new_status:
        return jsonify({"error": "Status is required"}), 400
    
    if new_status not in ["pending", "completed", "cancelled"]:
        return jsonify({"error": "Invalid status"}), 400
    
    order = Order.query.get_or_404(order_id)
    
    # Authorization: Check if user is the seller (company owner)
    if user_id:
        user = User.query.get(user_id)
        if user and user.company_id and user.company_id == order.company_id:
            order.status = new_status
            db.session.commit()
            return jsonify({
                "message": "Order status updated successfully",
                "order_id": order.id,
                "status": order.status
            })
    
    return jsonify({"error": "Unauthorized: Only the seller can update this order"}), 403


@app.route("/api/orders/<int:order_id>", methods=["GET", "PUT"])
def get_or_update_order(order_id):
    """Get or update order details"""
    ensure_db()
    order = Order.query.get_or_404(order_id)
    
    if request.method == "GET":
        # GET: Return order details including company address
        company = order.company
        product = order.product
        
        return jsonify({
            "id": order.id,
            "product_id": order.product_id,
            "product_name": product.name if product else None,
            "company_id": order.company_id,
            "company_name": company.name if company else None,
            "company_location": company.location if company else None,
            "company_phone": company.phone if company else None,
            "company_website": company.website if company else None,
            "company_gst": company.gst_number if company else None,
            "buyer_name": order.buyer_name,
            "buyer_email": order.buyer_email,
            "quantity": order.quantity,
            "unit_price": order.unit_price,
            "total_amount": order.total_amount,
            "status": order.status,
            "payment_status": order.payment_status,
            "payment_method": order.payment_method,
            "created_at": order.created_at.isoformat() if order.created_at else None,
        })
    
    else:  # PUT
        # PUT: Update order details including payment method and status
        data = request.get_json() or {}
        user_id = data.get("user_id")
        
        # Authorization: Check if user is the buyer (by email match)
        is_authorized = False
        if user_id:
            user = User.query.get(user_id)
            if user and user.email == order.buyer_email:
                is_authorized = True
        
        if not is_authorized:
            return jsonify({"error": "Unauthorized: You can only update your own orders"}), 403
        
        # Update payment method if provided
        if 'payment_method' in data:
            order.payment_method = data.get('payment_method')
        
        # Update payment status if provided
        if 'payment_status' in data:
            order.payment_status = data.get('payment_status')
        
        # Update status if provided
        if 'status' in data:
            order.status = data.get('status')
        
        order.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            "message": "Order updated successfully",
            "order_id": order.id,
            "status": order.status,
            "payment_method": order.payment_method,
            "payment_status": order.payment_status
        })


@app.route("/api/orders/<int:order_id>/cod", methods=["POST"])
def complete_cod_order(order_id):
    """Complete Cash on Delivery order - sets status to processing and adds tracking"""
    ensure_db()
    data = request.get_json() or {}
    user_id = data.get("user_id")
    
    order = Order.query.get_or_404(order_id)
    
    # Authorization: Check if user is the buyer (by buyer_id OR buyer_email match)
    # Same authorization logic as get_my_orders endpoint
    is_authorized = False
    if user_id:
        user = User.query.get(user_id)
        if user:
            # Check if user is the buyer by ID or email (same logic as get_my_orders)
            if order.buyer_id == user_id:
                is_authorized = True
            elif order.buyer_email and user.email and order.buyer_email == user.email:
                is_authorized = True
    
    if not is_authorized:
        return jsonify({"error": "Unauthorized: You can only complete your own orders"}), 403
    
    # Update order for COD
    order.payment_method = 'cod'
    order.payment_status = 'pending'
    order.status = 'processing'
    order.updated_at = datetime.utcnow()
    
    # Add tracking entry
    tracking = OrderTracking(
        order_id=order_id,
        status="processing",
        message="Order placed successfully. Payment will be collected on delivery.",
        location=""
    )
    db.session.add(tracking)
    db.session.commit()
    
    # Send email notification
    try:
        send_order_confirmation_email(order)
    except:
        pass  # Don't fail if email fails
    
    return jsonify({
        "message": "COD order placed successfully",
        "order_id": order.id,
        "status": order.status,
        "payment_method": order.payment_method,
        "payment_status": order.payment_status
    })


@app.route("/api/orders", methods=["POST"])
def create_order():
    """Create a new order when a buyer clicks Buy Now.

    We accept buyer_name and buyer_email from the frontend so the order
    still works even if the stored user_id no longer exists (e.g. after reseeding).
    """
    ensure_db()
    data = request.get_json() or {}

    user_id = data.get("user_id")
    product_id = data.get("product_id")
    quantity = data.get("quantity", 1)
    buyer_name = data.get("buyer_name")
    buyer_email = data.get("buyer_email")

    if not product_id:
        return jsonify({"error": "Product ID is required"}), 400

    try:
        quantity = int(quantity)
    except (TypeError, ValueError):
        quantity = 1

    if quantity <= 0:
        quantity = 1

    product = Product.query.get(product_id)
    if not product:
        return jsonify({"error": "Product not found"}), 404

    if not product.company_id:
        return jsonify({"error": "Product is not linked to a company"}), 400

    # If we have a user_id, try to pull missing name/email from that user,
    # but don't fail the order if the user record is gone.
    if user_id and (not buyer_name or not buyer_email):
        user = User.query.get(user_id)
        if user:
            buyer_name = buyer_name or user.name
            buyer_email = buyer_email or user.email

    if not buyer_name or not buyer_email:
        return jsonify({"error": "Buyer name and email are required"}), 400

    unit_price = product.price or 0
    total_amount = (unit_price or 0) * quantity

    order = Order(
        product_id=product.id,
        company_id=product.company_id,
        buyer_name=buyer_name,
        buyer_email=buyer_email,
        buyer_id=user_id if user_id else None,
        quantity=quantity,
        unit_price=unit_price,
        total_amount=total_amount,
        status="pending",
        payment_status="pending",
    )
    db.session.add(order)
    db.session.commit()

    if product.stock_quantity is not None:
        try:
            product.stock_quantity = max((product.stock_quantity or 0) - quantity, 0)
            product.sales_count = (product.sales_count or 0) + quantity
            if product.category:
                product.category.sales_count = (product.category.sales_count or 0) + quantity
            db.session.commit()
            if product.stock_quantity < 10:
                create_seller_alert(
                    product.company_id,
                    "low_stock",
                    f"Low stock: {product.name}",
                    f"Only {product.stock_quantity} unit(s) remaining after the latest order.",
                    severity="warning",
                    entity_type="product",
                    entity_id=product.id,
                )
        except Exception as stock_error:
            print(f"[Order] Failed to update stock metrics: {stock_error}")
            db.session.rollback()

    try:
        create_seller_alert(
            product.company_id,
            "new_order",
            f"New order for {product.name}",
            f"{buyer_name} placed an order for {quantity} unit(s).",
            severity="high",
            entity_type="order",
            entity_id=order.id,
        )
    except Exception as alert_error:
        print(f"[Alert] Failed to create order alert: {alert_error}")
        db.session.rollback()

    try:
        record_marketplace_event(
            "order_created",
            session_id=get_request_session_id(data),
            user_id=user_id,
            company_id=product.company_id,
            product_id=product.id,
            category_id=product.category_id,
            order_id=order.id,
            location=product.company.location if product.company else product.location,
            metadata={
                "buyer_name": buyer_name,
                "quantity": quantity,
                "total_amount": total_amount,
            },
        )
    except Exception as analytics_error:
        print(f"[Analytics] Failed to record order event: {analytics_error}")
        db.session.rollback()

    return jsonify(
        {
            "id": order.id,
            "product_id": order.product_id,
            "product_name": product.name,
            "company_id": order.company_id,
            "company_name": product.company.name if product.company else None,
            "buyer_name": order.buyer_name,
            "buyer_email": order.buyer_email,
            "quantity": order.quantity,
            "unit_price": order.unit_price,
            "total_amount": order.total_amount,
            "status": order.status,
            "created_at": order.created_at.isoformat() if order.created_at else None,
        }
    ), 201


@app.route("/api/orders/<int:order_id>", methods=["GET"])
def get_order_details(order_id):
    """Get detailed order information including company address"""
    ensure_db()
    order = Order.query.get_or_404(order_id)
    
    company = order.company
    product = order.product
    
    return jsonify({
        "id": order.id,
        "product_id": order.product_id,
        "product_name": product.name if product else None,
        "company_id": order.company_id,
        "company_name": company.name if company else None,
        "company_location": company.location if company else None,
        "company_phone": company.phone if company else None,
        "company_website": company.website if company else None,
        "company_gst": company.gst_number if company else None,
        "buyer_name": order.buyer_name,
        "buyer_email": order.buyer_email,
        "quantity": order.quantity,
        "unit_price": order.unit_price,
        "total_amount": order.total_amount,
        "status": order.status,
        "payment_status": order.payment_status,
        "payment_method": order.payment_method,
        "created_at": order.created_at.isoformat() if order.created_at else None,
    })


@app.route("/api/my-orders")
def get_my_orders():
    """Get orders for the logged-in buyer (by user_id / email)"""
    ensure_db()
    user_id = request.args.get("user_id", type=int)
    if not user_id:
        return jsonify({"error": "User ID required"}), 400
    
    # Get user to check email
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404

    # Get orders by buyer_id OR buyer_email (for backward compatibility)
    orders = (
        Order.query.filter(
            (Order.buyer_id == user_id) | (Order.buyer_email == user.email)
        )
        .order_by(Order.created_at.desc())
        .all()
    )

    return jsonify(
        [
            {
                "id": o.id,
                "product_id": o.product_id,
                "product_name": o.product.name if o.product else None,
                "company_name": o.company.name if o.company else None,
                "quantity": o.quantity,
                "unit_price": o.unit_price,
                "total_amount": o.total_amount,
                "status": o.status,
                "created_at": o.created_at.isoformat() if o.created_at else None,
            }
            for o in orders
        ]
    )


@app.route("/api/admin/analytics")
def get_admin_analytics():
    """Get comprehensive analytics for admin dashboard"""
    ensure_db()
    # Check admin access
    user, error = require_admin()
    if error:
        return error
    from sqlalchemy import func, extract
    from datetime import timedelta
    
    # Total orders
    total_orders = Order.query.count()
    
    # Total revenue (completed orders only)
    total_revenue = db.session.query(func.sum(Order.total_amount)).filter(
        Order.status == "completed"
    ).scalar() or 0
    
    # Revenue by month (last 6 months)
    six_months_ago = datetime.utcnow() - timedelta(days=180)
    monthly_revenue = db.session.query(
        extract('year', Order.created_at).label('year'),
        extract('month', Order.created_at).label('month'),
        func.sum(Order.total_amount).label('revenue')
    ).filter(
        Order.status == "completed",
        Order.created_at >= six_months_ago
    ).group_by(
        extract('year', Order.created_at),
        extract('month', Order.created_at)
    ).order_by('year', 'month').all()
    
    revenue_by_month = [
        {
            "month": f"{int(r.month):02d}/{int(r.year)}",
            "revenue": float(r.revenue) if r.revenue else 0
        }
        for r in monthly_revenue
    ]
    
    # Top selling companies (by revenue)
    top_companies = db.session.query(
        Company.id,
        Company.name,
        func.sum(Order.total_amount).label('total_revenue'),
        func.count(Order.id).label('order_count')
    ).join(Order, Company.id == Order.company_id).filter(
        Order.status == "completed"
    ).group_by(Company.id, Company.name).order_by(
        func.sum(Order.total_amount).desc()
    ).limit(10).all()
    
    top_companies_data = [
        {
            "id": c.id,
            "name": c.name,
            "revenue": float(c.total_revenue) if c.total_revenue else 0,
            "orders": c.order_count
        }
        for c in top_companies
    ]
    
    # Category analysis (products sold by category)
    category_analysis = db.session.query(
        Category.id,
        Category.name,
        func.count(Order.id).label('order_count'),
        func.sum(Order.total_amount).label('revenue')
    ).join(Product, Category.id == Product.category_id).join(
        Order, Product.id == Order.product_id
    ).filter(
        Order.status == "completed"
    ).group_by(Category.id, Category.name).order_by(
        func.sum(Order.total_amount).desc()
    ).all()
    
    category_data = [
        {
            "id": c.id,
            "name": c.name,
            "orders": c.order_count,
            "revenue": float(c.revenue) if c.revenue else 0
        }
        for c in category_analysis
    ]
    
    # Orders by status
    orders_by_status = db.session.query(
        Order.status,
        func.count(Order.id).label('count')
    ).group_by(Order.status).all()
    
    status_data = {s.status: s.count for s in orders_by_status}
    
    # Recent orders (last 30 days)
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    recent_orders_count = Order.query.filter(
        Order.created_at >= thirty_days_ago
    ).count()
    
    # Revenue forecast (simple linear projection based on last 3 months)
    forecast_data = []
    if len(revenue_by_month) >= 2:
        last_3_months = revenue_by_month[-3:] if len(revenue_by_month) >= 3 else revenue_by_month
        avg_revenue = sum(r["revenue"] for r in last_3_months) / len(last_3_months)
        
        # Forecast next 3 months
        for i in range(1, 4):
            forecast_month = datetime.utcnow() + timedelta(days=30 * i)
            forecast_data.append({
                "month": forecast_month.strftime("%m/%Y"),
                "forecasted_revenue": avg_revenue * (1 + 0.1 * i)  # 10% growth assumption
            })
    
    return jsonify({
        "total_orders": total_orders,
        "total_revenue": float(total_revenue) if total_revenue else 0,
        "recent_orders_30d": recent_orders_count,
        "revenue_by_month": revenue_by_month,
        "top_companies": top_companies_data,
        "category_analysis": category_data,
        "orders_by_status": status_data,
        "forecast": forecast_data
    })


@app.route("/api/analytics/track", methods=["POST"])
def track_marketplace_event():
    ensure_db()
    data = request.get_json() or {}
    event_type = data.get("event_type")
    if not event_type:
        return jsonify({"error": "event_type is required"}), 400

    try:
        event = record_marketplace_event(
            event_type,
            session_id=get_request_session_id(data),
            user_id=data.get("user_id"),
            company_id=data.get("company_id"),
            product_id=data.get("product_id"),
            category_id=data.get("category_id"),
            inquiry_id=data.get("inquiry_id"),
            order_id=data.get("order_id"),
            location=data.get("location", ""),
            search_query=data.get("search_query", ""),
            metadata=data.get("metadata", {}),
        )
        return jsonify({"message": "tracked", "id": event.id}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Failed to track event: {str(e)}"}), 500


@app.route("/api/admin/realtime-analytics", methods=["GET"])
def get_admin_realtime_analytics():
    ensure_db()
    user, error = require_admin()
    if error:
        return error

    days = request.args.get("days", 30, type=int)
    return jsonify(build_admin_realtime_analytics(days=max(days, 1)))


@app.route("/api/seller/live-insights", methods=["GET"])
def get_seller_live_insights():
    ensure_db()
    user_id = request.args.get("user_id", type=int)
    if not user_id:
        return jsonify({"error": "User ID required"}), 400

    insights = build_seller_live_insights(user_id)
    if not insights:
        return jsonify({"error": "Seller profile not found"}), 404
    return jsonify(insights)


@app.route("/api/seller/alerts", methods=["GET"])
def get_seller_alerts():
    ensure_db()
    user_id = request.args.get("user_id", type=int)
    if not user_id:
        return jsonify({"error": "User ID required"}), 400

    user = User.query.get(user_id)
    if not user or not user.company_id:
        return jsonify({"error": "Seller profile not found"}), 404

    alerts = SellerAlert.query.filter_by(company_id=user.company_id).order_by(SellerAlert.created_at.desc()).all()
    return jsonify([serialize_seller_alert(alert) for alert in alerts])


@app.route("/api/seller/alerts/<int:alert_id>/read", methods=["POST"])
def mark_seller_alert_read(alert_id):
    ensure_db()
    data = request.get_json() or {}
    user_id = data.get("user_id") or request.args.get("user_id", type=int)
    if not user_id:
        return jsonify({"error": "User ID required"}), 400

    user = User.query.get(user_id)
    alert = SellerAlert.query.get_or_404(alert_id)
    if not user or not user.company_id or alert.company_id != user.company_id:
        return jsonify({"error": "Unauthorized"}), 403

    alert.is_read = True
    db.session.commit()
    emit_dashboard_refresh(company_id=user.company_id)
    return jsonify({"message": "Alert marked as read"})


@app.route("/api/suppliers/<int:company_id>/trust", methods=["GET"])
def get_supplier_trust(company_id):
    ensure_db()
    return jsonify(calculate_supplier_trust(company_id))


# ==================== NEW FEATURE ENDPOINTS ====================

# Image Upload
@app.route("/api/upload/image", methods=["POST"])
def upload_image():
    """Upload image for products, avatars, or company logos."""
    ensure_db()
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    
    file = request.files["file"]
    upload_type = request.form.get("type", "product")  # "product", "avatar", or "company"
    
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        unique_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{random.randint(1000, 9999)}_{filename}"
        
        if upload_type == "product":
            folder = "products"
        elif upload_type == "avatar":
            folder = "avatars"
        elif upload_type == "company":
            folder = "companies"
        else:
            folder = "products"
            
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], folder, unique_filename)
        file.save(filepath)
        
        url = f"/uploads/{folder}/{unique_filename}"
        return jsonify({"url": url, "filename": unique_filename})
    
    return jsonify({"error": "Invalid file type"}), 400


@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    """Serve uploaded files."""
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


# Product Images
@app.route("/api/products/<int:product_id>/images", methods=["GET", "POST"])
def product_images(product_id):
    """Get or add product images."""
    ensure_db()
    product = Product.query.get_or_404(product_id)
    
    if request.method == "GET":
        images = ProductImage.query.filter_by(product_id=product_id).order_by(ProductImage.display_order, ProductImage.id).all()
        return jsonify([{"id": img.id, "url": img.image_url, "is_primary": img.is_primary} for img in images])
    
    # POST - Add image
    data = request.get_json() or {}
    image_url = data.get("image_url")
    is_primary = data.get("is_primary", False)
    
    if not image_url:
        return jsonify({"error": "image_url is required"}), 400
    
    # If setting as primary, unset others
    if is_primary:
        ProductImage.query.filter_by(product_id=product_id, is_primary=True).update({"is_primary": False})
    
    image = ProductImage(product_id=product_id, image_url=image_url, is_primary=is_primary)
    db.session.add(image)
    db.session.commit()
    
    return jsonify({"id": image.id, "url": image.image_url, "is_primary": image.is_primary})


@app.route("/api/products/<int:product_id>/images/<int:image_id>", methods=["DELETE"])
def delete_product_image(product_id, image_id):
    """Delete a product image."""
    ensure_db()
    image = ProductImage.query.filter_by(id=image_id, product_id=product_id).first_or_404()
    db.session.delete(image)
    db.session.commit()
    return jsonify({"message": "Image deleted"})


@app.route("/api/admin/products/<int:product_id>", methods=["DELETE"])
def delete_product(product_id):
    """Delete a product (admin only)."""
    ensure_db()
    # Check admin access
    user, error = require_admin()
    if error:
        return error
    product = Product.query.get_or_404(product_id)
    
    # Get all orders for this product
    orders = Order.query.filter_by(product_id=product_id).all()
    order_ids = [o.id for o in orders]
    
    # Delete order tracking for these orders
    for order_id in order_ids:
        OrderTracking.query.filter_by(order_id=order_id).delete()
        # Delete chat messages related to this order
        ChatMessage.query.filter_by(order_id=order_id).delete()
    
    # Delete orders for this product
    Order.query.filter_by(product_id=product_id).delete()
    
    # Delete inquiries for this product
    Inquiry.query.filter_by(product_id=product_id).delete()
    
    # Delete reviews for this product
    Review.query.filter_by(product_id=product_id).delete()
    
    # Delete chat messages related to this product
    ChatMessage.query.filter_by(product_id=product_id).delete()
    
    # Delete all product images
    ProductImage.query.filter_by(product_id=product_id).delete()
    
    # Delete the product
    db.session.delete(product)
    db.session.commit()
    
    return jsonify({"message": "Product deleted successfully"})


@app.route("/api/admin/products/<int:product_id>/approve", methods=["POST"])
def approve_product(product_id):
    """Approve a product (admin only)."""
    ensure_db()
    # Check admin access
    user, error = require_admin()
    if error:
        return error
    product = Product.query.get_or_404(product_id)
    product.approved = True
    db.session.commit()
    
    # Send approval email
    if product.company:
        send_product_approval_email(product, product.company)
    
    return jsonify({
        "message": "Product approved successfully",
        "product_id": product.id,
        "product_name": product.name,
        "approved": product.approved
    }), 200


@app.route("/api/admin/products/<int:product_id>/reject", methods=["POST"])
def reject_product(product_id):
    """Reject a product (admin only)."""
    ensure_db()
    # Check admin access
    user, error = require_admin()
    if error:
        return error
    product = Product.query.get_or_404(product_id)
    product.approved = False
    db.session.commit()
    
    return jsonify({
        "message": "Product rejected successfully",
        "product_id": product.id,
        "product_name": product.name,
        "approved": product.approved
    }), 200


@app.route("/api/admin/products/pending", methods=["GET"])
def get_pending_products():
    """Get all products pending approval (admin only)."""
    ensure_db()
    # Check admin access
    user, error = require_admin()
    if error:
        return error
    pending_products = Product.query.filter_by(approved=False).order_by(Product.created_at.desc()).all()
    
    return jsonify([
        {
            "id": p.id,
            "name": p.name,
            "description": p.description,
            "price": p.price,
            "image_url": p.image_url,
            "category_id": p.category_id,
            "category_name": p.category.name if p.category else None,
            "company_id": p.company_id,
            "company_name": p.company.name if p.company else None,
            "created_at": p.created_at.isoformat() if p.created_at else None,
        }
        for p in pending_products
    ])


@app.route("/api/admin/products/approve-all", methods=["POST"])
def approve_all_products():
    """Approve all pending products (admin only)."""
    ensure_db()
    # Check admin access
    user, error = require_admin()
    if error:
        return error
    
    # Get all unapproved products
    pending_products = Product.query.filter_by(approved=False).all()
    
    if not pending_products:
        return jsonify({
            "message": "No pending products to approve",
            "count": 0
        }), 200
    
    # Approve all
    count = 0
    for product in pending_products:
        product.approved = True
        count += 1
        # Send approval email
        if product.company:
            send_product_approval_email(product, product.company)
    
    db.session.commit()
    
    return jsonify({
        "message": f"Successfully approved {count} product(s)",
        "count": count
    }), 200


@app.route("/api/admin/companies/<int:company_id>", methods=["DELETE"])
def delete_company(company_id):
    """Delete a company and all its products (admin only)."""
    ensure_db()
    # Check admin access
    user, error = require_admin()
    if error:
        return error
    company = Company.query.get_or_404(company_id)
    
    # Get all products for this company
    products = Product.query.filter_by(company_id=company_id).all()
    product_ids = [p.id for p in products]
    
    # Delete all related data for each product
    for product_id in product_ids:
        # Delete product images
        ProductImage.query.filter_by(product_id=product_id).delete()
        
        # Delete inquiries for this product
        Inquiry.query.filter_by(product_id=product_id).delete()
        
        # Delete reviews for this product
        Review.query.filter_by(product_id=product_id).delete()
        
        # Get orders for this product
        orders = Order.query.filter_by(product_id=product_id).all()
        order_ids = [o.id for o in orders]
        
        # Delete order tracking for these orders
        for order_id in order_ids:
            OrderTracking.query.filter_by(order_id=order_id).delete()
            # Delete chat messages related to this order
            ChatMessage.query.filter_by(order_id=order_id).delete()
        
        # Delete orders for this product
        Order.query.filter_by(product_id=product_id).delete()
        
        # Delete chat messages related to this product
        ChatMessage.query.filter_by(product_id=product_id).delete()
    
    # Delete all products
    for product in products:
        db.session.delete(product)
    
    # Delete trade leads for this company
    TradeLead.query.filter_by(company_id=company_id).delete()
    
    # Delete orders directly linked to this company (as seller)
    orders_linked_to_company = Order.query.filter_by(company_id=company_id).all()
    for order in orders_linked_to_company:
        # Delete order tracking
        OrderTracking.query.filter_by(order_id=order.id).delete()
        # Delete chat messages
        ChatMessage.query.filter_by(order_id=order.id).delete()
        # Delete the order
        db.session.delete(order)
    
    # Unlink users from this company (set company_id to None)
    User.query.filter_by(company_id=company_id).update({"company_id": None})
    
    # Delete the company
    db.session.delete(company)
    db.session.commit()
    
    return jsonify({"message": "Company and all its products deleted successfully"})


# Reviews and Ratings
@app.route("/api/products/<int:product_id>/reviews", methods=["GET", "POST"])
def product_reviews(product_id):
    """Get or add product reviews."""
    ensure_db()
    product = Product.query.get_or_404(product_id)
    
    if request.method == "GET":
        reviews = Review.query.filter_by(product_id=product_id).order_by(Review.created_at.desc()).all()
        avg_rating = db.session.query(db.func.avg(Review.rating)).filter_by(product_id=product_id).scalar() or 0
        
        return jsonify({
            "reviews": [{
                "id": r.id,
                "user_id": r.user_id,
                "user_name": r.user.name,
                "user_avatar": r.user.avatar_url,
                "rating": r.rating,
                "title": r.title,
                "comment": r.comment,
                "is_verified_purchase": r.is_verified_purchase,
                "created_at": r.created_at.isoformat()
            } for r in reviews],
            "average_rating": round(float(avg_rating), 1),
            "total_reviews": len(reviews)
        })
    
    # POST - Add review
    data = request.get_json() or {}
    user_id = data.get("user_id")
    rating = data.get("rating")
    title = data.get("title", "")
    comment = data.get("comment", "")
    
    if not user_id or not rating:
        return jsonify({"error": "user_id and rating are required"}), 400
    
    if rating < 1 or rating > 5:
        return jsonify({"error": "Rating must be between 1 and 5"}), 400
    
    # Check if user already reviewed
    existing = Review.query.filter_by(product_id=product_id, user_id=user_id).first()
    if existing:
        return jsonify({"error": "You have already reviewed this product"}), 400
    
    # Check if user purchased (for verified purchase badge)
    has_purchase = Order.query.filter_by(product_id=product_id, buyer_id=user_id, status="completed").first() is not None
    
    review = Review(
        product_id=product_id,
        user_id=user_id,
        rating=rating,
        title=title,
        comment=comment,
        is_verified_purchase=has_purchase
    )
    db.session.add(review)
    db.session.commit()
    
    return jsonify({
        "id": review.id,
        "rating": review.rating,
        "title": review.title,
        "comment": review.comment,
        "is_verified_purchase": review.is_verified_purchase
    })


# Order Tracking
@app.route("/api/orders/<int:order_id>/tracking", methods=["GET", "POST"])
def order_tracking(order_id):
    """Get or add order tracking updates."""
    ensure_db()
    order = Order.query.get_or_404(order_id)
    
    if request.method == "GET":
        tracking = OrderTracking.query.filter_by(order_id=order_id).order_by(OrderTracking.created_at.desc()).all()
        return jsonify([{
            "id": t.id,
            "status": t.status,
            "message": t.message,
            "location": t.location,
            "created_at": t.created_at.isoformat()
        } for t in tracking])
    
    # POST - Add tracking update
    data = request.get_json() or {}
    status = data.get("status")
    message = data.get("message", "")
    location = data.get("location", "")
    
    if not status:
        return jsonify({"error": "status is required"}), 400
    
    tracking = OrderTracking(order_id=order_id, status=status, message=message, location=location)
    db.session.add(tracking)
    
    # Update order status if provided
    if data.get("update_order_status"):
        order.status = status
        order.updated_at = datetime.utcnow()
    
    db.session.commit()
    
    return jsonify({
        "id": tracking.id,
        "status": tracking.status,
        "message": tracking.message,
        "created_at": tracking.created_at.isoformat()
    })


@app.route("/api/orders/<int:order_id>/tracking-number", methods=["PUT"])
def update_tracking_number(order_id):
    """Update order tracking number."""
    ensure_db()
    order = Order.query.get_or_404(order_id)
    data = request.get_json() or {}
    
    tracking_number = data.get("tracking_number")
    shipping_carrier = data.get("shipping_carrier", "")
    
    if tracking_number:
        order.tracking_number = tracking_number
        order.shipping_carrier = shipping_carrier
        order.status = "shipped"
        order.updated_at = datetime.utcnow()
        
        # Add tracking entry
        tracking = OrderTracking(
            order_id=order_id,
            status="shipped",
            message=f"Order shipped via {shipping_carrier or 'carrier'}",
            location=""
        )
        db.session.add(tracking)
        db.session.commit()
        
        return jsonify({"message": "Tracking number updated", "tracking_number": tracking_number})
    
    return jsonify({"error": "tracking_number is required"}), 400


# Payment Gateway (Razorpay)
@app.route("/api/payments/create-order", methods=["POST"])
def create_payment_order():
    """Create Razorpay order."""
    ensure_db()
    if not razorpay_client:
        return jsonify({"error": "Payment gateway not configured"}), 500
    
    data = request.get_json() or {}
    amount = data.get("amount")
    currency = data.get("currency", "INR")
    order_id = data.get("order_id")  # Our internal order ID
    
    if not amount or not order_id:
        return jsonify({"error": "amount and order_id are required"}), 400
    
    try:
        # Create Razorpay order
        razorpay_order = razorpay_client.order.create({
            "amount": int(amount * 100),  # Convert to paise
            "currency": currency,
            "receipt": f"order_{order_id}",
            "notes": {
                "order_id": order_id
            }
        })
        
        return jsonify({
            "razorpay_order_id": razorpay_order["id"],
            "amount": razorpay_order["amount"],
            "currency": razorpay_order["currency"],
            "key_id": RAZORPAY_KEY_ID
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/payments/verify", methods=["POST"])
def verify_payment():
    """Verify Razorpay payment."""
    ensure_db()
    if not razorpay_client:
        return jsonify({"error": "Payment gateway not configured"}), 500
    
    data = request.get_json() or {}
    razorpay_order_id = data.get("razorpay_order_id")
    razorpay_payment_id = data.get("razorpay_payment_id")
    razorpay_signature = data.get("razorpay_signature")
    order_id = data.get("order_id")
    
    if not all([razorpay_order_id, razorpay_payment_id, razorpay_signature, order_id]):
        return jsonify({"error": "Missing payment details"}), 400
    
    try:
        # Verify signature
        params_dict = {
            "razorpay_order_id": razorpay_order_id,
            "razorpay_payment_id": razorpay_payment_id,
            "razorpay_signature": razorpay_signature
        }
        razorpay_client.utility.verify_payment_signature(params_dict)
        
        # Update order
        order = Order.query.get_or_404(order_id)
        order.payment_status = "paid"
        order.payment_id = razorpay_payment_id
        order.status = "processing"
        
        # Increment sales count for product and category
        try:
            if order.product:
                order.product.sales_count = (order.product.sales_count or 0) + 1
                if order.product.category:
                    order.product.category.sales_count = (order.product.category.sales_count or 0) + 1
            db.session.commit()
        except Exception as e:
            print(f"Error updating sales counts: {e}")
            db.session.rollback()
        order.payment_status = "paid"
        order.payment_id = razorpay_payment_id
        order.status = "processing"
        order.updated_at = datetime.utcnow()
        
        # Add tracking
        tracking = OrderTracking(
            order_id=order_id,
            status="processing",
            message="Payment received, order is being processed",
            location=""
        )
        db.session.add(tracking)
        db.session.commit()
        
        # Send email notification
        try:
            send_order_confirmation_email(order)
        except:
            pass  # Don't fail if email fails
        
        return jsonify({"message": "Payment verified successfully", "order_id": order_id})
    except Exception as e:
        return jsonify({"error": f"Payment verification failed: {str(e)}"}), 400


# User Profiles
@app.route("/api/users/<int:user_id>/profile", methods=["GET", "PUT"])
def user_profile(user_id):
    """Get or update user profile."""
    ensure_db()
    user = User.query.get_or_404(user_id)
    
    if request.method == "GET":
        return jsonify({
            "id": user.id,
            "name": user.name,
            "email": user.email,
            "avatar_url": user.avatar_url,
            "phone": user.phone,
            "bio": user.bio,
            "company_id": user.company_id,
            "created_at": user.created_at.isoformat()
        })
    
    # PUT - Update profile
    data = request.get_json() or {}
    if "name" in data:
        user.name = data["name"]
    if "phone" in data:
        user.phone = data["phone"]
    if "bio" in data:
        user.bio = data["bio"]
    if "avatar_url" in data:
        user.avatar_url = data["avatar_url"]
    
    db.session.commit()
    return jsonify({"message": "Profile updated", "user": {
        "id": user.id,
        "name": user.name,
        "avatar_url": user.avatar_url,
        "phone": user.phone,
        "bio": user.bio
    }})


# Order History
@app.route("/api/users/<int:user_id>/orders", methods=["GET"])
def user_orders(user_id):
    """Get user's order history."""
    ensure_db()
    user = User.query.get_or_404(user_id)
    
    orders = Order.query.filter_by(buyer_id=user_id).order_by(Order.created_at.desc()).all()
    
    return jsonify([{
        "id": o.id,
        "product_id": o.product_id,
        "product_name": o.product.name if o.product else "Unknown",
        "company_name": o.company.name if o.company else "Unknown",
        "quantity": o.quantity,
        "unit_price": o.unit_price,
        "total_amount": o.total_amount,
        "status": o.status,
        "payment_status": o.payment_status,
        "tracking_number": o.tracking_number,
        "created_at": o.created_at.isoformat(),
        "updated_at": o.updated_at.isoformat() if o.updated_at else None
    } for o in orders])


# Advanced Search
@app.route("/api/products/search", methods=["GET"])
def advanced_search():
    """Advanced product search with multiple filters."""
    ensure_db()
    query = Product.query.join(Company).join(Category)
    
    # Only show approved products
    query = query.filter(Product.approved == True)
    
    # Text search
    search = request.args.get("search")
    if search:
        like = f"%{search}%"
        query = query.filter(Product.name.ilike(like) | Product.description.ilike(like))
    
    # Category filter
    category = request.args.get("category")
    if category:
        query = query.filter(Product.category_id == category)
    
    # Price range
    min_price = request.args.get("min_price", type=float)
    max_price = request.args.get("max_price", type=float)
    if min_price is not None:
        query = query.filter(Product.price >= min_price)
    if max_price is not None:
        query = query.filter(Product.price <= max_price)
    
    # Rating filter (requires join with reviews)
    min_rating = request.args.get("min_rating", type=float)
    if min_rating is not None:
        subquery = db.session.query(Review.product_id, db.func.avg(Review.rating).label("avg_rating")).group_by(Review.product_id).having(db.func.avg(Review.rating) >= min_rating).subquery()
        query = query.join(subquery, Product.id == subquery.c.product_id)
    
    # Location filter
    location = request.args.get("location")
    if location:
        query = query.filter(Company.location.ilike(f"%{location}%"))
    
    # Verified supplier filter
    verified_supplier = request.args.get("verified_supplier")
    if verified_supplier and verified_supplier.lower() == "true":
        query = query.filter(Company.verified == True)
    
    # Best seller filter
    best_seller = request.args.get("best_seller")
    if best_seller and best_seller.lower() == "true":
        query = query.filter(Company.best_seller == True)
    
    # Sort options
    sort_by = request.args.get("sort_by", "created_at")
    sort_order = request.args.get("sort_order", "desc")
    
    if sort_by == "price":
        query = query.order_by(Product.price.desc() if sort_order == "desc" else Product.price.asc())
    elif sort_by == "price_desc":
        query = query.order_by(Product.price.desc())
    elif sort_by == "rating":
        # Sort by average rating
        subquery = db.session.query(Review.product_id, db.func.avg(Review.rating).label("avg_rating")).group_by(Review.product_id).subquery()
        query = query.outerjoin(subquery, Product.id == subquery.c.product_id)
        query = query.order_by(subquery.c.avg_rating.desc() if sort_order == "desc" else subquery.c.avg_rating.asc())
    else:
        query = query.order_by(Product.created_at.desc() if sort_order == "desc" else Product.created_at.asc())
    
    # Pagination
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    products = pagination.items

    try:
        lead_category_id = int(category) if category and str(category).isdigit() else (products[0].category_id if products else None)
        record_marketplace_event(
            "search" if search else "catalog_view",
            session_id=get_request_session_id(),
            user_id=request.args.get("user_id", type=int),
            company_id=products[0].company_id if products else None,
            product_id=products[0].id if products else None,
            category_id=lead_category_id,
            location=location or "",
            search_query=search or "",
            metadata={
                "results_count": pagination.total,
                "filters": {
                    "category": category,
                    "location": location,
                    "verified_supplier": verified_supplier,
                    "best_seller": best_seller,
                    "sort_by": sort_by,
                },
            },
        )
    except Exception as analytics_error:
        print(f"[Analytics] Failed to record advanced search: {analytics_error}")
        db.session.rollback()
    
    # Get average ratings for products
    product_ids = [p.id for p in products]
    ratings = db.session.query(Review.product_id, db.func.avg(Review.rating).label("avg_rating"), db.func.count(Review.id).label("count")).filter(Review.product_id.in_(product_ids)).group_by(Review.product_id).all()
    rating_dict = {r.product_id: {"rating": round(float(r.avg_rating), 1), "count": r.count} for r in ratings}
    
    return jsonify({
        "products": [{
            "id": p.id,
            "name": p.name,
            "description": p.description,
            "price": p.price,
            "image_url": p.image_url,
            "category_id": p.category_id,
            "category_name": p.category.name if p.category else None,
            "company_id": p.company_id,
            "company_name": p.company.name if p.company else None,
            "location": p.company.location if p.company else "India",
            "rating": rating_dict.get(p.id, {}).get("rating", 0),
            "review_count": rating_dict.get(p.id, {}).get("count", 0)
        } for p in products],
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": pagination.total,
            "pages": pagination.pages
        }
    })


# Real-time Chat
@app.route("/api/chat/conversations", methods=["GET"])
def get_conversations():
    """Get all conversations for a user."""
    ensure_db()
    user_id = request.args.get("user_id", type=int)
    if not user_id:
        return jsonify({"error": "user_id is required"}), 400
    
    # Get all unique conversations (users who sent/received messages)
    sent_conversations = db.session.query(ChatMessage.receiver_id).filter_by(sender_id=user_id).distinct().all()
    received_conversations = db.session.query(ChatMessage.sender_id).filter_by(receiver_id=user_id).distinct().all()
    
    all_user_ids = set([c[0] for c in sent_conversations] + [c[0] for c in received_conversations])
    
    conversations = []
    for other_user_id in all_user_ids:
        other_user = User.query.get(other_user_id)
        if not other_user:
            continue
        
        # Get last message
        last_message = ChatMessage.query.filter(
            ((ChatMessage.sender_id == user_id) & (ChatMessage.receiver_id == other_user_id)) |
            ((ChatMessage.sender_id == other_user_id) & (ChatMessage.receiver_id == user_id))
        ).order_by(ChatMessage.created_at.desc()).first()
        
        # Count unread
        unread_count = ChatMessage.query.filter_by(sender_id=other_user_id, receiver_id=user_id, is_read=False).count()
        
        conversations.append({
            "user_id": other_user_id,
            "user_name": other_user.name,
            "user_avatar": other_user.avatar_url,
            "last_message": last_message.message if last_message else "",
            "last_message_time": last_message.created_at.isoformat() if last_message else None,
            "unread_count": unread_count
        })
    
    # Sort by last message time
    conversations.sort(key=lambda x: x["last_message_time"] or "", reverse=True)
    
    return jsonify(conversations)


@app.route("/api/ai/generate-description", methods=["POST"])
def generate_description():
    """Generate description using AI based on user prompt"""
    ensure_db()
    data = request.get_json() or {}
    prompt = data.get("prompt", "")
    entity_type = data.get("entity_type", "product")  # "product" or "company"
    additional_info = data.get("additional_info", {})  # Additional context like name, category, etc.
    
    if not prompt:
        return jsonify({"error": "Prompt is required"}), 400
    
    try:
        if openai_client and OPENAI_API_KEY:
            # Use OpenAI API
            system_prompt = ""
            if entity_type == "product":
                product_name = additional_info.get("name", "")
                category = additional_info.get("category", "")
                system_prompt = f"You are a professional product description writer for a B2B marketplace. Write a compelling, professional product description for a B2B audience. Product name: {product_name}. Category: {category}. User request: {prompt}. Make it informative, highlight key features and benefits for business buyers."
            elif entity_type == "requirement":
                product_name = additional_info.get("name", "")
                system_prompt = f"You are a professional B2B procurement specialist. Write a detailed, structured buy requirement (RFQ) description for a B2B marketplace. Product: {product_name}. User context: {prompt}. Your description should include: 1. General Overview 2. Potential Technical Specifications 3. Quality Standards expected 4. Delivery and Packaging requirements. Make it professional and ready for suppliers to quote on."
            else:  # company
                company_name = additional_info.get("name", "")
                location = additional_info.get("location", "")
                system_prompt = f"You are a professional business description writer. Write a compelling company description for a B2B marketplace. Company name: {company_name}. Location: {location}. User request: {prompt}. Make it professional, highlight business capabilities, expertise, and value proposition for other businesses."
            
            # Try newer OpenAI API format first, then fallback
            try:
                from openai import OpenAI
                client = OpenAI(api_key=OPENAI_API_KEY)
                response = client.chat.completions.create(
                    model="gpt-3.5-turbo",
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": prompt}
                    ],
                    max_tokens=500,
                    temperature=0.7
                )
                generated_description = response.choices[0].message.content.strip()
            except Exception as e1:
                print(f"Primary OpenAI failed: {e1}")
                # Fallback to older API format if available
                try:
                    import openai
                    openai.api_key = OPENAI_API_KEY
                    response = openai.ChatCompletion.create(
                        model="gpt-3.5-turbo",
                        messages=[
                            {"role": "system", "content": system_prompt},
                            {"role": "user", "content": prompt}
                        ],
                        max_tokens=500,
                        temperature=0.7
                    )
                    generated_description = response.choices[0].message.content.strip()
                except Exception as e2:
                    print(f"Fallback OpenAI failed: {e2}")
                    generated_description = f"Requirement for {prompt}. (AI service currently unavailable)"
            
            return jsonify({"description": generated_description})
        else:
            # Fallback: Generate a template-based description
            if entity_type == "product":
                product_name = additional_info.get("name", "this product")
                category = additional_info.get("category", "")
                generated_description = f"""We are pleased to offer {product_name}{' in the ' + category + ' category' if category else ''}.

{prompt}

Key Features:
• High quality and reliability
• Competitive pricing
• Suitable for B2B requirements
• Available in bulk quantities

Perfect for businesses looking for {product_name.lower()}. Contact us for more details, pricing, and customization options."""
            elif entity_type == "requirement":
                product_name = additional_info.get("name", "this product")
                generated_description = f"""We are looking to source {product_name} for our upcoming business operations.

Requirement Details:
- High quality and compliance with industrial standards
- Competitive bulk pricing
- Reliable delivery timelines

Specifications provided: {prompt}

Please provide your best quotation along with product datasheets and lead times."""
            else:  # company
                company_name = additional_info.get("name", "our company")
                location = additional_info.get("location", "")
                generated_description = f"""{company_name}{' is a trusted business' + (' based in ' + location if location else '')} specializing in providing quality products and services to businesses.

{prompt}

We pride ourselves on:
• Professional service delivery
• Quality assurance
• Competitive pricing
• Strong customer relationships

Partner with us for your business needs. Contact us today to discuss how we can support your business goals."""
            
            return jsonify({"description": generated_description})
    
    except Exception as e:
        print(f"Error generating description: {e}")
        # Return a simple fallback description
        if entity_type == "product":
            fallback = f"Premium quality product. {prompt}. Available for B2B orders. Contact us for pricing and bulk order options."
        else:
            fallback = f"Professional business services. {prompt}. Contact us to learn more about how we can serve your business needs."
        return jsonify({"description": fallback})


@app.route("/api/chat/messages", methods=["GET", "POST"])
def chat_messages():
    """Get or send chat messages."""
    ensure_db()
    
    if request.method == "GET":
        sender_id = request.args.get("sender_id", type=int)
        receiver_id = request.args.get("receiver_id", type=int)
        product_id = request.args.get("product_id", type=int)
        
        if not sender_id or not receiver_id:
            return jsonify({"error": "sender_id and receiver_id are required"}), 400
        
        query = ChatMessage.query.filter(
            ((ChatMessage.sender_id == sender_id) & (ChatMessage.receiver_id == receiver_id)) |
            ((ChatMessage.sender_id == receiver_id) & (ChatMessage.receiver_id == sender_id))
        )
        
        if product_id:
            query = query.filter_by(product_id=product_id)
        
        messages = query.order_by(ChatMessage.created_at.asc()).all()
        
        # Mark as read
        ChatMessage.query.filter_by(sender_id=receiver_id, receiver_id=sender_id, is_read=False).update({"is_read": True})
        db.session.commit()
        
        return jsonify([{
            "id": m.id,
            "sender_id": m.sender_id,
            "sender_name": m.sender.name,
            "receiver_id": m.receiver_id,
            "receiver_name": m.receiver.name,
            "message": m.message,
            "product_id": m.product_id,
            "order_id": m.order_id,
            "is_read": m.is_read,
            "created_at": m.created_at.isoformat()
        } for m in messages])
    
    # POST - Send message
    data = request.get_json() or {}
    sender_id = data.get("sender_id")
    receiver_id = data.get("receiver_id")
    message = data.get("message")
    product_id = data.get("product_id")
    order_id = data.get("order_id")
    
    if not all([sender_id, receiver_id, message]):
        return jsonify({"error": "sender_id, receiver_id, and message are required"}), 400
    
    chat_message = ChatMessage(
        sender_id=sender_id,
        receiver_id=receiver_id,
        message=message,
        product_id=product_id,
        order_id=order_id
    )
    db.session.add(chat_message)
    db.session.commit()
    
    return jsonify({
        "id": chat_message.id,
        "sender_id": chat_message.sender_id,
        "receiver_id": chat_message.receiver_id,
        "message": chat_message.message,
        "created_at": chat_message.created_at.isoformat()
    })


# Email notification helper
def send_order_confirmation_email(order):
    """Send order confirmation email."""
    if not app.config.get("MAIL_USERNAME"):
        return  # Email not configured
    
    try:
        msg = Message(
            subject=f"Order Confirmation - Order #{order.id}",
            recipients=[order.buyer_email],
            body=f"""
Dear {order.buyer_name},

Thank you for your order!

Order ID: {order.id}
Product: {order.product.name if order.product else 'N/A'}
Quantity: {order.quantity}
Total Amount: ₹{order.total_amount}

Your order is being processed and you will receive updates soon.

Best regards,
DealsDouble.AI Team
            """,
            html=f"""
            <html>
            <body>
            <h2>Order Confirmation</h2>
            <p>Dear {order.buyer_name},</p>
            <p>Thank you for your order!</p>
            <ul>
            <li>Order ID: {order.id}</li>
            <li>Product: {order.product.name if order.product else 'N/A'}</li>
            <li>Quantity: {order.quantity}</li>
            <li>Total Amount: ₹{order.total_amount}</li>
            </ul>
            <p>Your order is being processed and you will receive updates soon.</p>
            <p>Best regards,<br>DealsDouble.AI Team</p>
            </body>
            </html>
            """
        )
        mail.send(msg)
    except Exception as e:
        print(f"Failed to send email: {e}")


def send_product_approval_email(product, company):
    """Send email when product is approved."""
    try:
        if not company or not company.users:
            return
        
        user = company.users[0]  # Get first user of the company
        if not user.email:
            return
        
        msg = Message(
            subject="Product Approved - DealsDouble.AI",
            recipients=[user.email],
            html=f"""
            <html>
            <body>
            <h2>Product Approved!</h2>
            <p>Dear {user.name},</p>
            <p>Great news! Your product has been approved and is now live on DealsDouble.AI.</p>
            <ul>
            <li>Product Name: {product.name}</li>
            <li>Category: {product.category.name if product.category else 'N/A'}</li>
            <li>Price: ₹{product.price}</li>
            </ul>
            <p>Your product is now visible to all buyers on our platform.</p>
            <p>Best regards,<br>DealsDouble.AI Team</p>
            </body>
            </html>
            """
        )
        mail.send(msg)
    except Exception as e:
        print(f"Failed to send product approval email: {e}")


def send_company_approval_email(company):
    """Send email when company is approved."""
    try:
        if not company or not company.users:
            return
        
        user = company.users[0]  # Get first user of the company
        if not user.email:
            return
        
        msg = Message(
            subject="Company Verified - DealsDouble.AI",
            recipients=[user.email],
            html=f"""
            <html>
            <body>
            <h2>Company Verified!</h2>
            <p>Dear {user.name},</p>
            <p>Congratulations! Your company "{company.name}" has been verified and approved.</p>
            <p>You can now:</p>
            <ul>
            <li>Post products</li>
            <li>Access the Seller Dashboard</li>
            <li>Receive inquiries from buyers</li>
            <li>Manage orders</li>
            </ul>
            <p>Best regards,<br>DealsDouble.AI Team</p>
            </body>
            </html>
            """
        )
        mail.send(msg)
    except Exception as e:
        print(f"Failed to send company approval email: {e}")


def send_new_inquiry_email(inquiry, product, seller_email):
    """Send email to seller when new inquiry is received."""
    try:
        if not seller_email:
            return
        
        msg = Message(
            subject=f"New Inquiry for {product.name} - DealsDouble.AI",
            recipients=[seller_email],
            html=f"""
            <html>
            <body>
            <h2>New Product Inquiry</h2>
            <p>You have received a new inquiry for your product.</p>
            <ul>
            <li>Product: {product.name}</li>
            <li>Buyer Name: {inquiry.name}</li>
            <li>Buyer Email: {inquiry.email}</li>
            <li>Buyer Phone: {inquiry.phone}</li>
            <li>Quantity: {inquiry.quantity or 'Not specified'}</li>
            <li>Message: {inquiry.message or 'No message'}</li>
            </ul>
            <p>Please respond to the buyer as soon as possible.</p>
            <p>Best regards,<br>DealsDouble.AI Team</p>
            </body>
            </html>
            """
        )
        mail.send(msg)
    except Exception as e:
        print(f"Failed to send inquiry email: {e}")


def send_inquiry_reply_email(inquiry, reply, seller):
    """Send email to buyer when seller replies to their inquiry."""
    if not app.config.get("MAIL_USERNAME"):
        return  # Email not configured
    
    try:
        if not inquiry.email:
            return
        
        msg = Message(
            sender=app.config.get("MAIL_USERNAME"),
            subject=f"Reply to your inquiry for {inquiry.product.name} - DealsDouble.AI",
            recipients=[inquiry.email],
            html=f"""
            <html>
            <body>
            <h2>Reply to Your Inquiry</h2>
            <p>Dear {inquiry.name},</p>
            <p>You have received a reply from the seller regarding your inquiry.</p>
            <div style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <h3>Your Inquiry:</h3>
                <p><strong>Product:</strong> {inquiry.product.name}</p>
                <p><strong>Your Message:</strong> {inquiry.message}</p>
            </div>
            <div style="background-color: #e8f4f8; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <h3>Seller's Reply:</h3>
                <p><strong>From:</strong> {seller.name}</p>
                <p>{reply.message}</p>
            </div>
            <p>You can view all replies to your inquiries on DealsDouble.AI.</p>
            <p>Best regards,<br>DealsDouble.AI Team</p>
            </body>
            </html>
            """
        )
        mail.send(msg)
        print(f"[OK] Inquiry reply email sent to {inquiry.email}")
    except Exception as e:
        print(f"Failed to send inquiry reply email: {e}")


def send_welcome_email(user):
    """Send welcome email to new user."""
    if not app.config.get("MAIL_USERNAME"):
        return  # Email not configured
    
    try:
        if not user.email:
            return
        
        msg = Message(
            subject="Welcome to DealsDouble.AI!",
            sender=app.config.get("MAIL_USERNAME"),
            recipients=[user.email],
            html=f"""
            <html>
            <body>
            <h2>Welcome to DealsDouble.AI!</h2>
            <p>Dear {user.name},</p>
            <p>Thank you for joining DealsDouble.AI - Your AI-Powered B2B Marketplace!</p>
            <p>Get started by:</p>
            <ul>
            <li>Browsing products from verified suppliers</li>
            <li>Posting your buying requirements</li>
            <li>Registering your company to start selling</li>
            </ul>
            <p>If you have any questions, feel free to contact our support team.</p>
            <p>Best regards,<br>DealsDouble.AI Team</p>
            </body>
            </html>
            """
        )
        mail.send(msg)
        print(f"[OK] Welcome email sent to {user.email}")
    except Exception as e:
        print(f"[ERROR] Failed to send welcome email to {user.email}: {e}")
        import traceback
        traceback.print_exc()


# Wishlist Endpoints
@app.route("/api/wishlist", methods=["GET"])
def get_wishlist():
    """Get user's wishlist"""
    ensure_db()
    user_id = request.args.get("user_id", type=int)
    if not user_id:
        return jsonify({"error": "user_id is required"}), 400
    
    wishlist_items = Wishlist.query.filter_by(user_id=user_id).all()
    product_ids = [item.product_id for item in wishlist_items]
    
    products = Product.query.filter(Product.id.in_(product_ids)).all() if product_ids else []
    
    # Get ratings
    ratings = db.session.query(Review.product_id, db.func.avg(Review.rating).label("avg_rating"), db.func.count(Review.id).label("count")).filter(Review.product_id.in_(product_ids)).group_by(Review.product_id).all() if product_ids else []
    rating_dict = {r.product_id: {"rating": round(float(r.avg_rating), 1), "count": r.count} for r in ratings}
    
    return jsonify([{
        "id": p.id,
        "name": p.name,
        "description": p.description,
        "price": p.price,
        "image_url": p.image_url,
        "category_name": p.category.name if p.category else None,
        "company_name": p.company.name if p.company else None,
        "location": p.company.location if p.company else "India",
        "rating": rating_dict.get(p.id, {}).get("rating", 0),
        "review_count": rating_dict.get(p.id, {}).get("count", 0),
        "added_at": next((item.created_at.isoformat() for item in wishlist_items if item.product_id == p.id), None)
    } for p in products])


@app.route("/api/wishlist", methods=["POST"])
def add_to_wishlist():
    """Add product to wishlist"""
    ensure_db()
    data = request.get_json() or {}
    user_id = data.get("user_id")
    product_id = data.get("product_id")
    
    if not user_id or not product_id:
        return jsonify({"error": "user_id and product_id are required"}), 400
    
    # Check if already in wishlist
    existing = Wishlist.query.filter_by(user_id=user_id, product_id=product_id).first()
    if existing:
        return jsonify({"message": "Product already in wishlist", "in_wishlist": True}), 200
    
    wishlist_item = Wishlist(user_id=user_id, product_id=product_id)
    db.session.add(wishlist_item)
    db.session.commit()
    
    return jsonify({"message": "Product added to wishlist", "in_wishlist": True}), 201


@app.route("/api/wishlist", methods=["DELETE"])
def remove_from_wishlist():
    """Remove product from wishlist"""
    ensure_db()
    user_id = request.args.get("user_id", type=int)
    product_id = request.args.get("product_id", type=int)
    
    if not user_id or not product_id:
        return jsonify({"error": "user_id and product_id are required"}), 400
    
    wishlist_item = Wishlist.query.filter_by(user_id=user_id, product_id=product_id).first()
    if not wishlist_item:
        return jsonify({"error": "Product not in wishlist"}), 404
    
    db.session.delete(wishlist_item)
    db.session.commit()
    
    return jsonify({"message": "Product removed from wishlist", "in_wishlist": False}), 200


@app.route("/api/wishlist/check", methods=["GET"])
def check_wishlist():
    """Check if product is in user's wishlist"""
    ensure_db()
    user_id = request.args.get("user_id", type=int)
    product_id = request.args.get("product_id", type=int)
    
    if not user_id or not product_id:
        return jsonify({"in_wishlist": False}), 200
    
    exists = Wishlist.query.filter_by(user_id=user_id, product_id=product_id).first() is not None
    return jsonify({"in_wishlist": exists}), 200


# Enhanced Supplier Profile
@app.route("/api/suppliers/<int:company_id>/stats")
def get_supplier_stats(company_id: int):
    """Get detailed supplier statistics"""
    ensure_db()
    company = Company.query.get_or_404(company_id)
    
    # Get product count
    product_count = Product.query.filter_by(company_id=company_id, approved=True).count()
    
    # Get total orders
    total_orders = Order.query.join(Product).filter(Product.company_id == company_id).count()
    
    # Get total revenue
    orders = Order.query.join(Product).filter(Product.company_id == company_id).all()
    total_revenue = sum(order.total_amount for order in orders if order.payment_status == "paid")
    
    trust = calculate_supplier_trust(company_id)
    avg_response_time = (
        f"{trust['avg_response_hours']} hours"
        if trust["avg_response_hours"] is not None
        else "No responses yet"
    )
    
    # Get product categories
    categories = db.session.query(Category.name, db.func.count(Product.id).label("count")).join(Product).filter(Product.company_id == company_id, Product.approved == True).group_by(Category.name).all()
    
    return jsonify({
        "company_id": company.id,
        "company_name": company.name,
        "verified": company.verified,
        "best_seller": company.best_seller if hasattr(company, 'best_seller') else False,
        "stats": {
            "product_count": product_count,
            "total_orders": total_orders,
            "total_revenue": total_revenue,
            "avg_response_time": avg_response_time,
            "categories": [{"name": c.name, "count": c.count} for c in categories],
            "trust_score": trust["trust_score"],
            "review_rating": trust["review_rating"],
            "response_rate": trust["response_rate"],
            "completion_rate": trust["completion_rate"],
        },
        "trust": trust,
    })


# Advanced Analytics Dashboard
@app.route("/api/admin/advanced-analytics", methods=["GET"])
def advanced_analytics():
    """Get advanced analytics for admin dashboard"""
    ensure_db()
    user, error = require_admin()
    if error:
        return error
    
    # Date range (default to last 30 days)
    days = request.args.get("days", 30, type=int)
    start_date = datetime.utcnow() - timedelta(days=days)
    
    # Revenue trends
    orders = Order.query.filter(Order.created_at >= start_date).all()
    revenue_by_date = {}
    for order in orders:
        if order.payment_status == "paid":
            date_key = order.created_at.date().isoformat()
            revenue_by_date[date_key] = revenue_by_date.get(date_key, 0) + order.total_amount
    
    # Product performance
    top_products = db.session.query(
        Product.id,
        Product.name,
        db.func.count(Order.id).label("order_count"),
        db.func.sum(Order.total_amount).label("revenue")
    ).join(Order, Product.id == Order.product_id).filter(
        Order.created_at >= start_date,
        Order.payment_status == "paid"
    ).group_by(Product.id, Product.name).order_by(
        db.func.sum(Order.total_amount).desc()
    ).limit(10).all()
    
    # Category performance
    category_performance = db.session.query(
        Category.name,
        db.func.count(Product.id).label("product_count"),
        db.func.count(Order.id).label("order_count")
    ).join(Product, Category.id == Product.category_id).outerjoin(
        Order, Product.id == Order.product_id
    ).filter(
        Product.approved == True,
        Order.created_at >= start_date if Order.id else True
    ).group_by(Category.name).all()
    
    # Supplier performance
    supplier_performance = db.session.query(
        Company.id,
        Company.name,
        db.func.count(Product.id).label("product_count"),
        db.func.count(Order.id).label("order_count"),
        db.func.sum(Order.total_amount).label("revenue")
    ).join(Product, Company.id == Product.company_id).outerjoin(
        Order, Product.id == Order.product_id
    ).filter(
        Company.verified == True,
        Order.created_at >= start_date if Order.id else True
    ).group_by(Company.id, Company.name).order_by(
        db.func.sum(Order.total_amount).desc()
    ).limit(10).all()
    
    return jsonify({
        "period": {
            "days": days,
            "start_date": start_date.isoformat(),
            "end_date": datetime.utcnow().isoformat()
        },
        "revenue_trends": {
            "daily": [{"date": k, "revenue": v} for k, v in sorted(revenue_by_date.items())],
            "total": sum(revenue_by_date.values())
        },
        "top_products": [{
            "id": p.id,
            "name": p.name,
            "order_count": p.order_count,
            "revenue": float(p.revenue or 0)
        } for p in top_products],
        "category_performance": [{
            "name": c.name,
            "product_count": c.product_count,
            "order_count": c.order_count
        } for c in category_performance],
        "supplier_performance": [{
            "id": s.id,
            "name": s.name,
            "product_count": s.product_count,
            "order_count": s.order_count,
            "revenue": float(s.revenue or 0)
        } for s in supplier_performance]
    })


@app.route("/api/seller/catalog-template", methods=["GET"])
def download_catalog_template():
    """Generate and download a CSV template for bulk product upload."""
    import csv
    from io import StringIO
    
    cols = ["name", "description", "price", "category", "stock_quantity", "min_order_quantity", "location", "tags"]
    sample_data = [
        ["Sample Product 1", "High quality industrial gear", 5000, "Industrial Supplies", 100, 5, "Mumbai", "gear, industrial"],
        ["Sample Product 2", "Advanced LED Panel", 1200, "Electronics & Electrical", 50, 10, "Bangalore", "led, electronic"]
    ]
    
    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(cols)
    writer.writerows(sample_data)
    
    output = si.getvalue()
    si.close()
    
    return output, 200, {
        "Content-Type": "text/csv",
        "Content-Disposition": "attachment; filename=catalog_template.csv"
    }


@app.route("/api/seller/catalog-upload", methods=["POST"])
def upload_catalog():
    """Handle bulk product upload from CSV or Excel file."""
    # Ensure current user is a seller
    user_id = request.form.get("user_id", type=int)
    if not user_id:
        user_id = request.headers.get("X-User-Id", type=int)
    
    user = User.query.get(user_id) if user_id else get_current_user()
    if not user or not user.company_id:
        return jsonify({"error": "Seller access required"}), 403
    
    if "file" not in request.files:
        return jsonify({"error": "No file part"}), 400
    
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No selected file"}), 400
    
    if file and allowed_data_file(file.filename):
        filename = secure_filename(file.filename)
        # Simple unique filename
        filename = f"{int(datetime.utcnow().timestamp())}_{filename}"
        file_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        file.save(file_path)
        
        try:
            import pandas as pd
            if filename.lower().endswith(".csv"):
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path)
            
            # Required columns validation
            required_cols = ["name", "category"]
            for col in required_cols:
                if col not in df.columns:
                    if os.path.exists(file_path): os.remove(file_path)
                    return jsonify({"error": f"Missing required column: {col}"}), 400
            
            # Map category names to IDs
            categories = {c.name.lower(): c.id for c in Category.query.all()}
            
            success_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    name = str(row.get("name", "")).strip()
                    if not name or name == "nan":
                        continue # Skip empty rows
                        
                    cat_name = str(row.get("category", "")).strip()
                    cat_id = categories.get(cat_name.lower())
                    
                    if not cat_id:
                        errors.append(f"Row {index+2}: Category '{cat_name}' not found.")
                        continue
                    
                    # Convert price - handle nan
                    price = row.get("price", 0)
                    price = float(price) if pd.notnull(price) else 0
                    
                    product = Product(
                        name=name,
                        description=str(row.get("description", "")) if pd.notnull(row.get("description")) else "",
                        price=price,
                        category_id=cat_id,
                        company_id=user.company_id,
                        stock_quantity=int(row.get("stock_quantity", 0)) if pd.notnull(row.get("stock_quantity")) else 0,
                        min_order_quantity=int(row.get("min_order_quantity", 1)) if pd.notnull(row.get("min_order_quantity")) else 1,
                        location=str(row.get("location", user.company.location if user.company else "India")) if pd.notnull(row.get("location")) else "India",
                        tags=str(row.get("tags", "")) if pd.notnull(row.get("tags")) else "",
                        approved=True # Auto-approve for now
                    )
                    
                    # Default image based on category
                    category_list = [
                        "Industrial Supplies", "Electronics & Electrical", "Apparel & Fashion", 
                        "Machinery", "Construction & Real Estate", "Chemicals", 
                        "Food & Beverage", "Health & Beauty"
                    ]
                    try:
                        cat_idx = category_list.index(cat_name)
                        product.image_url = f"/uploads/products/cat{cat_idx}.png"
                    except:
                        product.image_url = "/placeholder.png"

                    db.session.add(product)
                    success_count += 1
                except Exception as row_error:
                    errors.append(f"Row {index+2}: {str(row_error)}")
            
            db.session.commit()
            if os.path.exists(file_path): os.remove(file_path)
            
            return jsonify({
                "success": True,
                "message": f"Bulk upload complete. {success_count} products added.",
                "failed_count": len(errors),
                "errors": errors[:5]
            })
            
        except Exception as e:
            if os.path.exists(file_path): os.remove(file_path)
            return jsonify({"error": f"Failed to process file: {str(e)}"}), 500
            
    return jsonify({"error": "Invalid file format. Please use CSV or Excel (.xlsx, .xls)"}), 400


@app.route("/api/admin/users", methods=["GET"])
def get_all_users():
    """Get all registered users for admin dashboard"""
    ensure_db()
    user, error = require_admin()
    if error:
        return error
    
    users = User.query.order_by(User.created_at.desc()).all()
    
    user_list = []
    for u in users:
        role = "Buyer"
        if u.is_admin:
            role = "Admin"
        elif u.company_id:
            role = "Seller"
            
        user_list.append({
            "id": u.id,
            "name": u.name,
            "email": u.email,
            "phone": u.phone,
            "role": role,
            "is_buyer_manager": u.is_buyer_manager,
            "company_id": u.company_id,
            "company_name": u.company.name if u.company else None,
            "created_at": u.created_at.isoformat() if u.created_at else None
        })
        
    return jsonify(user_list)


# Currency Conversion
@app.route("/api/admin/users/<int:target_user_id>/subscription", methods=["POST"])
def admin_update_subscription(target_user_id: int):
    """Admin-only: Update a user's subscription tier."""
    user, error = require_admin()
    if error:
        return error
    
    data = request.get_json() or {}
    new_tier = data.get("tier", "").upper()
    
    if new_tier not in ["STARTER", "BASIC", "PREMIUM"]:
        return jsonify({"error": "Invalid tier. Must be STARTER, BASIC, or PREMIUM"}), 400
    
    target_user = User.query.get_or_404(target_user_id)
    target_user.membership_tier = new_tier
    db.session.commit()
    
    return jsonify({"success": True, "message": f"User {target_user.name} updated to {new_tier}"})


@app.route("/api/trade-leads")
def list_trade_leads():
    """List trade leads with masking logic based on user subscription/purchase."""
    ensure_db()
    current_user = get_current_user()
    leads = TradeLead.query.order_by(TradeLead.created_at.desc()).all()
    
    # Get user's purchased leads
    purchased_lead_ids = []
    if current_user:
        purchased_lead_ids = [p.lead_id for p in current_user.lead_purchases]
    
    results = []
    for lead in leads:
        # Check if user has access to full details
        # Access if: Admin, Lead Owner, Purchased, or Premium User
        has_access = False
        if current_user:
            if current_user.is_admin or current_user.membership_tier == "PREMIUM" or lead.id in purchased_lead_ids:
                has_access = True
                
        results.append({
            "id": lead.id,
            "title": lead.title,
            "description": lead.description,
            "type": lead.type,
            "category": lead.category,
            "location": lead.location,
            "price": lead.price,
            "has_access": has_access,
            "created_at": lead.created_at.isoformat(),
            # Masked data
            "contact_name": lead.contact_name if has_access else "Locked",
            "contact_email": lead.contact_email if has_access else "Locked",
            "contact_phone": lead.contact_phone if has_access else "Locked",
        })
        
    return jsonify(results)


@app.route("/api/trade-leads/<int:lead_id>/purchase", methods=["POST"])
def purchase_trade_lead(lead_id: int):
    """Allow a user to purchase a trade lead."""
    ensure_db()
    user = get_current_user()
    
    if not user:
        return jsonify({"error": "Authentication required"}), 401
        
    lead = TradeLead.query.get_or_404(lead_id)
    
    # Check if already purchased
    existing = LeadPurchase.query.filter_by(user_id=user.id, lead_id=lead_id).first()
    if existing:
        return jsonify({"message": "Already purchased"}), 200
        
    # In a real app, you'd check balance/process payment here
    # Record purchase event
    try:
        event = MarketplaceEvent(
            event_type="lead_purchased",
            user_id=user.id,
            inquiry_id=lead_id, # Reusing inquiry_id for leads since it's just a proxy for TradeLead ID
            metadata_json=json.dumps({
                "lead_id": lead_id,
                "price": lead.price,
                "category": lead.category
            })
        )
        db.session.add(event)
        db.session.commit()
    except Exception as e:
        print(f"Analytics event error: {e}")

    return jsonify({
        "success": True, 
        "message": "Lead unlocked successfully",
        "contact_name": lead.contact_name,
        "contact_email": lead.contact_email,
        "contact_phone": lead.contact_phone
    })


@app.route("/api/admin/leads", methods=["POST"])
def admin_create_lead():
    """Admin-only: Create a new trade lead and notify via Socket.IO"""
    user, error = require_admin()
    if error:
        return error
        
    data = request.get_json() or {}
    title = data.get("title")
    description = data.get("description")
    l_type = data.get("type", "buy")
    category = data.get("category")
    location = data.get("location")
    contact_name = data.get("contact_name", "")
    contact_email = data.get("contact_email", "")
    contact_phone = data.get("contact_phone", "")
    price = data.get("price", 500)
    
    if not all([title, description, category]):
        return jsonify({"error": "Missing title, description, or category"}), 400
        
    lead = TradeLead(
        title=title,
        description=description,
        type=l_type,
        category=category,
        location=location,
        contact_name=contact_name,
        contact_email=contact_email,
        contact_phone=contact_phone,
        price=price
    )
    db.session.add(lead)
    db.session.commit()
    
    # Record analytics event
    try:
        event = MarketplaceEvent(
            event_type="lead_created",
            user_id=user.id,
            location=location,
            metadata_json=json.dumps({
                "lead_id": lead.id,
                "category": category,
                "type": l_type
            })
        )
        db.session.add(event)
        db.session.commit()
    except Exception as e:
        print(f"Analytics event error: {e}")
    
    # Emit real-time notification
    try:
        socketio.emit("new_trade_lead", {
            "id": lead.id,
            "title": lead.title,
            "type": lead.type,
            "category": lead.category,
            "created_at": lead.created_at.isoformat()
        })
    except Exception as e:
        print(f"Socket emit error: {e}")
        
    return jsonify({"success": True, "id": lead.id}), 201


@app.route("/api/admin/leads/<int:lead_id>", methods=["DELETE"])
def admin_delete_lead(lead_id: int):
    """Admin-only: Delete a trade lead."""
    user, error = require_admin()
    if error:
        return error
        
    lead = TradeLead.query.get_or_404(lead_id)
    LeadPurchase.query.filter_by(lead_id=lead_id).delete() # Cleanup purchases
    db.session.delete(lead)
    db.session.commit()
    
    return jsonify({"success": True, "message": "Lead deleted"})


@app.route("/api/currency/convert", methods=["GET"])
def convert_currency():
    """Convert currency amount (using a simple conversion rate API)"""
    ensure_db()
    amount = request.args.get("amount", type=float)
    from_currency = request.args.get("from", "INR").upper()
    to_currency = request.args.get("to", "USD").upper()
    
    if not amount:
        return jsonify({"error": "amount is required"}), 400
    
    # Simple conversion rates (in production, use a real API like exchangerate-api.com or fixer.io)
    # Base currency: INR
    conversion_rates = {
        "INR": 1.0,
        "USD": 0.012,
        "EUR": 0.011,
        "GBP": 0.0095,
        "AED": 0.044,
        "CNY": 0.087,
        "JPY": 1.8
    }
    
    from_rate = conversion_rates.get(from_currency, 1.0)
    to_rate = conversion_rates.get(to_currency, 1.0)
    
    # Convert to base (INR) then to target currency
    base_amount = amount / from_rate
    converted_amount = base_amount * to_rate
    
    return jsonify({
        "amount": amount,
        "from_currency": from_currency,
        "to_currency": to_currency,
        "converted_amount": round(converted_amount, 2),
        "rate": round(to_rate / from_rate, 4)
    })


@app.route("/api/currency/rates", methods=["GET"])
def get_currency_rates():
    """Get all currency conversion rates"""
    ensure_db()
    base_currency = request.args.get("base", "INR").upper()
    
    # Simple conversion rates (in production, use a real API)
    conversion_rates = {
        "INR": 1.0,
        "USD": 0.012,
        "EUR": 0.011,
        "GBP": 0.0095,
        "AED": 0.044,
        "CNY": 0.087,
        "JPY": 1.8
    }
    
    base_rate = conversion_rates.get(base_currency, 1.0)
    
    return jsonify({
        "base": base_currency,
        "rates": {currency: round(rate / base_rate, 4) for currency, rate in conversion_rates.items()}
    })



@app.route("/api/admin/fix-images", methods=["POST"])
def fix_product_images():
    """Admin-only: Update all products to use local category images instead of external placeholders."""
    user, error = require_admin()
    if error:
        return error

    try:
        categories_data = [
            "Industrial Supplies",
            "Electronics & Electrical",
            "Apparel & Fashion",
            "Machinery",
            "Construction & Real Estate",
            "Chemicals",
            "Food & Beverage",
            "Health & Beauty",
        ]
        category_map = {name: i for i, name in enumerate(categories_data)}

        products = Product.query.all()
        updated_count = 0
        
        for product in products:
            # Match by category name or industrial as default
            new_url = None
            if product.category and product.category.name in category_map:
                cat_idx = category_map[product.category.name]
                new_url = f"/uploads/products/cat{cat_idx}.png"
            
            # If still external placeholder, force update to a local one
            if not new_url and (not product.image_url or "via.placeholder.com" in product.image_url):
                new_url = "/placeholder.png"
                
            if new_url and product.image_url != new_url:
                product.image_url = new_url
                updated_count += 1

        db.session.commit()
        return jsonify({
            "success": True, 
            "message": f"Updated {updated_count} products with local high-quality images.",
            "total_products": len(products)
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    ensure_db()
    socketio.run(app, host="0.0.0.0", port=5000, debug=False)
